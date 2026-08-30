"""
tests/test_core.py — tests unitaires des modules logique pure.

Couverture :
  - naming       : validate_meta, canonical_session_dirname,
                   vigiechiro_wav_prefix, compute_new_wav_name
  - taxons       : classify_taxon (tous les groupes + fallback prefix)
  - cleanup      : decide_contact (toutes les branches)
  - manifest     : save/load idempotence, flags, Action(status)
  - registry     : upsert, thread-safety, batch commit, migration
  - repair       : coverage pure, suggestions, dry-run/apply, garde-fous
  - export_sessions : plan/run, options Data/Data_k, dry-run, structure

Ces modules sont les plus "à risque" : une régression y corrompt des
données (renommage incohérent, cleanup trop aggressif, etc.). Tests
légers, rapides (< 2s au total), sans réseau, sans GUI.

Run :  pytest -q tests/
"""

from __future__ import annotations

import sqlite3
import tempfile
from datetime import datetime
from pathlib import Path

import pytest


# =========================================================================
# naming
# =========================================================================

class TestNaming:
    def _make_meta(self, **kwargs):
        from naming import SessionMeta
        defaults = dict(
            date_debut=datetime(2025, 9, 3),
            n_site_tadarida="212097",
            n_point_fixe="Z3",
            n_passage=2,
            n_enregistreur=7,
            n_serie="SMU03126",
            nom_contrat="TestContrat",
        )
        defaults.update(kwargs)
        return SessionMeta(**defaults)

    def test_validate_meta_ok(self):
        from naming import validate_meta
        errs = validate_meta(self._make_meta())
        assert errs == [], f"attendait [] mais reçu {errs}"

    def test_validate_meta_missing_date(self):
        from naming import validate_meta
        errs = validate_meta(self._make_meta(date_debut=None))
        assert any("date_debut" in e for e in errs)

    def test_validate_meta_missing_site(self):
        from naming import validate_meta
        errs = validate_meta(self._make_meta(n_site_tadarida=None))
        assert any("n_site_tadarida" in e for e in errs)

    def test_validate_meta_bad_point(self):
        from naming import validate_meta
        # Point "99" sans lettre → invalide
        errs = validate_meta(self._make_meta(n_point_fixe="99"))
        assert len(errs) > 0

    def test_canonical_session_dirname(self):
        from naming import canonical_session_dirname
        name = canonical_session_dirname(self._make_meta())
        assert name == "20250903_site212097_Z3_Pass2_enr07"

    def test_canonical_session_dirname_uppercase_point(self):
        from naming import canonical_session_dirname
        name = canonical_session_dirname(self._make_meta(n_point_fixe="z3"))
        # Normalisé en majuscules
        assert "_Z3_" in name

    def test_canonical_session_dirname_invalid_raises(self):
        from naming import canonical_session_dirname
        with pytest.raises(ValueError):
            canonical_session_dirname(self._make_meta(n_site_tadarida=None))

    def test_canonical_session_enr_padding(self):
        """n_enregistreur doit être zéro-paddé à 2 chiffres."""
        from naming import canonical_session_dirname
        name = canonical_session_dirname(self._make_meta(n_enregistreur=3))
        assert name.endswith("_enr03")

    def test_vigiechiro_wav_prefix(self):
        from naming import vigiechiro_wav_prefix
        pref = vigiechiro_wav_prefix(self._make_meta())
        assert pref == "Car212097-2025-Pass2-Z3-SMU03126"

    def test_compute_new_wav_name_with_timestamp(self):
        """Fichier brut SM4BAT → nom canonique Vigie-Chiro."""
        from naming import compute_new_wav_name
        meta = self._make_meta()
        new = compute_new_wav_name(meta, "S4U04784_20250903_210523.wav")
        # Le nouveau nom doit commencer par le préfixe canonique + timestamp
        assert new is not None
        assert "Car212097-2025-Pass2-Z3-SMU03126" in new
        assert "_20250903_210523" in new

    def test_compute_new_wav_name_already_canonical(self):
        """Fichier déjà au nom canonique → retourne le même nom (skip silencieux)."""
        from naming import compute_new_wav_name
        meta = self._make_meta()
        original = "Car212097-2025-Pass2-Z3-SMU03126_20250903_210523_000.wav"
        new = compute_new_wav_name(meta, original)
        assert new == original

    def test_compute_new_wav_name_unreadable(self):
        """Fichier sans timestamp parseable → None."""
        from naming import compute_new_wav_name
        meta = self._make_meta()
        assert compute_new_wav_name(meta, "garbage_no_date.wav") is None


class TestRememberWorkspace:
    def test_disabled_keeps_session_path_skips_recents(self, tmp_path):
        from gui_config import Settings, remember_workspace
        s = Settings(remember_last_workspace=False)
        remember_workspace(str(tmp_path), s)
        assert Path(s.last_workspace) == tmp_path.resolve()
        assert s.recent_workspaces == []

    def test_enabled_updates_recents(self, tmp_path):
        from gui_config import Settings, remember_workspace
        s = Settings(remember_last_workspace=True)
        remember_workspace(str(tmp_path), s)
        assert s.last_workspace is not None
        assert s.recent_workspaces[0] == s.last_workspace


# =========================================================================
# taxons
# =========================================================================

class TestTaxons:
    def test_classify_noise(self):
        from taxons import classify_taxon
        assert classify_taxon("noise") == "noise"
        assert classify_taxon("Noise") == "noise"  # case-insensitive
        assert classify_taxon("NOISE") == "noise"

    def test_classify_chiro_codes(self):
        from taxons import classify_taxon
        # Codes taxons chiros classiques
        # (on teste les préfixes de fallback car la table MNHN peut être absente en CI)
        assert classify_taxon("Pippip") in ("chiros",)  # Pipistrellus pipistrellus
        assert classify_taxon("Nycnoc") in ("chiros",)  # Nyctalus noctula
        assert classify_taxon("Myodau") in ("chiros",)  # Myotis daubentonii

    def test_classify_ortho(self):
        from taxons import classify_taxon
        # Orthoptères (préfixe T...)
        assert classify_taxon("Tetvir") in ("orthos",)  # Tettigonia viridissima
        assert classify_taxon("Yposp") in ("orthos",)

    def test_classify_unknown(self):
        from taxons import classify_taxon
        assert classify_taxon("ZZZnon_existant") == "unknown"
        assert classify_taxon(None) == "unknown"
        assert classify_taxon("") == "unknown"

    def test_classify_empty_and_none(self):
        from taxons import classify_taxon
        assert classify_taxon(None) == "unknown"


# =========================================================================
# cleanup.decide_contact
# =========================================================================

class TestCleanupDecision:
    def _thresholds(self):
        return {"chiros": 0.5, "orthos": 0.5, "micromam": 0.5, "oiseaux": 0.5}

    def test_noise_always_deleted(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("noise", 0.9,
                                    self._thresholds(), disabled=set())
        assert g == "noise"
        assert d == "deleted_noise"

    def test_chiro_above_threshold_kept(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("Pippip", 0.8,
                                    self._thresholds(), disabled=set())
        assert g == "chiros"
        assert d == "kept"

    def test_chiro_below_threshold_deleted(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("Pippip", 0.3,
                                    self._thresholds(), disabled=set())
        assert g == "chiros"
        assert d == "deleted_low_confidence"

    def test_chiro_exactly_at_threshold_kept(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("Pippip", 0.5,
                                    self._thresholds(), disabled=set())
        # >= seuil → conservé
        assert d == "kept"

    def test_group_disabled(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("Pippip", 0.9,
                                    self._thresholds(), disabled={"chiros"})
        assert d == "deleted_group_disabled"

    def test_proba_none_kept_by_default(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("Pippip", None,
                                    self._thresholds(), disabled=set())
        assert d == "kept"

    def test_unknown_taxon_keep_default(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("ZZZ_inconnu", 0.9,
                                    self._thresholds(), disabled=set(),
                                    unknown_action="keep")
        assert g == "unknown"
        assert d == "kept"

    def test_unknown_taxon_delete_optional(self):
        from cleanup import decide_contact
        g, d, _r = decide_contact("ZZZ_inconnu", 0.9,
                                    self._thresholds(), disabled=set(),
                                    unknown_action="delete")
        assert d == "deleted_unknown_group"


# =========================================================================
# manifest
# =========================================================================

class TestManifest:
    def test_save_load_roundtrip(self, tmp_path):
        from manifest import Manifest
        session = tmp_path / "session1"
        session.mkdir()
        m = Manifest.load_or_create(session)
        m.canonical_name = "20250903_site212097_Z3_Pass2_enr07"
        m.set_meta(n_site_tadarida="212097", n_point_fixe="Z3")
        m.save(session)
        assert (session / "_session_manifest.json").is_file()

        # Reload
        m2 = Manifest.load(session)
        assert m2 is not None
        assert m2.canonical_name == "20250903_site212097_Z3_Pass2_enr07"
        assert m2.meta.get("n_site_tadarida") == "212097"

    def test_idempotence_flag(self, tmp_path):
        """Un flag mis à True reste True après save/load."""
        from manifest import Manifest
        session = tmp_path / "session2"
        session.mkdir()
        m = Manifest.load_or_create(session)
        m.flags["renamed"] = True
        m.save(session)

        m2 = Manifest.load(session)
        assert m2.flags.get("renamed") is True
        assert m2.is_done("rename")

    def test_action_status_failure_not_done(self, tmp_path):
        """Une action avec status != 'ok' ne doit pas être considérée done."""
        from manifest import Manifest
        session = tmp_path / "session3"
        session.mkdir()
        m = Manifest.load_or_create(session)
        # Action échouée : flag ne doit pas passer à True
        m.record_action("rename", status="error", tool_version="test")
        m.save(session)
        # L'implémentation gère ça via flags (qui ne passe à True que sur "ok")
        assert m.flags.get("renamed", False) is False

    def test_partial_upload_not_done(self, tmp_path):
        """Régression reprise d'upload : une action 'partial' NE DOIT PAS marquer
        l'étape comme faite (sinon le 2e clic court-circuite l'upload et les WAV
        manquants ne repartent jamais)."""
        from manifest import Manifest
        session = tmp_path / "sess_partial"
        session.mkdir()
        m = Manifest.load_or_create(session)
        m.record_action("upload", status="partial", tool_version="test")
        assert m.flags.get("uploaded", False) is False
        assert m.is_done("upload") is False            # le bug corrigé
        # une fois l'upload complet réussi → l'étape est faite
        m.record_action("upload", status="ok", tool_version="test")
        assert m.is_done("upload") is True

    def test_is_done_unknown_type_needs_ok(self, tmp_path):
        """Type sans flag dédié : seule une action 'ok' compte comme faite."""
        from manifest import Manifest
        session = tmp_path / "sess_unknown"
        session.mkdir()
        m = Manifest.load_or_create(session)
        m.record_action("customstep", status="partial", tool_version="test")
        assert m.is_done("customstep") is False
        m.record_action("customstep", status="ok", tool_version="test")
        assert m.is_done("customstep") is True


# =========================================================================
# export_sessions — plan + run (portable Data/Data_k + meta)
# =========================================================================

class TestExportSessions:
    def _make_session(self, root: Path, name: str, *, with_data=True,
                      with_data_k=True, sibling_data_k=False):
        """Crée une mini-session sous root/<campagne>/<name>/."""
        camp = root / "MonContrat"
        session = camp / name
        session.mkdir(parents=True)
        (session / "_session_manifest.json").write_text(
            '{"schema_version": 1, "flags": {"uploaded": true}}', encoding="utf-8")
        (session / "participation-abc-observations.xlsx").write_bytes(b"PK\x03\x04xlsx")
        (session / "participation-abc-observations.sync.json").write_text(
            '{"entries":[]}', encoding="utf-8")
        (session / "_stats_before_cleanup.json").write_text("{}", encoding="utf-8")
        (session / "SMU_Summary.txt").write_text(
            "DATE,TIME,LAT,LON,TEMP\n", encoding="utf-8")
        if with_data:
            data = session / "Data"
            data.mkdir()
            (data / "raw1.wav").write_bytes(b"R" * 100)
            (data / "raw2.wav").write_bytes(b"R" * 200)
        if with_data_k and not sibling_data_k:
            dk = session / "Data_k"
            dk.mkdir()
            (dk / "seg_000.wav").write_bytes(b"K" * 50)
            (dk / "seg_001.wav").write_bytes(b"K" * 60)
        if sibling_data_k:
            sdk = camp / "Data_k" / name
            sdk.mkdir(parents=True)
            (sdk / "sib_000.wav").write_bytes(b"S" * 40)
        chi = session / "ChiroSurf_nuits" / "nuit1"
        chi.mkdir(parents=True)
        (chi / "note.txt").write_text("surf", encoding="utf-8")
        return session

    def test_plan_meta_and_data_k_default(self, tmp_path):
        from export_sessions import ExportSessionSpec, plan_export
        session = self._make_session(tmp_path, "20250903_site212097_Z3_Pass2_enr07")
        dest = tmp_path / "usb"
        dest.mkdir()
        plan = plan_export(
            [ExportSessionSpec(session_path=session)],  # data_k=True, data=False
            dest, stamp="20260101_120000",
        )
        kinds = {f.kind for f in plan.files}
        assert "meta" in kinds
        assert "data_k" in kinds
        assert "data" not in kinds
        assert "chirosurf" in kinds
        rels = {f.rel_dst for f in plan.files}
        assert any(r.endswith("_session_manifest.json") for r in rels)
        assert any("/Data_k/" in r or r.endswith("Data_k/seg_000.wav")
                   or "Data_k/seg_000.wav" in r for r in rels)
        assert not any("/Data/" in r.replace("\\", "/") and "raw" in r for r in rels)
        assert plan.estimated_bytes > 0
        assert plan.dest_root.endswith("ChiroTool_export_20260101_120000")

    def test_plan_include_data(self, tmp_path):
        from export_sessions import ExportSessionSpec, plan_export
        session = self._make_session(tmp_path, "sess_data")
        plan = plan_export(
            [ExportSessionSpec(session_path=session, include_data=True,
                               include_data_k=False)],
            tmp_path / "out", stamp="t1",
        )
        kinds = {f.kind for f in plan.files}
        assert "data" in kinds
        assert "data_k" not in kinds
        assert any("Data/raw1.wav" in f.rel_dst.replace("\\", "/") for f in plan.files)

    def test_plan_sibling_data_k_normalized(self, tmp_path):
        """Source campagne/Data_k/<session>/ → export .../<session>/Data_k/."""
        from export_sessions import ExportSessionSpec, plan_export
        session = self._make_session(
            tmp_path, "sess_sib", with_data_k=False, sibling_data_k=True,
        )
        plan = plan_export(
            [ExportSessionSpec(session_path=session, include_data_k=True)],
            tmp_path / "out", stamp="t2",
        )
        rels = [f.rel_dst.replace("\\", "/") for f in plan.files if f.kind == "data_k"]
        assert rels
        assert all("/Data_k/" in r for r in rels)
        assert any(r.endswith("sib_000.wav") for r in rels)

    def test_dry_run_writes_nothing(self, tmp_path):
        from export_sessions import ExportSessionSpec, plan_export, run_export
        session = self._make_session(tmp_path, "sess_dry")
        dest = tmp_path / "usb"
        dest.mkdir()
        plan = plan_export(
            [ExportSessionSpec(session_path=session)], dest, stamp="dry1",
        )
        root = Path(plan.dest_root)
        res = run_export(plan, dry_run=True)
        assert res["dry_run"] is True
        assert res["n_copied"] == 0
        assert not root.exists()  # rien créé en dry-run

    def test_run_export_structure_and_manifest(self, tmp_path):
        from export_sessions import ExportSessionSpec, plan_export, run_export
        import json
        session = self._make_session(tmp_path, "sess_run")
        dest = tmp_path / "usb"
        dest.mkdir()
        plan = plan_export(
            [ExportSessionSpec(session_path=session, include_data=False,
                               include_data_k=True)],
            dest, stamp="run1",
        )
        progress_calls = []

        def prog(done, total, label):
            progress_calls.append((done, total, label))

        res = run_export(plan, dry_run=False, progress=prog)
        assert res["n_errors"] == 0
        assert res["n_copied"] > 0
        root = Path(plan.dest_root)
        assert (root / "EXPORT_README.txt").is_file()
        assert (root / "export_manifest.json").is_file()
        man = json.loads((root / "export_manifest.json").read_text(encoding="utf-8"))
        assert man["schema"] == 1
        assert man["totals"]["n_copied"] == res["n_copied"]
        # Arborescence relative campagne/session
        sess_dir = root / "MonContrat" / "sess_run"
        assert (sess_dir / "_session_manifest.json").is_file()
        assert (sess_dir / "Data_k" / "seg_000.wav").is_file()
        assert not (sess_dir / "Data").exists()
        assert (sess_dir / "ChiroSurf_nuits" / "nuit1" / "note.txt").is_file()
        assert progress_calls  # progress appelé

    def test_collision_session_names(self, tmp_path):
        from export_sessions import ExportSessionSpec, plan_export
        # Deux sessions même nom sous campagnes différentes → rel uniques
        s1 = self._make_session(tmp_path / "a", "SameName")
        # force même nom sous autre parent
        camp2 = tmp_path / "b" / "AutreContrat"
        s2 = camp2 / "SameName"
        s2.mkdir(parents=True)
        (s2 / "_session_manifest.json").write_text("{}", encoding="utf-8")
        plan = plan_export(
            [
                ExportSessionSpec(session_path=s1, campaign="C1"),
                ExportSessionSpec(session_path=s2, campaign="C1"),  # même campagne forcée
            ],
            tmp_path / "out", stamp="col",
        )
        rels = {s["rel_path"] for s in plan.sessions}
        assert len(rels) == 2  # SameName et SameName_2

    def test_wizard_group_helpers(self):
        """Helpers du wizard (groupement contrats) sans ouvrir de GUI."""
        from gui_export_wizard import _group_sessions, _session_label, _fmt_bytes
        groups = _group_sessions([
            {"nom_contrat": "Alpha", "date_debut": "2025-09-01",
             "n_point_fixe": "Z1", "n_passage": 1, "canonical_name": "n1"},
            {"campaign": "BetaCamp", "date_debut": "2025-08-01",
             "n_point_fixe": "Z2", "n_passage": 2, "id": "n2"},
        ])
        assert "Alpha" in groups
        assert "BetaCamp" in groups
        lbl = _session_label(groups["Alpha"][0])
        assert "Z1" in lbl and "Pass1" in lbl
        assert "MB" in _fmt_bytes(2_000_000) or "KB" in _fmt_bytes(2_000_000)


class TestFinishUploadWithTrigger:
    """Branche all_already_present / trigger : toujours record_action, jamais
    de flag posé en silence si le compute échoue."""

    class _ClientOK:
        def trigger_compute(self, pid):
            return {"ok": True}

    class _ClientFail:
        def trigger_compute(self, pid):
            raise RuntimeError("HTTP 503 compute down")

    def test_trigger_ok_sets_uploaded_via_record_action(self, tmp_path):
        from manifest import Manifest
        from pipeline import _finish_upload_with_trigger

        session = tmp_path / "sess_ok"
        session.mkdir()
        m = Manifest.load_or_create(session)
        out: dict = {"phase": "upload", "steps": []}

        result = _finish_upload_with_trigger(
            self._ClientOK(), m, session, out,
            part_id="pid1",
            stats={"n_wavs": 3, "uploaded_now": 0, "already_present": 3,
                   "failed": 0, "all_already_present": True},
            notes_prefix="3 déjà présents (reprise)",
        )

        assert "error" not in result
        assert any(s.get("step") == "trigger_compute" and not s.get("error")
                   for s in result["steps"])
        m2 = Manifest.load(session)
        assert m2 is not None
        assert m2.flags.get("uploaded") is True
        assert m2.is_done("upload") is True
        last = m2.last_action("upload")
        assert last is not None
        assert last.status == "ok"
        assert last.stats.get("trigger_ok") is True
        assert last.stats.get("all_already_present") is True
        # Pas de mutation manuelle hors record_action : une seule action upload
        assert sum(1 for a in m2.actions if a.type == "upload") == 1

    def test_trigger_fail_no_uploaded_flag_error_surfaced(self, tmp_path):
        from manifest import Manifest
        from pipeline import _finish_upload_with_trigger

        session = tmp_path / "sess_fail"
        session.mkdir()
        m = Manifest.load_or_create(session)
        out: dict = {"phase": "upload", "steps": []}

        result = _finish_upload_with_trigger(
            self._ClientFail(), m, session, out,
            part_id="pid2",
            stats={"n_wavs": 5, "uploaded_now": 0, "already_present": 5,
                   "failed": 0, "all_already_present": True},
            notes_prefix="5 déjà présents (reprise)",
        )

        assert result.get("error")
        assert "trigger" in result["error"].lower() or "Tadarida" in result["error"]
        assert any(s.get("step") == "trigger_compute" and s.get("error")
                   for s in result["steps"])
        m2 = Manifest.load(session)
        assert m2 is not None
        assert m2.flags.get("uploaded") is False
        assert m2.is_done("upload") is False   # relance / repair possibles
        last = m2.last_action("upload")
        assert last is not None
        assert last.status == "error"
        assert last.stats.get("trigger_ok") is False
        assert "503" in (last.notes or "")


# =========================================================================
# taxons + cleanup intégration (règle OR WAV)
# =========================================================================

class TestCleanupFileDecision:
    """La règle OR : un WAV est gardé dès qu'UN contact est kept."""

    def test_or_rule_keeps_file_with_one_chiro(self):
        """Un WAV avec 1 chiro kept + 10 noise → gardé."""
        # Simule les contacts d'un même WAV
        from cleanup import decide_contact
        thresholds = {"chiros": 0.5, "orthos": 0.5,
                       "micromam": 0.5, "oiseaux": 0.5}
        # Contact 1 : chiro valide
        _, d1, _ = decide_contact("Pippip", 0.9, thresholds, set())
        # Contacts 2..11 : noise
        decisions = [d1] + ["deleted_noise"] * 10
        # Règle OR : au moins un kept → fichier gardé
        any_kept = any(d == "kept" for d in decisions)
        assert any_kept is True

    def test_or_rule_deletes_file_with_only_noise(self):
        from cleanup import decide_contact
        thresholds = {"chiros": 0.5, "orthos": 0.5,
                       "micromam": 0.5, "oiseaux": 0.5}
        _, d, _ = decide_contact("noise", 0.9, thresholds, set())
        decisions = [d] * 5
        any_kept = any(dd == "kept" for dd in decisions)
        assert any_kept is False


# =========================================================================
# registry
# =========================================================================

class TestRegistry:
    def test_fresh_create_empty(self, tmp_path):
        from registry import Registry
        r = Registry(tmp_path)
        try:
            sessions = list(r.list_sessions()) if hasattr(r, "list_sessions") \
                else []
            assert len(sessions) == 0
        finally:
            r.close()

    def test_upsert_inserted_then_updated(self, tmp_path):
        from registry import Registry
        r = Registry(tmp_path)
        try:
            action1 = r.upsert_session({
                "id": "test_session_1",
                "canonical_name": "20250903_site212097_Z3_Pass2_enr07",
                "nom_contrat": "Contrat A",
            })
            assert action1 == "inserted"

            action2 = r.upsert_session({
                "id": "test_session_1",
                "canonical_name": "20250903_site212097_Z3_Pass2_enr07",
                "nom_contrat": "Contrat A",
            })
            assert action2 == "updated"
        finally:
            r.close()

    def test_thread_safe_upsert(self, tmp_path):
        """Upsert depuis un autre thread ne doit pas lever ProgrammingError."""
        import threading

        from registry import Registry
        r = Registry(tmp_path)
        try:
            results = []
            def worker():
                try:
                    a = r.upsert_session({
                        "id": "cross_thread",
                        "canonical_name": "cross",
                    })
                    results.append(("ok", a))
                except Exception as e:
                    results.append(("err", str(e)))
            t = threading.Thread(target=worker)
            t.start()
            t.join()
            assert results == [("ok", "inserted")]
        finally:
            r.close()

    def test_batch_feed_from_scan(self, tmp_path):
        """feed_from_scan doit être transactionnel (pas 1 commit/session)."""
        from registry import Registry
        r = Registry(tmp_path)
        try:
            # Simule un SessionState-like via dict
            states = [
                {"name": f"session_{i}", "path": f"/tmp/session_{i}",
                 "campaign": "TestCamp", "n_wav": i, "total_bytes": i * 1000}
                for i in range(10)
            ]
            result = r.feed_from_scan(states)
            assert result["inserted"] == 10
            assert result["updated"] == 0
            # Re-feed → tout en update
            result2 = r.feed_from_scan(states)
            assert result2["inserted"] == 0
            assert result2["updated"] == 10
        finally:
            r.close()


# =========================================================================
# materiels
# =========================================================================

class TestMateriels:
    def test_save_load_roundtrip(self, tmp_path, monkeypatch):
        import materiels as mm
        # Redirige le stockage vers un tempdir
        fake_path = tmp_path / "materiels.json"
        monkeypatch.setattr(mm, "_materiels_path", lambda: fake_path)

        ms = [
            mm.Materiel(id=1, marque="WA", modele="SM4BAT",
                         serie_enr="S4U001", micro_modele="SMX-U1"),
            mm.Materiel(id=7, marque="Audiomoth", serie_enr="AM007"),
        ]
        mm.save_materiels(ms)
        assert fake_path.is_file()

        loaded = mm.load_materiels()
        assert len(loaded) == 2
        assert loaded[0].id == 1
        assert loaded[0].modele == "SM4BAT"
        assert loaded[1].id == 7

    def test_find_by_id_and_serial(self, tmp_path, monkeypatch):
        import materiels as mm
        monkeypatch.setattr(mm, "_materiels_path",
                            lambda: tmp_path / "materiels.json")
        ms = [
            mm.Materiel(id=7, serie_enr="S4U04784", modele="SM4BAT"),
        ]
        mm.save_materiels(ms)
        loaded = mm.load_materiels()
        assert mm.find_by_id(loaded, 7).modele == "SM4BAT"
        assert mm.find_by_id(loaded, 99) is None
        assert mm.find_by_serial(loaded, "S4U04784").id == 7
        assert mm.find_by_serial(loaded, "s4u04784").id == 7  # case-insensitive

    def test_save_empty_refuses_to_overwrite(self, tmp_path, monkeypatch):
        """Le garde-fou anti-effacement : save([]) ne doit pas écraser
        une liste existante sauf si allow_clear=True."""
        import materiels as mm
        fake_path = tmp_path / "materiels.json"
        monkeypatch.setattr(mm, "_materiels_path", lambda: fake_path)

        # 1. On sauvegarde 2 matériels
        ms = [mm.Materiel(id=1, modele="SM4BAT"),
              mm.Materiel(id=7, modele="Audiomoth")]
        mm.save_materiels(ms)
        assert len(mm.load_materiels()) == 2

        # 2. Tentative d'écraser par []: REFUS
        mm.save_materiels([])
        assert len(mm.load_materiels()) == 2, \
            "liste vide ne doit PAS écraser une liste non-vide"

        # 3. Suppression explicite : OK
        mm.save_materiels([], allow_clear=True)
        assert len(mm.load_materiels()) == 0

    def test_save_creates_backup(self, tmp_path, monkeypatch):
        """save_materiels doit créer un .bak avant écrasement."""
        import materiels as mm
        fake_path = tmp_path / "materiels.json"
        monkeypatch.setattr(mm, "_materiels_path", lambda: fake_path)

        mm.save_materiels([mm.Materiel(id=1, modele="A")])
        # Second save → .bak créé
        mm.save_materiels([mm.Materiel(id=1, modele="B")])
        bak = fake_path.with_suffix(fake_path.suffix + ".bak")
        assert bak.is_file()

    def test_next_free_id(self):
        import materiels as mm
        ms = [mm.Materiel(id=1), mm.Materiel(id=3), mm.Materiel(id=5)]
        assert mm.next_free_id(ms) == 2
        assert mm.next_free_id(ms, start=4) == 4
        assert mm.next_free_id(ms, start=5) == 6

    def test_empty_materiel(self):
        import materiels as mm
        m = mm.Materiel(id=1)
        assert m.is_empty() is True
        m2 = mm.Materiel(id=2, modele="Test")
        assert m2.is_empty() is False


# =========================================================================
# rename : auto-cicatrisation des doublons de casse (régression v0.3.1)
# =========================================================================

class TestRenameAutoHeal:
    """Garde-fou contre la régression « Terminé avec erreurs » :

    un run de prep interrompu peut laisser, pour un même timestamp, la source
    (casse minuscule) ET sa cible canonique déjà renommée, byte-identiques.
    Le re-run NE DOIT PAS bloquer toute la session : la source redondante est
    mise en quarantaine, la cible conservée. Un vrai conflit (contenu
    différent) doit en revanche TOUJOURS bloquer (zéro écrasement).
    """

    def _meta(self):
        from naming import SessionMeta
        return SessionMeta(
            date_debut=datetime(2025, 9, 3), n_site_tadarida="212097",
            n_point_fixe="Z3", n_passage=2, n_enregistreur=7,
            n_serie="SMU03126", nom_contrat="T",
        )

    def _wav(self, path: Path, frames: int = 100):
        import wave
        with wave.open(str(path), "wb") as w:
            w.setnchannels(1)
            w.setsampwidth(2)
            w.setframerate(384000)
            w.writeframes(b"\x01\x02" * frames)

    def test_identical_existing_target_is_quarantined(self, tmp_path):
        import shutil

        from naming import compute_new_wav_name
        from rename import rename_session
        meta = self._meta()
        sess = tmp_path / "rawsess"
        sess.mkdir()
        # 1 fichier brut qui se renomme proprement
        self._wav(sess / "SMU03126_20250903_210000.wav")
        # 1 fichier brut (casse minuscule) dont la cible canonique existe DÉJÀ,
        # identique → doublon résiduel à cicatriser
        src_dup = sess / "smu03126_20250903_210523.wav"
        self._wav(src_dup)
        canon = compute_new_wav_name(meta, src_dup.name)
        assert canon is not None
        shutil.copy(src_dup, sess / canon)   # cible identique pré-existante

        res = rename_session(sess, meta, dry_run=False, rename_folder=False)

        assert not res.get("errors"), f"ne doit pas bloquer : {res.get('errors')}"
        assert res.get("quarantined", 0) == 1
        assert (sess / "_doublons_casse").is_dir()
        assert not src_dup.exists(), "la source redondante doit être déplacée"
        assert (sess / canon).is_file(), "la cible canonique doit être conservée"

    def test_different_existing_target_blocks(self, tmp_path):
        from naming import compute_new_wav_name
        from rename import rename_session
        meta = self._meta()
        sess = tmp_path / "rawsess2"
        sess.mkdir()
        src = sess / "smu03126_20250903_210523.wav"
        self._wav(src, frames=100)
        canon = compute_new_wav_name(meta, src.name)
        self._wav(sess / canon, frames=50)   # cible DIFFÉRENTE → vrai conflit

        res = rename_session(sess, meta, dry_run=False, rename_folder=False)

        assert res.get("errors"), "un conflit réel (contenu différent) doit bloquer"
        assert src.exists(), "la source ne doit pas être touchée en cas de conflit"


# =========================================================================
# synthesis : récap par espèce d'une nuit
# =========================================================================

class TestSynthesis:
    HEADERS = ["nom du fichier", "tadarida_taxon", "observateur_taxon"]

    def _rows(self):
        return [
            ["f1.wav", "Pippip", None],      # Tadarida chiros
            ["f1.wav", "Pippip", None],      # même espèce, même fichier
            ["f2.wav", "Nycnoc", "Pippip"],  # validé Pippip → écrase Nycnoc
            ["f3.wav", "noise", None],       # bruit
            ["f4.wav", None, None],          # aucune espèce → ignoré
            ["f5.wav", "", ""],              # vide → ignoré
        ]

    def test_counts_and_totals(self):
        from synthesis import compute_night_synthesis
        res = compute_night_synthesis(self.HEADERS, self._rows())
        by_taxon = {s["taxon"]: s for s in res["species"]}

        # Pippip : 3 contacts (2 Tadarida + 1 validé), 2 fichiers (f1, f2)
        assert by_taxon["Pippip"]["n_contacts"] == 3
        assert by_taxon["Pippip"]["n_fichiers"] == 2
        assert by_taxon["Pippip"]["validated"] is True
        assert by_taxon["Pippip"]["groupe"] == "chiros"

        # Nycnoc écrasé par la validation observateur → absent
        assert "Nycnoc" not in by_taxon

        # noise compté à part
        assert by_taxon["noise"]["n_contacts"] == 1

        assert res["total_contacts"] == 4          # 3 Pippip + 1 noise
        assert res["total_fichiers"] == 3          # f1, f2, f3
        assert res["by_group"]["chiros"] == 3
        assert res["by_group"]["noise"] == 1

    def test_sorted_desc(self):
        from synthesis import compute_night_synthesis
        res = compute_night_synthesis(self.HEADERS, self._rows())
        counts = [s["n_contacts"] for s in res["species"]]
        assert counts == sorted(counts, reverse=True)

    def test_missing_columns_graceful(self):
        from synthesis import compute_night_synthesis
        res = compute_night_synthesis(["a", "b"], [[1, 2], [3, 4]])
        assert res["species"] == []
        assert res["total_contacts"] == 0
        assert res["total_fichiers"] == 0

    def test_empty_rows(self):
        from synthesis import compute_night_synthesis
        res = compute_night_synthesis(self.HEADERS, [])
        assert res["total_contacts"] == 0


# =========================================================================
# validation : filtrage pur + menu taxons dynamique
# =========================================================================

class TestValidationFilters:
    CI = {"nom du fichier": 0, "tadarida_taxon": 1,
          "tadarida_probabilite": 2, "observateur_taxon": 3}

    def _rows(self):
        # [fichier, tadarida, proba, observateur]
        return [
            ["f1.wav", "Pippip", 0.9, None],
            ["f2.wav", "Pippip", 0.3, None],
            ["f3.wav", "Nycnoc", 0.8, None],
            ["f4.wav", "barbar", 0.2, "barbar"],   # patrimonial + validé
            ["f5.wav", "noise", 0.95, None],
        ]

    def test_proba_filter(self):
        from gui_validation import row_passes_filters as rp
        r = ["f", "Pippip", 0.3, None]
        assert rp(r, self.CI, proba_min=0.0) is True
        assert rp(r, self.CI, proba_min=0.5) is False

    def test_taxon_filter(self):
        from gui_validation import row_passes_filters as rp
        r = ["f", "Pippip", 0.9, None]
        assert rp(r, self.CI, taxon_filter="Pippip") is True
        assert rp(r, self.CI, taxon_filter="Nycnoc") is False
        assert rp(r, self.CI, taxon_filter="Tous") is True

    def test_chiros_only(self):
        from gui_validation import row_passes_filters as rp
        assert rp(["f", "Pippip", 0.9, None], self.CI, only_chiros=True) is True
        assert rp(["f", "noise", 0.9, None], self.CI, only_chiros=True) is False

    def test_hide_cleaned(self):
        from gui_validation import row_passes_filters as rp
        r = ["f", "Pippip", 0.9, None]
        assert rp(r, self.CI, hide_cleaned=True, wav_present=True) is True
        assert rp(r, self.CI, hide_cleaned=True, wav_present=False) is False
        assert rp(r, self.CI, hide_cleaned=False, wav_present=False) is True

    def test_only_unvalidated(self):
        from gui_validation import row_passes_filters as rp
        assert rp(["f", "Pippip", 0.9, "Pippip"], self.CI,
                  only_unvalidated=True) is False
        assert rp(["f", "Pippip", 0.9, None], self.CI,
                  only_unvalidated=True) is True

    def test_only_observer_taxon(self):
        from gui_validation import row_passes_filters as rp
        assert rp(["f", "Pippip", 0.9, "Pippip"], self.CI,
                  only_observer_taxon=True) is True
        assert rp(["f", "Pippip", 0.9, None], self.CI,
                  only_observer_taxon=True) is False
        assert rp(["f", "Pippip", 0.9, ""], self.CI,
                  only_observer_taxon=True) is False
        # Les deux filtres opposés : rien ne passe
        assert rp(["f", "Pippip", 0.9, "Pippip"], self.CI,
                  only_unvalidated=True, only_observer_taxon=True) is False

    def test_only_patrimonial(self):
        from gui_validation import row_passes_filters as rp, PATRIMONIAL_CODES
        assert rp(["f", "barbar", 0.9, None], self.CI, only_patrimonial=True,
                  patrimonial_codes=PATRIMONIAL_CODES) is True
        assert rp(["f", "Pippip", 0.9, None], self.CI, only_patrimonial=True,
                  patrimonial_codes=PATRIMONIAL_CODES) is False

    def test_eligible_taxons_all(self):
        from gui_validation import eligible_taxons
        t = eligible_taxons(self._rows(), self.CI, proba_min=0.0)
        assert set(t) == {"Pippip", "Nycnoc", "barbar", "noise"}

    def test_eligible_taxons_reflects_proba(self):
        from gui_validation import eligible_taxons
        # proba ≥ 0.5 : barbar (0.2) et Pippip/f2 (0.3) sortent ; Pippip reste (f1=0.9)
        t = eligible_taxons(self._rows(), self.CI, proba_min=0.5)
        assert set(t) == {"Pippip", "Nycnoc", "noise"}
        assert "barbar" not in t

    def test_eligible_taxons_reflects_hide_cleaned(self):
        from gui_validation import eligible_taxons
        # f3 (Nycnoc) supprimé au nettoyage → disparaît de la liste
        wav_present = [True, True, False, True, True]
        t = eligible_taxons(self._rows(), self.CI, hide_cleaned=True,
                            wav_present=wav_present)
        assert "Nycnoc" not in t
        assert "Pippip" in t

    def test_eligible_taxons_observer_only(self):
        from gui_validation import eligible_taxons
        t = eligible_taxons(self._rows(), self.CI, only_observer_taxon=True)
        assert t == ["barbar"]

    def test_sort_proba_desc_missing_last(self):
        from gui_validation import sort_filtered_indexes
        ci = {"tadarida_probabilite": 0, "nom du fichier": 1}
        rows = [[0.2, "a"], [0.9, "b"], [None, "c"], [0.5, "d"]]
        out = sort_filtered_indexes([0, 1, 2, 3], rows, ci, "tad_proba",
                                    descending=True)
        assert out == [1, 3, 0, 2]          # 0.9, 0.5, 0.2, puis manquant

    def test_sort_taxon_asc(self):
        from gui_validation import sort_filtered_indexes
        ci = {"tadarida_taxon": 0}
        rows = [["Pippip"], ["barbar"], ["Nycnoc"]]
        out = sort_filtered_indexes([0, 1, 2], rows, ci, "tad_taxon")
        assert out == [1, 2, 0]              # barbar, Nycnoc, Pippip

    def test_count_observer_progress(self):
        from gui_validation import count_observer_progress
        headers = ["nom du fichier", "tadarida_taxon", "observateur_taxon"]
        rows = [
            ["a", "pippip", "pippip"],
            ["b", "noise", ""],
            ["c", "barbar", None],
            [None, None, None],
        ]
        d = count_observer_progress(headers, rows)
        assert d["n_total"] == 3
        assert d["n_validated"] == 1

    def test_benjamin_vu_coherent_with_synthesis(self):
        """_Vu issue #3 : synthèse « validés seulement » = compteur observateur."""
        from pathlib import Path
        from chirosurf_nights import read_csv
        from synthesis import compute_night_synthesis
        from gui_validation import count_observer_progress
        sample = Path(__file__).resolve().parent.parent / (
            "samples/issue3_benjamin/Nuit_1-observations_Vu.csv")
        if not sample.is_file():
            import pytest
            pytest.skip("samples issue #3 absents")
        headers, rows = read_csv(sample)
        c = count_observer_progress(headers, rows)
        s_all = compute_night_synthesis(headers, rows)
        s_vo = compute_night_synthesis(headers, rows, validated_only=True)
        assert c["n_total"] == 8000
        assert c["n_validated"] == 16
        assert s_all["validated_contacts"] == 16
        assert s_vo["total_contacts"] == 16
        assert s_all["total_contacts"] == 8000


# =========================================================================
# activity_graph : cascade des filtres (site → point → passage → nuit)
# =========================================================================

class TestActivityCascade:
    def _agg(self):
        # clé = (site, point, passage, night, taxon) -> bins
        return {
            ("212097", "Z1", 1, "2026-06-01", "Pippip"): [1],
            ("212097", "Z2", 1, "2026-06-01", "Nycnoc"): [1],
            ("212097", "Z1", 2, "2026-06-02", "Pippip"): [1],
            ("999999", "A1", 1, "2026-05-01", "Barbar"): [1],
        }

    def test_no_selection_all_options(self):
        from activity_graph import cascade_options
        o = cascade_options(self._agg())
        assert o["sites"] == ["212097", "999999"]
        assert set(o["points"]) == {"Z1", "Z2", "A1"}
        assert o["passages"] == [1, 2]
        assert o["nights"] == ["2026-05-01", "2026-06-01", "2026-06-02"]

    def test_cascade_by_site(self):
        from activity_graph import cascade_options
        o = cascade_options(self._agg(), sel_sites={"212097"})
        assert set(o["points"]) == {"Z1", "Z2"}     # A1 exclu (autre site)
        assert o["passages"] == [1, 2]
        assert o["nights"] == ["2026-06-01", "2026-06-02"]  # 2026-05-01 exclu

    def test_cascade_by_site_and_point(self):
        from activity_graph import cascade_options
        o = cascade_options(self._agg(), sel_sites={"212097"}, sel_points={"Z1"})
        # passages/nuits limités à Z1 du site 212097
        assert o["passages"] == [1, 2]
        assert o["nights"] == ["2026-06-01", "2026-06-02"]

    def test_cascade_full_chain(self):
        from activity_graph import cascade_options
        o = cascade_options(self._agg(), sel_sites={"212097"},
                            sel_points={"Z1"}, sel_passages={1})
        # Z1 + passage 1 => uniquement la nuit du 2026-06-01
        assert o["nights"] == ["2026-06-01"]

    def test_placeholder_ordering(self):
        from activity_graph import cascade_options
        agg = {
            ("212097", "Z1", 1, "2026-06-01", "Pippip"): [1],
            ("?", "?", None, "2026-06-01", "noise"): [1],
        }
        o = cascade_options(agg)
        assert o["sites"] == ["212097", "?"]          # "?" repoussé en fin


class TestActivityAggregate:
    HEADERS = ["nom du fichier", "tadarida_taxon", "observateur_taxon",
               "validateur_taxon"]

    def _fname(self, hms="210000"):
        return f"Car212097-2026-Pass1-Z1-SMU03126_20260821_{hms}.wav"

    def test_observer_taxon_groups_under_observer_code(self):
        from activity_graph import aggregate_rows
        rows = [
            [self._fname(), "Nycnoc", "Nyclas", None],
            [self._fname("210500"), "Pippip", None, None],
        ]
        all_t = aggregate_rows(self.HEADERS, rows, bin_minutes=30)
        taxons = {k[-1] for k in all_t}
        assert "Nyclas" in taxons
        assert "Nycnoc" not in taxons  # observateur prime
        obs_only = aggregate_rows(
            self.HEADERS, rows, bin_minutes=30, use_observer_taxon=True)
        assert {k[-1] for k in obs_only} == {"Nyclas"}

    def test_chiros_only_drops_noise(self):
        from activity_graph import aggregate_rows
        rows = [
            [self._fname(), "Pippip", None, None],
            [self._fname("211000"), "noise", None, None],
        ]
        out = aggregate_rows(
            self.HEADERS, rows, bin_minutes=30, chiros_only=True)
        assert {k[-1] for k in out} == {"Pippip"}

    def test_human_validated_accepts_observateur(self):
        from activity_graph import aggregate_rows
        rows = [
            [self._fname(), "Nycnoc", "Nyclas", None],
            [self._fname("211000"), "Pippip", None, None],
        ]
        out = aggregate_rows(
            self.HEADERS, rows, bin_minutes=30, use_only_validated=True)
        assert {k[-1] for k in out} == {"Nyclas"}

    def test_synthesis_chiros_only(self):
        from synthesis import compute_night_synthesis
        headers = ["nom du fichier", "tadarida_taxon", "observateur_taxon"]
        rows = [
            ["a.wav", "Pippip", None],
            ["b.wav", "noise", None],
        ]
        res = compute_night_synthesis(headers, rows, chiros_only=True)
        assert [s["taxon"] for s in res["species"]] == ["Pippip"]
        assert res["total_contacts"] == 1


# =========================================================================
# te10 : robustesse backend Python (statuts, erreurs, pas d'avortement)
# =========================================================================

class TestTe10Robustness:
    def _wav(self, path, frames=100, sr=384000):
        import wave
        with wave.open(str(path), "wb") as w:
            w.setnchannels(1)
            w.setsampwidth(2)
            w.setframerate(sr)
            w.writeframes(b"\x01\x02" * frames)

    def test_write_segment_states(self, tmp_path):
        from te10 import write_segment, plan_file
        src = tmp_path / "s_20250101_000000.wav"
        self._wav(src)
        out = tmp_path / "out"
        plans = plan_file(src, out, 10, 5.0)
        assert plans
        st, _ = write_segment(plans[0])
        assert st == "written"
        assert plans[0].dst.is_file()
        st2, _ = write_segment(plans[0])   # existe déjà
        assert st2 == "skipped"

    def test_write_segment_error_never_raises(self, tmp_path):
        """Une source illisible renvoie 'error' — jamais d'exception (sinon tout
        le lot avorterait)."""
        from te10 import write_segment, SegmentPlan
        bad = tmp_path / "bad.wav"
        bad.write_text("pas un wav")
        plan = SegmentPlan(src=bad, dst=tmp_path / "out" / "x_000.wav",
                           start_frame=0, n_frames=10, sr_te=38400)
        st, msg = write_segment(plan)      # ne doit pas lever
        assert st == "error"
        assert not plan.dst.exists()

    def test_process_folder_counts_errors_not_skipped(self, tmp_path):
        """Backend Python : un WAV corrompu est compté en 'errors' (pas 'skipped')
        et n'avorte PAS la session ; le WAV valide est bien produit."""
        from te10 import process_folder
        src = tmp_path / "src"
        src.mkdir()
        self._wav(src / "a_20250101_000000.wav")           # valide
        (src / "b_20250101_000100.wav").write_text("corrompu")  # illisible
        stats = process_folder(src, tmp_path / "out", jobs=1, force_python=True)
        assert stats["written"] >= 1        # le valide est traité malgré l'autre
        assert stats["errors"] >= 1         # le corrompu compté en erreur
        assert stats["engine"] == "python"

    def test_plan_file_splits_15s_wav_not_truncated(self, tmp_path):
        """Un WAV de 15 s raw doit donner 3 segments de 5 s (tout le son, pas
        seulement les 5 premières secondes). Noms distincts (timestamp +5 s)."""
        from te10 import plan_file
        sr = 8000
        src = tmp_path / "s_20250101_120000.wav"
        self._wav(src, frames=15 * sr, sr=sr)
        plans = plan_file(src, tmp_path / "out", 10, 5.0)
        assert len(plans) == 3
        assert [p.n_frames for p in plans] == [5 * sr, 5 * sr, 5 * sr]
        names = [p.dst.name for p in plans]
        assert len(set(names)) == 3
        assert "120000" in names[0]
        assert "120005" in names[1]
        assert "120010" in names[2]
        assert all(n.endswith("_000.wav") for n in names)

    def test_process_folder_writes_all_segments(self, tmp_path):
        from te10 import process_folder
        sr = 8000
        src = tmp_path / "src"
        src.mkdir()
        self._wav(src / "s_20250101_120000.wav", frames=15 * sr, sr=sr)
        out = tmp_path / "out"
        stats = process_folder(src, out, jobs=1, force_python=True)
        assert stats["n_planned_segments"] == 3
        assert stats["written"] == 3
        assert stats["errors"] == 0
        assert len(list(out.glob("*.wav"))) == 3


# =========================================================================
# pipeline : fenetre temporelle de participation (bug date_fin a minuit)
# =========================================================================

class TestParticipationWindow:
    def test_midnight_becomes_night(self):
        from pipeline import _participation_window
        d0, d1 = _participation_window(datetime(2026, 6, 23))   # minuit
        assert d0 == datetime(2026, 6, 23, 20, 0)     # décalé à 20:00
        assert d1 == datetime(2026, 6, 24, 6, 0)      # fin le lendemain 06:00
        assert d1 > d0

    def test_evening_start_ends_next_morning(self):
        from pipeline import _participation_window
        dd = datetime(2026, 6, 23, 21, 30)
        d0, d1 = _participation_window(dd)
        assert d0 == dd
        assert d1 == datetime(2026, 6, 24, 6, 0)

    def test_explicit_date_fin_untouched(self):
        from pipeline import _participation_window
        dd = datetime(2026, 6, 23)
        df = datetime(2026, 6, 24, 5, 0)
        assert _participation_window(dd, df) == (dd, df)

    def test_early_morning_same_day_end(self):
        from pipeline import _participation_window
        dd = datetime(2026, 6, 23, 2, 0)
        d0, d1 = _participation_window(dd)
        assert d0 == dd
        assert d1 == datetime(2026, 6, 23, 6, 0)      # 06:00 le jour même > 02:00
        assert d1 > d0

    def test_none_date_debut(self):
        from pipeline import _participation_window
        assert _participation_window(None) == (None, None)


# =========================================================================
# cleanup : appariement WAV<->xlsx insensible a la casse (regression)
# =========================================================================

class TestCleanupCaseInsensitive:
    def _wav(self, path, frames=50):
        import wave
        with wave.open(str(path), "wb") as w:
            w.setnchannels(1)
            w.setsampwidth(2)
            w.setframerate(38400)
            w.writeframes(b"\x01\x02" * frames)

    def test_norm_stem_lowercases(self):
        from cleanup import _norm_stem
        assert _norm_stem("2MU08078_X.wav") == _norm_stem("2mu08078_x")

    def test_case_divergent_wav_kept(self, tmp_path):
        """Un WAV dont la casse diffère entre le xlsx et le disque, mais porteur
        d'un contact 'kept', NE DOIT PAS être supprimé (avant : il tombait dans
        silent_files → supprimé avec silent_policy='delete'). 3 WAV pour que le
        divergent soit minoritaire (le garde-fou mass-delete ne le protège pas)."""
        from cleanup import cleanup_session
        session = tmp_path / "sess"
        dk = session / "Data_k"
        dk.mkdir(parents=True)
        self._wav(dk / "a_20250101_210000_000.wav")
        self._wav(dk / "b_20250101_210100_000.wav")
        self._wav(dk / "2MU_20250101_210200_000.wav")     # disque : MAJUSCULES
        csv_path = session / "participation-test-observations.csv"
        csv_path.write_text(
            "nom du fichier,tadarida_taxon,tadarida_probabilite\n"
            "a_20250101_210000_000.wav,Pippip,0.9\n"
            "b_20250101_210100_000.wav,Pippip,0.9\n"
            "2mu_20250101_210200_000.wav,Pippip,0.9\n",   # xlsx : minuscules
            encoding="utf-8",
        )
        res = cleanup_session(
            session,
            thresholds={"chiros": 0.5, "orthos": 0.5,
                        "micromam": 0.5, "oiseaux": 0.5},
            disabled=set(), silent_policy="delete", dry_run=False,
        )
        assert not res.get("errors"), res.get("errors")
        # Les 3 WAV portent un contact kept → aucun ne doit être supprimé.
        assert (dk / "2MU_20250101_210200_000.wav").exists()
        assert (dk / "a_20250101_210000_000.wav").exists()
        assert (dk / "b_20250101_210100_000.wav").exists()


# =========================================================================
# securite : stockage token (fallback) + base_url https
# =========================================================================

class TestCredentialsFallback:
    TOKEN = "ABCD1234ABCD1234ABCD1234ABCD1234"

    def test_fallback_roundtrip_and_not_plaintext(self, tmp_path, monkeypatch):
        import os as _os

        import credentials as cr
        monkeypatch.setattr(cr, "_try_keyring", lambda: None)   # force le fichier
        monkeypatch.setattr(cr, "_fallback_dir", lambda: tmp_path)

        assert cr.save_token(self.TOKEN) == "file"
        assert cr.load_token() == self.TOKEN

        content = (tmp_path / "credentials.json").read_text(encoding="utf-8")
        if _os.name == "nt":
            # Sous Windows, DPAPI doit avoir chiffré : pas de token en clair.
            assert self.TOKEN not in content
            assert "token_dpapi" in content

        assert cr.delete_token() is True
        assert cr.load_token() is None

    def test_legacy_plaintext_still_readable(self, tmp_path, monkeypatch):
        """Compat : un ancien credentials.json en clair reste lisible."""
        import json

        import credentials as cr
        monkeypatch.setattr(cr, "_try_keyring", lambda: None)
        monkeypatch.setattr(cr, "_fallback_dir", lambda: tmp_path)
        (tmp_path / "credentials.json").write_text(
            json.dumps({"token": self.TOKEN}), encoding="utf-8")
        assert cr.load_token() == self.TOKEN


class TestApiBaseUrlSecurity:
    TOKEN = "A" * 32

    def test_rejects_http_non_localhost(self):
        from vigiechiro_api import VigieChiroClient
        with pytest.raises(ValueError):
            VigieChiroClient(self.TOKEN,
                             base_url="http://vigiechiro.example.com/api/v1")

    def test_allows_http_localhost(self):
        from vigiechiro_api import VigieChiroClient
        c = VigieChiroClient(self.TOKEN, base_url="http://localhost:8080/api/v1")
        assert c.base_url.startswith("http://localhost")

    def test_allows_https(self):
        from vigiechiro_api import VigieChiroClient
        c = VigieChiroClient(self.TOKEN,
                             base_url="https://vigiechiro.example.com/api/v1")
        assert c.base_url.startswith("https://")


# =========================================================================
# rename : mode compatible antivirus (_av_safe_pace via variable d'env)
# =========================================================================

class TestAvSafePace:
    def test_env_activation(self, monkeypatch):
        import rename
        # Valeurs explicites via env : prioritaires (independant du marqueur).
        monkeypatch.setenv("CHIROTOOL_AV_SAFE", "1")
        assert rename._av_safe_pace() == (40, 0.25)
        monkeypatch.setenv("CHIROTOOL_AV_SAFE", "25:400")
        assert rename._av_safe_pace() == (25, 0.4)
        monkeypatch.setenv("CHIROTOOL_AV_SAFE", "0")   # desactivation explicite
        assert rename._av_safe_pace() is None


# =========================================================================
# coherence : User-Agent Nominatim derive de version.py
# =========================================================================

class TestNominatimUserAgent:
    def test_ua_reflects_version(self):
        import gui_map
        from version import __version__
        assert gui_map.NOMINATIM_UA.startswith("ChiroTool/")
        assert __version__ in gui_map.NOMINATIM_UA


# =========================================================================
# carte : points sur carres d'autres observateurs (sites externes)
# =========================================================================

class TestMapExternalSites:
    def test_points_from_raw_latlon_order(self):
        import gui_map
        raw = {"localites": [
            {"nom": "Z1", "geometries": {"geometries": [
                {"type": "Point", "coordinates": [45.5, 4.2]}]}},   # [lat, lon]
            {"nom": "Z2", "geometries": {"geometries": [
                {"type": "LineString", "coordinates": [[1, 2], [3, 4]]}]}},
        ]}
        pts = gui_map._points_from_raw(raw)
        assert pts == [{"nom": "Z1", "lat": 45.5, "lon": 4.2}]

    def test_external_site_ids_persist(self, tmp_path, monkeypatch):
        import gui_map
        monkeypatch.setattr(gui_map, "_external_sites_path",
                            lambda: tmp_path / "external_sites.json")
        assert gui_map._load_external_site_ids() == []
        gui_map._remember_external_site_id("abc123")
        gui_map._remember_external_site_id("abc123")   # idempotent
        gui_map._remember_external_site_id("def456")
        assert gui_map._load_external_site_ids() == ["abc123", "def456"]


# =========================================================================
# carte / wizard : point actif (continuité create|reuse → meta)
# =========================================================================

class TestActivePoint:
    def test_save_load_roundtrip(self, tmp_path, monkeypatch):
        import gui_map
        monkeypatch.setattr(gui_map, "_active_point_path",
                            lambda: tmp_path / "active_point.json")
        assert gui_map.load_active_point() is None
        saved = gui_map.save_active_point(
            site_id="sid1", numero="381009", point="z1",
            lat=45.1, lon=5.2, is_mine=False, source="reuse",
        )
        assert saved["point"] == "Z1"
        assert saved["numero"] == "381009"
        assert saved["is_mine"] is False
        assert saved["source"] == "reuse"
        assert saved["selected_at"]
        loaded = gui_map.load_active_point()
        assert loaded is not None
        assert loaded["site_id"] == "sid1"
        assert loaded["point"] == "Z1"
        assert loaded["numero"] == "381009"
        assert loaded["is_mine"] is False

    def test_freshness_window(self, tmp_path, monkeypatch):
        import gui_map
        from datetime import datetime, timedelta
        monkeypatch.setattr(gui_map, "_active_point_path",
                            lambda: tmp_path / "active_point.json")
        now = datetime(2026, 8, 3, 12, 0, 0)
        gui_map.save_active_point(
            site_id="x", numero="123456", point="A1",
            is_mine=True, source="create",
        )
        ap = gui_map.load_active_point()
        # Forcer selected_at pour le test
        ap["selected_at"] = (now - timedelta(days=3)).isoformat(timespec="seconds")
        assert gui_map.active_point_is_fresh(ap, now=now) is True
        ap["selected_at"] = (now - timedelta(days=8)).isoformat(timespec="seconds")
        assert gui_map.active_point_is_fresh(ap, now=now) is False
        assert gui_map.active_point_is_fresh(None) is False

    def test_should_prefill_priority(self):
        """N'écrase pas un guess complet ; comble si site ou point manquant."""
        import gui_map
        from datetime import datetime, timedelta
        now = datetime(2026, 8, 3, 12, 0, 0)
        ap = {
            "numero": "381009", "point": "Z1", "is_mine": False,
            "selected_at": now.isoformat(timespec="seconds"),
        }
        # Session déjà complète → non
        assert gui_map.should_prefill_from_active(
            "111111", "A2", ap, now=now) is False
        # Site seul → oui (point manquant)
        assert gui_map.should_prefill_from_active(
            "111111", None, ap, now=now) is True
        # Rien → oui
        assert gui_map.should_prefill_from_active(
            None, None, ap, now=now) is True
        # Actif trop vieux → non même sans meta
        old = dict(ap)
        old["selected_at"] = (now - timedelta(days=10)).isoformat(
            timespec="seconds")
        assert gui_map.should_prefill_from_active(
            None, None, old, now=now) is False

    def test_merge_recent_active_first_and_dedupe(self):
        import gui_map
        from datetime import datetime
        now = datetime(2026, 8, 3, 12, 0, 0)
        ap = {
            "site_id": "ext1", "numero": "381009", "point": "Z1",
            "is_mine": False, "source": "reuse",
            "selected_at": now.isoformat(timespec="seconds"),
            "lat": 45.0, "lon": 5.0,
        }
        mine = [
            {"numero": "212097", "point": "A1", "is_mine": True,
             "updated": "2026-07-01"},
            # Doublon du point actif (même numero/point) — l'actif gagne
            {"numero": "381009", "point": "Z1", "is_mine": True,
             "updated": "2026-07-02"},
        ]
        external = [
            {"numero": "381009", "point": "Z2", "is_mine": False,
             "updated": "2026-06-01"},
        ]
        merged = gui_map.merge_recent_points(mine, external, ap, now=now)
        assert merged[0].get("is_active") is True
        assert merged[0]["numero"] == "381009"
        assert merged[0]["point"] == "Z1"
        # Pas de second 381009/Z1
        keys = [(p["numero"], p["point"]) for p in merged]
        assert keys.count(("381009", "Z1")) == 1
        assert ("212097", "A1") in keys
        assert ("381009", "Z2") in keys
        # Entrées externes marquées
        z2 = next(p for p in merged if p["point"] == "Z2")
        assert z2["is_mine"] is False

    def test_reuse_also_remembers_external(self, tmp_path, monkeypatch):
        """Simule le fix P0 : reuse d'un site non-mien → external_sites + active."""
        import gui_map
        monkeypatch.setattr(gui_map, "_external_sites_path",
                            lambda: tmp_path / "external_sites.json")
        monkeypatch.setattr(gui_map, "_active_point_path",
                            lambda: tmp_path / "active_point.json")
        # Comme _on_add_point_confirmed pour un reuse externe
        site_id = "foreign_site_id"
        is_mine = False
        if site_id and not is_mine:
            gui_map._remember_external_site_id(site_id)
        gui_map.save_active_point(
            site_id=site_id, numero="381009", point="Z1",
            lat=45.0, lon=5.0, is_mine=False, source="reuse",
        )
        assert "foreign_site_id" in gui_map._load_external_site_ids()
        ap = gui_map.load_active_point()
        assert ap["source"] == "reuse"
        assert ap["is_mine"] is False
        assert gui_map.should_prefill_from_active(None, None, ap) is True


# =========================================================================
# version : selection de la release (liste + /latest, anti-cache-obsolete)
# =========================================================================

class TestBestRelease:
    def test_latest_beats_stale_list(self):
        """Cas reel : la LISTE est servie en cache obsolete (max = 0.3.1) mais
        /releases/latest est a jour (0.4.0). Le combine doit renvoyer 0.4.0."""
        from version import _best_release
        candidates = [
            {"tag_name": "v0.3.1", "prerelease": False, "html_url": "u/031"},
            {"tag_name": "V0.3.0", "prerelease": False, "html_url": "u/030"},
            {"tag_name": "v0.4.0", "prerelease": False, "html_url": "u/040"},  # /latest
        ]
        best = _best_release(candidates)
        assert best["tag"] == "v0.4.0"
        assert best["prerelease"] is False

    def test_ignores_drafts(self):
        from version import _best_release
        candidates = [
            {"tag_name": "v0.4.0", "draft": True, "html_url": "u/040"},   # brouillon
            {"tag_name": "v0.3.1", "prerelease": False, "html_url": "u/031"},
        ]
        assert _best_release(candidates)["tag"] == "v0.3.1"

    def test_empty(self):
        from version import _best_release
        assert _best_release([]) is None


# =========================================================================
# AudioMoth : format date_time sans serie (issue #1)
# =========================================================================

class TestAudioMoth:
    """Format AudioMoth (issue #1). Validé sur de vrais fichiers CEREMA/TWAV_splitter :
    l'EXPANDÉ `date_time_ms.WAV` est traité ; le BRUT `date_time T.WAV` (déclenché,
    concaténé) est refusé (il doit d'abord être expandé)."""

    def _meta(self):
        from naming import SessionMeta
        return SessionMeta(
            date_debut=datetime(2022, 7, 21), n_site_tadarida="430658",
            n_point_fixe="Z1", n_passage=1, n_enregistreur=5,
            n_serie="24F319", nom_contrat="T")

    # --- format EXPANDÉ (date_time_ms) : traité ---
    def test_expanded_timestamp(self):
        from naming import extract_timestamp_from_name
        ts = extract_timestamp_from_name("20220721_033014_451.WAV")   # avec ms
        assert ts is not None and ts[0] == datetime(2022, 7, 21, 3, 30, 14)
        ts2 = extract_timestamp_from_name("20260615_212501.WAV")      # sans ms
        assert ts2 is not None and ts2[0] == datetime(2026, 6, 15, 21, 25, 1)

    def test_expanded_classified_raw_no_serial(self):
        from chiro_core import classify_wav_name
        from naming import extract_serial_from_name
        assert classify_wav_name("20220721_033014_451.WAV") == "raw"
        assert extract_serial_from_name("20220721_033014_451.WAV") is None

    def test_expanded_canonical_preserves_ms(self):
        from naming import compute_new_wav_name
        new = compute_new_wav_name(self._meta(), "20220721_033014_451.WAV")
        assert new == "Car430658-2022-Pass1-Z1-24F319_20220721_033014_451.wav"

    # --- format BRUT T.WAV : détecté et refusé ---
    def test_raw_twav_detected_not_processed(self):
        from chiro_core import classify_wav_name, is_audiomoth_twav
        from naming import extract_timestamp_from_name
        n = "20260615_212501T.WAV"
        assert is_audiomoth_twav(n) is True
        assert classify_wav_name(n) == "audiomoth_twav"
        assert extract_timestamp_from_name(n) is None   # PAS traité tel quel

    def test_rename_refuses_raw_twav(self, tmp_path):
        from rename import rename_session
        d = tmp_path / "sess" / "Data"
        d.mkdir(parents=True)
        (d / "20260615_212501T.WAV").write_bytes(b"RIFFxxxxWAVE")   # dummy
        res = rename_session(tmp_path / "sess", self._meta(),
                             dry_run=False, rename_folder=False)
        assert res.get("audiomoth_twav_raw") == 1
        assert res.get("errors")                       # refus explicite
        assert (d / "20260615_212501T.WAV").exists()   # non renommé

    def test_wildlife_still_works(self):
        """Non-régression : le format Wildlife (série en tête) reste géré."""
        from naming import compute_new_wav_name, extract_serial_from_name
        assert extract_serial_from_name("2MU08078_20260623_211115.wav") == "2MU08078"
        new = compute_new_wav_name(self._meta(), "2MU08078_20260623_211115.wav")
        assert "_20260623_211115" in new and new.startswith("Car430658-2022-Pass1-Z1-")


# =========================================================================
# Titley Swift / Ranger (issue #4, diagnostic Mickaël 2026-08-27)
# =========================================================================

class TestTitleyNaming:
    def _meta(self):
        from naming import SessionMeta
        return SessionMeta(
            date_debut=datetime(2026, 8, 21), n_site_tadarida="630230",
            n_point_fixe="Z1", n_passage=1, n_enregistreur=5,
            n_serie="SWIFT01", nom_contrat="T")

    def _wav(self, path: Path, frames: int = 100, sr: int = 8000):
        import wave
        with wave.open(str(path), "wb") as w:
            w.setnchannels(1)
            w.setsampwidth(2)
            w.setframerate(sr)
            w.writeframes(b"\x01\x02" * frames)

    def test_timestamp_space_and_underscore(self):
        from naming import extract_timestamp_from_name
        ts = extract_timestamp_from_name("2026-08-21 20-45-29.wav")
        assert ts is not None and ts[0] == datetime(2026, 8, 21, 20, 45, 29)
        ts2 = extract_timestamp_from_name("2026-08-21_20-45-29.wav")
        assert ts2 is not None and ts2[0] == datetime(2026, 8, 21, 20, 45, 29)

    def test_timestamp_with_device_prefix(self):
        from naming import extract_serial_from_name, extract_timestamp_from_name
        n = "669178 2026-08-21 20-45-29.wav"
        ts = extract_timestamp_from_name(n)
        assert ts is not None and ts[0] == datetime(2026, 8, 21, 20, 45, 29)
        assert extract_serial_from_name(n) is None  # pas une série Wildlife

    def test_classified_raw(self):
        from chiro_core import classify_wav_name
        assert classify_wav_name("2026-08-21 20-45-29.wav") == "raw"
        assert classify_wav_name("669178_2026-08-21_20-45-29.wav") == "raw"

    def test_compute_new_wav_name(self):
        from naming import compute_new_wav_name
        new = compute_new_wav_name(self._meta(), "2026-08-21 20-45-29.wav")
        assert new is not None
        assert new.startswith("Car630230-2026-Pass1-Z1-SWIFT01_")
        assert "_20260821_204529" in new

    def test_rename_titley_executes(self, tmp_path):
        from rename import rename_session
        sess = tmp_path / "swift"
        sess.mkdir()
        self._wav(sess / "2026-08-21 20-45-29.wav")
        res = rename_session(sess, self._meta(), dry_run=False, rename_folder=False)
        assert not res.get("errors"), res.get("errors")
        assert res.get("unreadable") == []
        assert res["executed"] == 1
        leftover = [p.name for p in sess.iterdir() if p.suffix.lower() == ".wav"]
        assert leftover == ["Car630230-2026-Pass1-Z1-SWIFT01_20260821_204529.wav"]

    def test_rename_stops_if_no_name_readable(self, tmp_path):
        from manifest import Manifest
        from rename import rename_session
        sess = tmp_path / "bad"
        sess.mkdir()
        self._wav(sess / "garbage.wav")
        res = rename_session(sess, self._meta(), dry_run=False, rename_folder=False)
        assert res.get("errors")
        assert "Aucun nom de fichier lisible" in res["errors"][0]
        m = Manifest.load(sess)
        assert m is None or not m.is_done("rename")

    def test_plan_file_titley_15s_distinct_names(self, tmp_path):
        """Défense TE×10 : même sans rename, les 3 tranches ont des noms distincts."""
        from te10 import plan_file
        sr = 8000
        src = tmp_path / "2026-08-21 20-45-29.wav"
        self._wav(src, frames=15 * sr, sr=sr)
        plans = plan_file(src, tmp_path / "out", 10, 5.0)
        assert len(plans) == 3
        names = [p.dst.name for p in plans]
        assert len(set(names)) == 3
        assert "20-45-29" in names[0]
        assert "20-45-34" in names[1]
        assert "20-45-39" in names[2]


# =========================================================================
# Dates Summary vs WAV (Jeanne)
# =========================================================================

class TestWavDatesOverSummary:
    def test_same_night_keeps_summary(self):
        from chiro_core import SummaryInfo, should_prefer_wav_dates
        s = SummaryInfo(
            path=Path("x"),
            start_dt=datetime(2026, 8, 21, 20, 0),
            end_dt=datetime(2026, 8, 22, 6, 0),
        )
        wav = datetime(2026, 8, 21, 20, 45)
        assert should_prefer_wav_dates(s, wav) is False

    def test_other_calendar_day_prefers_wav(self):
        from chiro_core import SummaryInfo, should_prefer_wav_dates
        s = SummaryInfo(
            path=Path("x"),
            start_dt=datetime(2026, 8, 18, 19, 0),
            end_dt=datetime(2026, 8, 22, 6, 0),
        )
        wav = datetime(2026, 8, 21, 20, 45)
        assert should_prefer_wav_dates(s, wav) is True

    def test_long_span_same_start_day_prefers_wav(self):
        from chiro_core import SummaryInfo, should_prefer_wav_dates
        s = SummaryInfo(
            path=Path("x"),
            start_dt=datetime(2026, 8, 21, 0, 0),
            end_dt=datetime(2026, 8, 23, 6, 0),
        )
        wav = datetime(2026, 8, 21, 20, 45)
        assert should_prefer_wav_dates(s, wav) is True

    def test_no_wav_keeps_summary(self):
        from chiro_core import SummaryInfo, should_prefer_wav_dates
        s = SummaryInfo(path=Path("x"), start_dt=datetime(2026, 8, 18, 19, 0))
        assert should_prefer_wav_dates(s, None) is False

    def test_warning_only_on_mismatch(self):
        from chiro_core import SummaryInfo, summary_vs_wav_warning
        same = SummaryInfo(
            path=Path("x"),
            start_dt=datetime(2026, 8, 21, 20, 0),
            end_dt=datetime(2026, 8, 22, 6, 0),
        )
        assert summary_vs_wav_warning(same, datetime(2026, 8, 21, 20, 45)) is None
        cumulated = SummaryInfo(
            path=Path("x"),
            start_dt=datetime(2026, 8, 18, 19, 0),
            end_dt=datetime(2026, 8, 22, 6, 0),
        )
        msg = summary_vs_wav_warning(
            cumulated, datetime(2026, 8, 21, 20, 45),
            datetime(2026, 8, 22, 5, 10))
        assert msg is not None
        assert "WAV" in msg and "SD" in msg

    def test_temps_sliced_to_wav_window(self):
        from chiro_core import SummaryInfo, summary_temps_in_window
        s = SummaryInfo(
            path=Path("x"),
            samples=[
                (datetime(2026, 8, 18, 19, 0), 18.0),
                (datetime(2026, 8, 18, 23, 0), 16.0),
                (datetime(2026, 8, 21, 20, 50), 14.0),
                (datetime(2026, 8, 22, 5, 0), 11.0),
            ],
        )
        t0, t1 = summary_temps_in_window(
            s, datetime(2026, 8, 21, 20, 0), datetime(2026, 8, 22, 6, 0))
        assert t0 == 14.0 and t1 == 11.0

    def test_try_auto_meta_uses_wav_when_summary_spans(self, tmp_path):
        from rename import try_auto_meta
        sess = tmp_path / "jeanne"
        sess.mkdir()
        import wave
        wav = sess / "SMU03126_20260821_204529.wav"
        with wave.open(str(wav), "wb") as w:
            w.setnchannels(1)
            w.setsampwidth(2)
            w.setframerate(8000)
            w.writeframes(b"\x00\x00" * 10)
        (sess / "SMU03126_Summary.txt").write_text(
            "DATE,TIME,LAT,NS,LON,EW,POWER(V),TEMP(C),#FSFILES,#ZCFILES,#SCRUBBED\n"
            "2026-08-18,19:00:00,45.0,N,5.0,E,4.1,15.0,1,0,0\n"
            "2026-08-22,06:00:00,45.0,N,5.0,E,4.0,12.0,1,0,0\n",
            encoding="utf-8",
        )
        meta, msgs = try_auto_meta(sess)
        assert meta is not None
        assert meta.date_debut is not None
        assert meta.date_debut.date() == datetime(2026, 8, 21).date()
        assert any("WAV" in m or "multi" in m.lower() for m in msgs)

    def test_stale_participation_id(self):
        from pipeline import _participation_id_is_stale
        meta = {
            "vigiechiro_participation_id": "abc",
            "vigiechiro_participation_date": "2026-08-18",
        }
        assert _participation_id_is_stale(meta, datetime(2026, 8, 21, 20, 45)) is True
        assert _participation_id_is_stale(meta, datetime(2026, 8, 18, 19, 0)) is False

    def test_stale_from_payload_only(self):
        from pipeline import _participation_id_is_stale
        meta = {
            "vigiechiro_participation_id": "abc",
            "participation_payload": {"date_debut": "2026-08-18T19:00:00"},
        }
        assert _participation_id_is_stale(meta, datetime(2026, 8, 21, 20, 45)) is True

    def test_no_stored_date_not_stale(self):
        from pipeline import _participation_id_is_stale
        meta = {"vigiechiro_participation_id": "abc"}
        assert _participation_id_is_stale(meta, datetime(2026, 8, 21)) is False


class TestObservationSidecar:
    """Lot 1 — sidecar de synchro (donnee_id + index natif) écrit à l'export.
    L'xlsx reste inchangé ; le mapping serveur vit dans <stem>.sync.json."""

    def _fake_client(self, donnees):
        from vigiechiro_api import VigieChiroClient
        c = VigieChiroClient("A" * 32)

        def fake_iter(pid, on_progress=None, **kw):
            for d in donnees:
                yield d
        c.iter_donnees = fake_iter
        return c

    def test_sidecar_maps_native_index(self, tmp_path):
        from vigiechiro_api import load_observation_sidecar
        donnees = [
            {"_id": "d1", "titre": "Car-001_x.wav", "observations": [
                {"_id": "o0", "temps_debut": 1.0, "tadarida_taxon": "pippip"},
                {"_id": "o1", "temps_debut": 2.0, "tadarida_taxon": "barbar"},
                {"_id": "o2", "temps_debut": 3.0, "tadarida_taxon": "nyclei"},
            ]},
            {"_id": "d2", "titre": "Car-002_silent.wav", "observations": []},
        ]
        c = self._fake_client(donnees)
        dst = tmp_path / "participation-abc-observations.xlsx"
        stats = c.download_observations_as_xlsx("abc", dst)
        assert (stats["n_files"], stats["n_contacts"], stats["n_silent_files"]) == (2, 3, 1)
        entries = load_observation_sidecar(dst)["entries"]
        assert len(entries) == 3                       # rien pour la donnée silencieuse
        assert [e["obs_index"] for e in entries] == [0, 1, 2]
        assert [e["donnee_id"] for e in entries] == ["d1", "d1", "d1"]
        assert [e["row"] for e in entries] == [2, 3, 4]   # header=1 → contacts 2,3,4
        assert entries[0]["nom_fichier"] == "Car-001_x.wav"
        assert entries[0]["temps_debut"] == 1.0
        assert load_observation_sidecar(dst)["sync"] == {}

    def test_xlsx_unchanged_11_columns(self, tmp_path):
        import openpyxl
        donnees = [{"_id": "d1", "titre": "f.wav", "observations": [
            {"temps_debut": 1.0, "tadarida_taxon": "pippip"}]}]
        c = self._fake_client(donnees)
        dst = tmp_path / "participation-abc-observations.xlsx"
        c.download_observations_as_xlsx("abc", dst)
        ws = openpyxl.load_workbook(dst).active
        header = [cell.value for cell in next(ws.iter_rows())]
        assert len(header) == 11
        assert header[0] == "nom du fichier"
        assert header[7] == "observateur_taxon"
        assert header[10] == "validateur_probabilite"

    def test_sidecar_canonical_fallback(self, tmp_path):
        """Après sauvegarde en ..._KG.xlsx, le mapping reste résoluble."""
        from vigiechiro_api import load_observation_sidecar
        donnees = [{"_id": "d1", "titre": "f.wav", "observations": [{"temps_debut": 1.0}]}]
        c = self._fake_client(donnees)
        c.download_observations_as_xlsx("abc", tmp_path / "participation-abc-observations.xlsx")
        saved = tmp_path / "participation-abc-observations_KG.xlsx"
        side = load_observation_sidecar(saved)
        assert len(side["entries"]) == 1 and side["entries"][0]["donnee_id"] == "d1"

    def test_missing_sidecar_tolerated(self, tmp_path):
        from vigiechiro_api import load_observation_sidecar
        assert load_observation_sidecar(tmp_path / "nope.xlsx") == {"entries": [], "sync": {}}


class TestResolveTaxon:
    """Lot 2 — résolution libelle_court → ObjectId + cache + bypass 24-hex."""

    def _client(self, taxons, counter=None):
        from vigiechiro_api import VigieChiroClient
        c = VigieChiroClient("A" * 32)

        def fake_iter(page_size=99):
            if counter is not None:
                counter.append(1)
            for t in taxons:
                yield t
        c.iter_taxons = fake_iter
        return c

    def test_maps_case_insensitive(self):
        c = self._client([{"_id": "a" * 24, "libelle_court": "Barbar"}])
        assert c.resolve_taxon_id("barbar") == "a" * 24
        assert c.resolve_taxon_id("BARBAR") == "a" * 24

    def test_cache_single_fetch(self):
        calls = []
        c = self._client([{"_id": "a" * 24, "libelle_court": "pippip"}], counter=calls)
        c.resolve_taxon_id("pippip")
        c.resolve_taxon_id("pippip")
        assert len(calls) == 1                      # /taxons balayé une seule fois

    def test_unknown_raises(self):
        from vigiechiro_api import NotFoundError
        c = self._client([{"_id": "a" * 24, "libelle_court": "pippip"}])
        with pytest.raises(NotFoundError):
            c.resolve_taxon_id("zzzzzz")

    def test_bypass_objectid(self):
        calls = []
        c = self._client([], counter=calls)
        oid = "5aef012345678901234567ab"
        assert c.resolve_taxon_id(oid) == oid
        assert c.resolve_taxon_id(oid.upper()) == oid   # normalisé en minuscule
        assert len(calls) == 0                      # aucun appel /taxons pour un id direct

    def test_collision_keeps_first_and_warns(self):
        c = self._client([
            {"_id": "a" * 24, "libelle_court": "dupdup"},
            {"_id": "b" * 24, "libelle_court": "dupdup"},
        ])
        with pytest.warns(UserWarning):
            assert c.resolve_taxon_id("dupdup") == "a" * 24


class TestPushObservation:
    """Lot 3 — PATCH d'une observation + gardes + typage des erreurs 403/422."""

    def _client(self):
        from vigiechiro_api import VigieChiroClient
        return VigieChiroClient("A" * 32)

    def test_payload_and_params(self):
        c = self._client()
        captured = {}

        def fake_request(method, path, **kw):
            captured.update(method=method, path=path, **kw)
            return {"ok": True}
        c._request = fake_request
        oid = "a" * 24
        c.push_observation("d1", 2, oid, "SUR")
        assert captured["method"] == "PATCH"
        assert captured["path"] == "/donnees/d1/observations/2"
        assert captured["json"] == {"observateur_taxon": oid, "observateur_probabilite": "SUR"}
        assert "validateur_taxon" not in captured["json"]       # jamais de validateur_*
        assert "validateur_probabilite" not in captured["json"]
        assert captured["params"] == {"no_bilan": "true"}
        assert captured["source"] == "foreground"

    def test_no_bilan_false_triggers_bilan(self):
        c = self._client()
        captured = {}
        c._request = lambda method, path, **kw: captured.update(kw) or {}
        c.push_observation("d1", 0, "a" * 24, "POSSIBLE", no_bilan=False)
        assert captured["params"] is None                       # bilan déclenché

    def test_guard_bad_enum(self):
        c = self._client()
        c._request = lambda *a, **k: {}
        with pytest.raises(ValueError):
            c.push_observation("d1", 0, "a" * 24, "possible")   # minuscule refusée

    def test_guard_bad_id(self):
        c = self._client()
        c._request = lambda *a, **k: {}
        with pytest.raises(ValueError):
            c.push_observation("d1", 0, "barbar", "SUR")        # pas un ObjectId

    def test_request_classifies_403_and_422(self):
        from vigiechiro_api import ForbiddenError, ValidationError

        class FakeResp:
            def __init__(self, code):
                self.status_code, self.text, self.content, self.headers = code, "x", b"x", {}
        for code, exc in ((403, ForbiddenError), (422, ValidationError)):
            c = self._client()
            c.session.request = lambda *a, _c=code, **k: FakeResp(_c)
            with pytest.raises(exc):
                c._request("PATCH", "/donnees/d/observations/0", json={})


class TestSyncState:
    """Lot 4 — machine à états de synchro + mapping ligne↔serveur (pur)."""

    def test_is_sendable(self):
        from sync_state import is_sendable
        assert is_sendable("barbar", "SUR")
        assert not is_sendable("barbar", "")      # sans confiance → non envoyable
        assert not is_sendable("", "SUR")

    def test_sync_label(self):
        from sync_state import sync_label, SYNC_SYNCED, SYNC_TO_RETRACT
        assert sync_label(SYNC_SYNCED)             # libellé non vide (tooltip)
        assert sync_label(SYNC_TO_RETRACT)
        assert sync_label(None) is None            # pas d'état → pas d'infobulle
        assert sync_label("bogus") is None

    def test_pending_when_validated_never_pushed(self):
        from sync_state import next_sync_state, SYNC_PENDING
        assert next_sync_state("barbar", "SUR", None) == SYNC_PENDING
        assert next_sync_state("barbar", "SUR", {}) == SYNC_PENDING

    def test_none_when_not_sendable_never_pushed(self):
        from sync_state import next_sync_state
        assert next_sync_state("", "", None) is None
        assert next_sync_state("barbar", "", None) is None    # taxon sans confiance

    def test_synced_and_modified(self):
        from sync_state import next_sync_state, SYNC_SYNCED, SYNC_MODIFIED
        rec = {"pushed_taxon": "barbar", "pushed_conf": "SUR"}
        assert next_sync_state("barbar", "SUR", rec) == SYNC_SYNCED
        assert next_sync_state("pippip", "SUR", rec) == SYNC_MODIFIED
        assert next_sync_state("barbar", "PROBABLE", rec) == SYNC_MODIFIED

    def test_modified_back_to_synced(self):
        from sync_state import next_sync_state, SYNC_SYNCED
        rec = {"pushed_taxon": "barbar", "pushed_conf": "SUR", "state": "modified"}
        assert next_sync_state("barbar", "SUR", rec) == SYNC_SYNCED

    def test_to_retract_when_cleared_after_push(self):
        from sync_state import next_sync_state, SYNC_TO_RETRACT
        rec = {"pushed_taxon": "barbar", "pushed_conf": "SUR"}
        assert next_sync_state("", "", rec) == SYNC_TO_RETRACT
        assert next_sync_state("barbar", "", rec) == SYNC_TO_RETRACT   # non-envoyable

    def test_build_row_key_map(self):
        from sync_state import build_row_key_map
        col_idx = {"nom du fichier": 0, "temps_debut": 1}
        rows = [["f1.wav", 1.0], ["f1.wav", 2.0], ["f2.wav", 0.5]]
        entries = [
            {"donnee_id": "d1", "obs_index": 0, "nom_fichier": "f1.wav", "temps_debut": 1.0},
            {"donnee_id": "d1", "obs_index": 1, "nom_fichier": "f1.wav", "temps_debut": 2.0},
        ]
        row_keys, n_unmapped = build_row_key_map(rows, col_idx, entries)
        assert row_keys == {0: "d1#0", 1: "d1#1"}
        assert n_unmapped == 1                     # f2.wav non présent dans entries

    def test_build_row_key_map_no_entries(self):
        from sync_state import build_row_key_map
        rk, n = build_row_key_map([["f.wav", 1.0]],
                                  {"nom du fichier": 0, "temps_debut": 1}, [])
        assert rk == {} and n == 1

    def test_build_row_key_map_collision_excluded(self):
        """Sécurité anti-corruption : 2 obs partageant (nom, deb, fin) → EXCLUES
        (non mappées) au lieu d'écrire toutes deux sur l'index 0 côté serveur."""
        from sync_state import build_row_key_map
        col_idx = {"nom du fichier": 0, "temps_debut": 1, "temps_fin": 2}
        rows = [["f.wav", 1.0, 2.0], ["f.wav", 1.0, 2.0], ["g.wav", 5.0, 6.0]]
        entries = [
            {"donnee_id": "d1", "obs_index": 0, "nom_fichier": "f.wav", "temps_debut": 1.0, "temps_fin": 2.0},
            {"donnee_id": "d1", "obs_index": 1, "nom_fichier": "f.wav", "temps_debut": 1.0, "temps_fin": 2.0},
            {"donnee_id": "d2", "obs_index": 0, "nom_fichier": "g.wav", "temps_debut": 5.0, "temps_fin": 6.0},
        ]
        row_keys, n_unmapped = build_row_key_map(rows, col_idx, entries)
        assert row_keys == {2: "d2#0"}       # seule la ligne non ambiguë est mappée
        assert n_unmapped == 2

    def test_build_row_key_map_temps_fin_disambiguates(self):
        """Même temps_debut mais temps_fin différent → restent appariables."""
        from sync_state import build_row_key_map
        col_idx = {"nom du fichier": 0, "temps_debut": 1, "temps_fin": 2}
        rows = [["f.wav", 1.0, 2.0], ["f.wav", 1.0, 3.0]]
        entries = [
            {"donnee_id": "d1", "obs_index": 0, "nom_fichier": "f.wav", "temps_debut": 1.0, "temps_fin": 2.0},
            {"donnee_id": "d1", "obs_index": 1, "nom_fichier": "f.wav", "temps_debut": 1.0, "temps_fin": 3.0},
        ]
        row_keys, n_unmapped = build_row_key_map(rows, col_idx, entries)
        assert row_keys == {0: "d1#0", 1: "d1#1"} and n_unmapped == 0


class TestRegistryRollup:
    """Lot 6 — rollup envoi identifications (schéma v3), sans perte pour l'existant."""

    def _reg(self, tmp_path):
        from registry import Registry
        return Registry(tmp_path)

    def _add(self, reg, tmp_path, sid="s1"):
        reg.upsert_session({"id": sid, "canonical_name": sid,
                            "session_path": str(tmp_path / sid)})
        return sid

    def test_columns_default_zero_and_writable(self, tmp_path):
        reg = self._reg(tmp_path)
        sid = self._add(reg, tmp_path)
        row = reg.get_session(sid)
        assert row["ident_pushed"] == 0 and row["ident_total"] == 0
        reg.update_fields(sid, {"ident_pushed": 3, "ident_total": 5})
        row = reg.get_session(sid)
        assert row["ident_pushed"] == 3 and row["ident_total"] == 5

    def test_rescan_does_not_reset_rollup(self, tmp_path):
        reg = self._reg(tmp_path)
        sid = self._add(reg, tmp_path)
        reg.update_fields(sid, {"ident_pushed": 4, "ident_total": 6})
        self._add(reg, tmp_path)                    # re-scan (upsert existant)
        row = reg.get_session(sid)
        assert row["ident_pushed"] == 4 and row["ident_total"] == 6

    def test_migration_idempotent_preserves_data(self, tmp_path):
        from registry import Registry
        reg = self._reg(tmp_path)
        self._add(reg, tmp_path)
        reg2 = Registry(tmp_path)                   # ré-ouverture → migration idempotente
        cols = {r[1] for r in reg2._get_conn().execute("PRAGMA table_info(sessions)")}
        assert {"ident_pushed", "ident_total"} <= cols
        assert reg2.get_session("s1") is not None    # données préservées


class TestActivityReference:
    """Référentiel d'activité Vigie-Chiro (Bas et al. 2020) — combiné + repli."""

    def test_load_real_reference(self):
        from activity_reference import load_reference
        ref = load_reference()
        row = ref[("national", "toutes", "pippip")]   # clé (referentiel, saison, code)
        assert row["q25"] == 13 and row["q98"] == 3737

    def test_classify_boundaries(self):
        from activity_reference import classify
        row = {"q25": 13, "q75": 411, "q98": 3737}
        assert classify(5, row) == "Faible"
        assert classify(13, row) == "Moyenne"        # ≥ Q25
        assert classify(411, row) == "Forte"         # ≥ Q75
        assert classify(3737, row) == "Très forte"   # ≥ Q98

    def test_region_from_dept_and_site(self):
        from activity_reference import region_for_dept, region_for_site
        assert region_for_dept("34") == "Occitanie"
        assert region_for_dept("2A") == "Corse"
        assert region_for_dept("1") == "Auvergne-Rhone-Alpes"   # zero-pad
        assert region_for_site("340123") == "Occitanie"          # 2 premiers = dept
        assert region_for_dept("999") is None

    def test_season_from_date(self):
        from datetime import datetime
        from activity_reference import season_for_date
        assert season_for_date(datetime(2026, 5, 1)) == "printemps"
        assert season_for_date(datetime(2026, 7, 15)) == "ete"
        assert season_for_date(datetime(2026, 10, 1)) == "automne"
        assert season_for_date(datetime(2026, 12, 25)) == "toutes"   # hors fenêtres
        assert season_for_date(None) == "toutes"

    def test_activity_for_real_region(self):
        from activity_reference import load_reference, activity_for
        ref = load_reference()
        a = activity_for("Pippip", 100, ref, saison="automne", region="Occitanie")
        assert a["referentiel"] == "region:Occitanie"   # Occitanie automne = fiable
        assert a["classe"] == "Forte"                    # Q75=87 ≤ 100 < Q98=1999
        assert activity_for("zzzzzz", 10, ref) is None   # code hors référentiel

    def test_fallback_prefers_reliable_national_over_weak_region(self):
        from activity_reference import activity_for
        # région peu fiable ('Faible') vs national fiable : on garde le national.
        ref = {
            ("region:Corse", "automne", "barbar"):
                {"q25": 1, "q75": 5, "q98": 50, "nbocc": 10, "confiance": "Faible"},
            ("national", "automne", "barbar"):
                {"q25": 2, "q75": 16, "q98": 181, "nbocc": 4000, "confiance": "Tres bonne"},
        }
        a = activity_for("barbar", 20, ref, saison="automne", region="Corse")
        assert a["referentiel"] == "national" and a["fiable"] is True

    def test_annotate_synthesis_with_context(self):
        from activity_reference import load_reference, annotate_synthesis
        ref = load_reference()
        synth = {"species": [
            {"taxon": "pippip", "groupe": "chiros", "n_contacts": 5000},
            {"taxon": "noise", "groupe": "noise", "n_contacts": 999},
        ]}
        annotate_synthesis(synth, ref, saison="ete", region="Occitanie")
        assert synth["species"][0]["activite"]["classe"] in (
            "Faible", "Moyenne", "Forte", "Très forte")
        assert synth["species"][1]["activite"] is None   # non-chiro

    def test_missing_reference_degrades(self, tmp_path):
        from activity_reference import load_reference
        assert load_reference(tmp_path / "nope.csv") == {}


class TestSynthesisValidatedTotal:
    """Enrichissement synthèse : contacts validés vs total + richesse."""

    def test_validated_vs_total_and_richesse(self):
        from synthesis import compute_night_synthesis
        headers = ["nom du fichier", "tadarida_taxon", "observateur_taxon"]
        rows = [
            ["f1.wav", "pippip", "pippip"],   # validé
            ["f2.wav", "pippip", ""],          # non validé (Tadarida seul)
            ["f3.wav", "barbar", "barbar"],    # validé
            ["f4.wav", "noise", ""],           # bruit
        ]
        s = compute_night_synthesis(headers, rows)
        assert s["total_contacts"] == 4
        assert s["validated_contacts"] == 2       # f1 + f3
        assert s["richesse_chiros"] == 2           # pippip + barbar
        assert s["richesse_totale"] == 2           # noise exclu

    def test_validated_only_filter(self):
        from synthesis import compute_night_synthesis
        headers = ["nom du fichier", "tadarida_taxon", "observateur_taxon"]
        rows = [
            ["f1.wav", "pippip", "pippip"],   # validé
            ["f2.wav", "pippip", ""],          # Tadarida seul → exclu si validated_only
            ["f3.wav", "barbar", "barbar"],    # validé
        ]
        s = compute_night_synthesis(headers, rows, validated_only=True)
        assert s["total_contacts"] == 2            # seulement f1 + f3
        assert {sp["taxon"] for sp in s["species"]} == {"pippip", "barbar"}
        assert next(sp for sp in s["species"] if sp["taxon"] == "pippip")["n_contacts"] == 1


class TestEntriesFromDonnees:
    """Reconstruction du mapping serveur (fallback xlsx sans sidecar)."""

    def test_entries_from_donnees(self):
        from vigiechiro_api import entries_from_donnees
        donnees = [
            {"_id": "d1", "titre": "f1.wav", "observations": [
                {"_id": "o0", "temps_debut": 1.0, "temps_fin": 2.0},
                {"_id": "o1", "temps_debut": 3.0, "temps_fin": 4.0}]},
            {"_id": "d2", "titre": "f2.wav", "observations": []},
        ]
        e = entries_from_donnees(donnees)
        assert len(e) == 2
        assert e[0] == {"donnee_id": "d1", "obs_index": 0, "obs_id": "o0",
                        "nom_fichier": "f1.wav", "temps_debut": 1.0, "temps_fin": 2.0}
        assert e[1]["obs_index"] == 1

    def test_entries_feed_row_key_map(self):
        # bout-en-bout : les entries reconstruites doivent mapper les lignes xlsx
        from vigiechiro_api import entries_from_donnees
        from sync_state import build_row_key_map
        donnees = [{"_id": "d1", "titre": "f1.wav", "observations": [
            {"temps_debut": 1.0, "temps_fin": 2.0}]}]
        entries = entries_from_donnees(donnees)
        rows = [["f1.wav", 1.0, 2.0]]
        col_idx = {"nom du fichier": 0, "temps_debut": 1, "temps_fin": 2}
        rk, n = build_row_key_map(rows, col_idx, entries)
        assert rk == {0: "d1#0"} and n == 0


class TestTaxonIndex:
    """Index /taxons Vigie-Chiro : validation, autocomplétion, code genre."""

    def test_load_and_is_known(self):
        from taxon_index import load_taxon_index, is_known
        idx = load_taxon_index()
        assert is_known("Pleaur", idx) and is_known("plesp", idx)   # casse-insensible
        assert not is_known("zzzzz", idx)
        assert not is_known("Pip35", idx)   # code Tadarida non accepté par le serveur

    def test_canonical_code(self):
        from taxon_index import load_taxon_index, canonical_code
        idx = load_taxon_index()
        assert canonical_code("pleaur", idx) == "Pleaur"
        assert canonical_code("inconnu", idx) is None

    def test_suggest_surfaces_species_and_genus(self):
        from taxon_index import load_taxon_index, suggest
        idx = load_taxon_index()
        codes = {c for c, _ in suggest("orei", idx)}     # recherche par libellé FR
        assert {"Pleaur", "Pleaus", "Plesp"} <= codes    # espèces + genre incertain
        assert suggest("", idx) == []

    def test_genus_code_for_uncertain(self):
        from taxon_index import load_taxon_index, genus_code_for
        idx = load_taxon_index()
        assert genus_code_for("Pleaur", idx) == "Plesp"  # Oreillard indéterminé
        assert genus_code_for("Pleaus", idx) == "Plesp"
        assert genus_code_for("Pippip", idx) == "Pipsp"
        assert genus_code_for("Myodau", idx) == "Myosp"
        assert genus_code_for("Plesp", idx) is None      # déjà un code genre
        assert genus_code_for("Barbar", idx) is None     # genre monospécifique (pas de Barsp)

    def test_missing_snapshot_degrades(self, tmp_path):
        from taxon_index import load_taxon_index
        assert load_taxon_index(tmp_path / "nope.csv") == {}


class TestTokenPrompt:
    """H2 — saisie du token sans l'exposer dans l'historique shell."""

    def test_env_variable(self, monkeypatch):
        from credentials import prompt_token, ENV_TOKEN
        monkeypatch.setenv(ENV_TOKEN, "  ENVTOKEN123  ")
        assert prompt_token() == "ENVTOKEN123"        # trim + priorité env

    def test_arg_is_deprecated_but_works(self, monkeypatch):
        from credentials import prompt_token, ENV_TOKEN
        monkeypatch.delenv(ENV_TOKEN, raising=False)
        with pytest.warns(UserWarning):               # avertit de l'exposition
            assert prompt_token("ARGTOKEN") == "ARGTOKEN"

    def test_interactive_masked_fallback(self, monkeypatch):
        from credentials import prompt_token, ENV_TOKEN
        monkeypatch.delenv(ENV_TOKEN, raising=False)
        # pas d'arg, pas d'env → saisie masquée (injectée pour le test)
        assert prompt_token(_input=lambda: "TYPEDTOKEN") == "TYPEDTOKEN"


class TestResourcePath:
    """resource_path : les CSV embarqués sont trouvés (dev + exe PyInstaller)."""

    def test_bundled_csvs_resolve(self):
        from resources import resource_path
        for name in ("SpeciesListComplete.csv", "activity_ref_PF.csv",
                     "taxons_vigiechiro.csv"):
            assert resource_path(name).is_file(), name

    def test_reference_loaders_still_work(self):
        # les loaders passent maintenant par resource_path → doivent charger
        from activity_reference import load_reference
        from taxon_index import load_taxon_index
        assert load_reference()          # non vide
        assert load_taxon_index()


class TestCleanupMassDeleteGuard:
    """Chemin destructif : garde-fou mass-delete (préserve les WAV par défaut)."""

    _HEADERS = ["nom du fichier", "temps_debut", "temps_fin", "frequence_mediane",
                "tadarida_taxon", "tadarida_probabilite", "tadarida_taxon_autre",
                "observateur_taxon", "observateur_probabilite",
                "validateur_taxon", "validateur_probabilite"]

    def _session(self, tmp_path, name, filenames, taxon, proba):
        import openpyxl
        session = tmp_path / name
        data_k = session / "Data_k"
        data_k.mkdir(parents=True)
        for fn in filenames:
            (data_k / fn).write_bytes(b"RIFFxxxxWAVE")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(self._HEADERS)
        for fn in filenames:
            ws.append([fn, 1.0, 2.0, 45.0, taxon, proba, "", "", "", "", ""])
        wb.save(session / "participation-abc-observations.xlsx")
        return session, data_k

    _THR = {"chiros": 0.5, "orthos": 0.5, "micromam": 0.5, "oiseaux": 0.5}

    def test_blocked_by_default_preserves_wavs(self, tmp_path):
        from cleanup import cleanup_session
        names = [f"f{i}.wav" for i in range(6)]
        session, data_k = self._session(tmp_path, "20250903_site212097_Z3_Pass2_enr07",
                                         names, "pippip", 0.10)   # chiros sous le seuil
        out = cleanup_session(session, thresholds=self._THR, disabled=set(),
                              silent_policy="delete", dry_run=False,
                              allow_mass_delete=False)
        assert out.get("errors")                      # suppression refusée
        assert out.get("n_deleted") == 0
        assert len(list(data_k.glob("*.wav"))) == 6   # WAV intacts

    def test_allowed_executes_deletion(self, tmp_path):
        from cleanup import cleanup_session
        names = [f"g{i}.wav" for i in range(6)]
        session, data_k = self._session(tmp_path, "20250903_site212097_Z3_Pass2_enr08",
                                         names, "pippip", 0.10)
        out = cleanup_session(session, thresholds=self._THR, disabled=set(),
                              silent_policy="delete", dry_run=False,
                              allow_mass_delete=True)   # override explicite
        assert out.get("n_deleted", 0) >= 1
        assert len(list(data_k.glob("*.wav"))) < 6


class TestFilterItems:
    """Recherche des filtres graphes : insensible à la casse ET aux accents."""

    def test_empty_query_returns_all(self):
        from activity_graph import filter_items
        items = ["Pippip", "Barbar", "Nyclei"]
        assert filter_items(items, "") == items
        assert filter_items(items, "   ") == items

    def test_case_and_accent_insensitive(self):
        from activity_graph import filter_items
        items = ["Pipistrelle", "Sérotine", "Noctule"]
        assert filter_items(items, "pip") == ["Pipistrelle"]
        assert filter_items(items, "sero") == ["Sérotine"]   # requête sans accent
        assert filter_items(items, "SÉRO") == ["Sérotine"]   # requête accentuée MAJ

    def test_label_fn(self):
        from activity_graph import filter_items
        assert filter_items(["212097", "741131"], "#7",
                            label_fn=lambda s: f"#{s}") == ["741131"]

    def test_no_match(self):
        from activity_graph import filter_items
        assert filter_items(["a", "b"], "zzz") == []


class TestFolderRenameNonFatal:
    """Issue #2 — le renommage du DOSSIER échoue (handle Windows / antivirus) mais
    les WAV sont renommés : ça ne doit PLUS bloquer la nuit, juste avertir."""

    def test_folder_rename_failure_is_non_fatal(self, tmp_path, monkeypatch):
        from datetime import datetime
        from pathlib import Path
        from naming import SessionMeta
        import rename as rename_mod
        from rename import rename_session

        sess = tmp_path / "20260519_raw"
        sess.mkdir()
        for t in ("203704", "203711"):
            (sess / f"SMU05451_20260519_{t}.wav").write_bytes(b"RIFFxxxxWAVE")
        meta = SessionMeta(date_debut=datetime(2026, 5, 19), n_site_tadarida="212097",
                           n_point_fixe="Z6", n_passage=2, n_enregistreur=7,
                           n_serie="SMU05451", nom_contrat="T")

        real_rename = Path.rename

        def fake_rename(self, target):
            if Path(self) == sess:               # échoue UNIQUEMENT sur le dossier
                raise OSError(32, "handle ouvert (simulé)")
            return real_rename(self, target)
        monkeypatch.setattr(Path, "rename", fake_rename)
        monkeypatch.setattr(rename_mod.time, "sleep", lambda *a: None)  # pas d'attente

        out = rename_session(sess, meta, dry_run=False, rename_folder=True)

        assert out["executed"] == 2                    # les 2 WAV sont renommés
        assert not out.get("errors")                    # échec dossier = NON fatal
        assert out.get("folder_rename_failed")          # mais signalé
        assert any("DOSSIER" in w for w in out.get("warnings", []))
        assert out["final_session_path"] == str(sess)   # on garde le chemin réel
        assert list(sess.glob("Car212097-2026-Pass2-Z6-SMU05451_*.wav"))  # WAV canoniques


# =========================================================================
# repair — diagnostic / alignement d'état (une nuit)
# =========================================================================

class TestFichiersListingPagination:
    def test_stop_on_empty_page(self):
        from vigiechiro_api import fichiers_listing_page_done
        assert fichiers_listing_page_done(
            n_page_items=0, n_names_so_far=0, meta_total=None) is True

    def test_continue_when_total_missing_and_full_page(self):
        """Régression : total=0/absent ne doit PAS couper après page 1."""
        from vigiechiro_api import fichiers_listing_page_done
        assert fichiers_listing_page_done(
            n_page_items=99, n_names_so_far=99, meta_total=0,
            page_size=99) is False
        assert fichiers_listing_page_done(
            n_page_items=99, n_names_so_far=99, meta_total=None,
            page_size=99) is False

    def test_stop_when_total_reached(self):
        from vigiechiro_api import fichiers_listing_page_done
        assert fichiers_listing_page_done(
            n_page_items=50, n_names_so_far=250, meta_total=250,
            page_size=99) is True

    def test_stop_on_short_page(self):
        from vigiechiro_api import fichiers_listing_page_done
        assert fichiers_listing_page_done(
            n_page_items=37, n_names_so_far=137, meta_total=None,
            page_size=99) is True


class TestRepairCoveragePure:
    """Helpers purs : compute_coverage + suggest_actions."""

    def test_coverage_full_match(self):
        from repair import compute_coverage
        cov = compute_coverage(["a.wav", "b.wav"], ["b.wav", "a.wav"])
        assert cov["coverage_ok"] is True
        assert cov["missing_on_server"] == []
        assert cov["extra_on_server"] == []
        assert cov["local_wav_count"] == 2
        assert cov["server_wav_count"] == 2

    def test_coverage_partial_missing(self):
        from repair import compute_coverage
        cov = compute_coverage(["a.wav", "b.wav", "c.wav"], ["a.wav"])
        assert cov["coverage_ok"] is False
        assert cov["missing_on_server"] == ["b.wav", "c.wav"]
        assert cov["extra_on_server"] == []

    def test_coverage_extra_on_server_still_ok(self):
        """Des fichiers en trop côté serveur n'empêchent pas coverage_ok.

        Cas typique post-nettoyage : des centaines de WAV encore en ligne
        mais purgés de Data_k — ce ne sont PAS des manquants d'upload.
        """
        from repair import compute_coverage
        cov = compute_coverage(["a.wav"], ["a.wav", "ghost.wav", "noise.wav"])
        assert cov["coverage_ok"] is True
        assert set(cov["extra_on_server"]) == {"ghost.wav", "noise.wav"}
        assert cov["missing_on_server"] == []

    def test_coverage_name_normalization(self):
        """Casse / extension manquante côté API ne doivent pas créer de faux manquants."""
        from repair import compute_coverage
        cov = compute_coverage(
            ["Car123.WAV", "b.wav"],
            ["car123.wav", "B"],  # B sans .wav
        )
        assert cov["coverage_ok"] is True
        assert cov["missing_on_server"] == []

    def test_coverage_listing_failed(self):
        from repair import compute_coverage
        cov = compute_coverage(["a.wav", "b.wav"], [], listing_ok=False)
        assert cov["coverage_ok"] is False
        assert cov["listing_ok"] is False
        # Pas de faux « à uploader » quand le listing a échoué
        assert cov["missing_on_server"] == []

    def test_format_report_highlights_token_401(self):
        from repair import format_repair_report
        text = format_repair_report({
            "session": "s",
            "participation_id": "p",
            "local_wav_count": 10,
            "server_wav_count": 0,
            "coverage_ok": False,
            "listing_ok": False,
            "listing_error": "token Vigie-Chiro invalide ou expiré (HTTP 401)",
            "missing_on_server": [],
            "extra_on_server": [],
            "local_flags": {},
            "suggested_actions": ["noop"],
            "errors": ["list_participation_files : token … 401"],
            "notes": [],
        })
        assert "non comparable" in text or "ÉCHEC listing" in text
        assert "À uploader" not in text
        assert "Préférences" in text or "token" in text.lower()

    def test_coverage_empty_local(self):
        from repair import compute_coverage
        cov = compute_coverage([], [])
        assert cov["coverage_ok"] is False

    def test_suggest_partial_resume_no_trigger(self):
        from repair import (
            suggest_actions, ACTION_RESUME_UPLOAD, ACTION_TRIGGER,
            ACTION_SET_UPLOADED,
        )
        acts = suggest_actions(
            coverage_ok=False,
            listing_ok=True,
            missing_on_server=["b.wav"],
            traitement_etat=None,
            has_xlsx=False,
            flag_uploaded=False,
            has_participation_id=True,
        )
        assert ACTION_RESUME_UPLOAD in acts
        assert ACTION_TRIGGER not in acts
        assert ACTION_SET_UPLOADED not in acts

    def test_suggest_full_coverage_set_uploaded_and_trigger(self):
        from repair import (
            suggest_actions, ACTION_SET_UPLOADED, ACTION_TRIGGER, ACTION_FETCH,
        )
        acts = suggest_actions(
            coverage_ok=True,
            listing_ok=True,
            missing_on_server=[],
            traitement_etat="",
            has_xlsx=False,
            flag_uploaded=False,
            has_participation_id=True,
        )
        assert ACTION_SET_UPLOADED in acts
        assert ACTION_TRIGGER in acts
        assert ACTION_FETCH not in acts

    def test_suggest_termine_sans_xlsx(self):
        from repair import suggest_actions, ACTION_FETCH, ACTION_TRIGGER
        acts = suggest_actions(
            coverage_ok=True,
            listing_ok=True,
            missing_on_server=[],
            traitement_etat="TERMINE",
            has_xlsx=False,
            flag_uploaded=True,
            has_participation_id=True,
        )
        assert ACTION_FETCH in acts
        assert ACTION_TRIGGER not in acts  # déjà terminé

    def test_suggest_en_cours_no_trigger(self):
        from repair import suggest_actions, ACTION_TRIGGER, ACTION_SET_UPLOADED
        acts = suggest_actions(
            coverage_ok=True,
            listing_ok=True,
            missing_on_server=[],
            traitement_etat="EN_COURS",
            has_xlsx=False,
            flag_uploaded=False,
            has_participation_id=True,
        )
        assert ACTION_SET_UPLOADED in acts
        assert ACTION_TRIGGER not in acts

    def test_suggest_noop_when_aligned(self):
        from repair import suggest_actions, ACTION_NOOP
        acts = suggest_actions(
            coverage_ok=True,
            listing_ok=True,
            missing_on_server=[],
            traitement_etat="TERMINE",
            has_xlsx=True,
            flag_uploaded=True,
            has_participation_id=True,
        )
        assert acts == [ACTION_NOOP]

    def test_suggest_no_participation(self):
        from repair import suggest_actions, ACTION_NOOP
        acts = suggest_actions(
            coverage_ok=False,
            listing_ok=False,
            missing_on_server=[],
            traitement_etat=None,
            has_xlsx=False,
            flag_uploaded=False,
            has_participation_id=False,
        )
        assert acts == [ACTION_NOOP]


class TestRepairLocalFs:
    def test_list_data_k_and_xlsx(self, tmp_path):
        from repair import list_local_data_k_wavs, find_local_observations_xlsx
        session = tmp_path / "sess"
        dk = session / "Data_k"
        dk.mkdir(parents=True)
        (dk / "a.wav").write_bytes(b"x")
        (dk / "b.WAV").write_bytes(b"x")
        (dk / "note.txt").write_text("nope")
        (session / "participation-abc-observations.xlsx").write_bytes(b"PK")
        (session / "participation-abc-observations_cleanup.xlsx").write_bytes(b"PK")

        names = list_local_data_k_wavs(session)
        assert names == ["a.wav", "b.WAV"] or set(names) == {"a.wav", "b.WAV"}
        xlsx = find_local_observations_xlsx(session)
        assert xlsx is not None
        assert "_cleanup" not in xlsx.name.lower()

    def test_format_repair_report_readable(self):
        from repair import format_repair_report
        text = format_repair_report({
            "session": "/tmp/s",
            "participation_id": "abc",
            "local_wav_count": 2,
            "server_wav_count": 1,
            "missing_on_server": ["b.wav"],
            "extra_on_server": [],
            "coverage_ok": False,
            "listing_ok": True,
            "traitement_etat": "EN_COURS",
            "has_xlsx": False,
            "local_flags": {"uploaded": False},
            "suggested_actions": ["resume_upload_missing"],
            "dry_run": True,
        })
        assert "Diagnostic" in text
        assert "b.wav" in text
        assert "EN_COURS" in text
        assert "dry-run" in text


class _FakeRepairClient:
    """Client minimal injecté dans diagnose_and_repair_session."""

    def __init__(self, *, etat="TERMINE", files=None, fail_list=False,
                 fail_status=False, fail_trigger=False, fail_fetch=False):
        self.etat = etat
        self.files = list(files if files is not None else [])
        self.fail_list = fail_list
        self.fail_status = fail_status
        self.fail_trigger = fail_trigger
        self.fail_fetch = fail_fetch
        self.trigger_calls = 0
        self.fetch_calls = 0

    def participation_status(self, participation_id: str) -> dict:
        if self.fail_status:
            raise RuntimeError("status boom")
        return {"id": participation_id, "etat": self.etat, "date": None, "has_bilan": False}

    def list_participation_files(self, participation_id: str) -> list[str]:
        if self.fail_list:
            raise RuntimeError("list boom")
        return list(self.files)

    def trigger_compute(self, participation_id: str) -> dict:
        self.trigger_calls += 1
        if self.fail_trigger:
            raise RuntimeError("trigger boom")
        return {"ok": True}

    def download_observations_as_xlsx(self, participation_id, dst, on_progress=None):
        self.fetch_calls += 1
        if self.fail_fetch:
            raise RuntimeError("fetch boom")
        from pathlib import Path
        dst = Path(dst)
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_bytes(b"fake-xlsx")
        if on_progress:
            on_progress(1, 1)
        return {"n_files": 1, "n_contacts": 0, "path": str(dst)}


def _session_with_manifest(tmp_path, *, wavs, part_id="pid123", flags=None, xlsx=False):
    from manifest import Manifest
    session = tmp_path / "20250903_site212097_Z3_Pass2_enr07"
    session.mkdir()
    dk = session / "Data_k"
    dk.mkdir()
    for w in wavs:
        (dk / w).write_bytes(b"RIFF")
    m = Manifest.load_or_create(session)
    m.set_meta(vigiechiro_participation_id=part_id)
    if flags:
        for k, v in flags.items():
            m.flags[k] = v
    m.save(session)
    if xlsx:
        (session / f"participation-{part_id}-observations.xlsx").write_bytes(b"PK")
    return session


class TestDiagnoseAndRepairSession:
    def test_empty_listing_with_xlsx_no_mass_reupload(self, tmp_path):
        """Listing 0 + xlsx/cleaned → ne pas proposer re-upload de tout Data_k."""
        from repair import diagnose_and_repair_session, ACTION_RESUME_UPLOAD
        wavs = [f"f{i}.wav" for i in range(25)]
        session = _session_with_manifest(
            tmp_path, wavs=wavs,
            flags={"uploaded": True, "cleaned": True},
            xlsx=True,
        )
        client = _FakeRepairClient(etat="TERMINE", files=[])  # listing vide
        report = diagnose_and_repair_session(
            session, token=None, apply=False, client=client,
        )
        assert report["listing_ok"] is False
        assert ACTION_RESUME_UPLOAD not in report["suggested_actions"]
        assert any("re-uploader" in n or "listing" in n.lower()
                   for n in (report.get("notes") or []))

    def test_dry_run_no_file_modification(self, tmp_path):
        from repair import diagnose_and_repair_session
        from manifest import Manifest
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav", "b.wav"], flags={"uploaded": False},
        )
        mtime_before = (session / "_session_manifest.json").stat().st_mtime_ns
        client = _FakeRepairClient(etat="", files=["a.wav", "b.wav"])

        report = diagnose_and_repair_session(
            session, token=None, apply=False, client=client,
        )

        assert report["dry_run"] is True
        assert report["coverage_ok"] is True
        assert "set_uploaded_true" in report["suggested_actions"]
        assert "trigger_compute" in report["suggested_actions"]
        assert report["applied_actions"] == []
        assert client.trigger_calls == 0
        assert client.fetch_calls == 0
        # Manifest inchangé
        mtime_after = (session / "_session_manifest.json").stat().st_mtime_ns
        assert mtime_after == mtime_before
        m = Manifest.load(session)
        assert m.flags.get("uploaded") is False

    def test_partial_coverage_no_set_uploaded_no_trigger(self, tmp_path):
        from repair import diagnose_and_repair_session
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav", "b.wav", "c.wav"], flags={"uploaded": False},
        )
        client = _FakeRepairClient(etat="", files=["a.wav"])  # b,c manquants

        report = diagnose_and_repair_session(
            session, apply=True, client=client,
            allow_trigger=True, confirm_trigger=True,
        )

        assert report["coverage_ok"] is False
        assert report["missing_on_server"] == ["b.wav", "c.wav"]
        assert "resume_upload_missing" in report["suggested_actions"]
        assert "set_uploaded_true" not in report["suggested_actions"]
        assert "trigger_compute" not in report["suggested_actions"]
        assert client.trigger_calls == 0
        assert "set_uploaded_true" not in report["applied_actions"]
        # resume suggéré mais non exécuté
        assert any(
            s["action"] == "resume_upload_missing" for s in report["skipped_actions"]
        )

    def test_apply_set_uploaded_on_full_coverage(self, tmp_path):
        from repair import diagnose_and_repair_session
        from manifest import Manifest
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav"], flags={"uploaded": False},
        )
        client = _FakeRepairClient(etat="EN_COURS", files=["a.wav"])

        report = diagnose_and_repair_session(session, apply=True, client=client)

        assert "set_uploaded_true" in report["applied_actions"]
        assert client.trigger_calls == 0  # EN_COURS → pas de trigger
        m = Manifest.load(session)
        assert m.flags.get("uploaded") is True
        types = [a.type for a in m.actions]
        assert "repair" in types

    def test_trigger_requires_double_confirmation(self, tmp_path):
        from repair import diagnose_and_repair_session
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav"], flags={"uploaded": True},
        )
        client = _FakeRepairClient(etat="", files=["a.wav"])

        r1 = diagnose_and_repair_session(
            session, apply=True, client=client, allow_trigger=True,
            confirm_trigger=False,
        )
        assert client.trigger_calls == 0
        assert any(s["action"] == "trigger_compute" for s in r1["skipped_actions"])

        r2 = diagnose_and_repair_session(
            session, apply=True, client=client, allow_trigger=False,
            confirm_trigger=True,
        )
        assert client.trigger_calls == 0

        r3 = diagnose_and_repair_session(
            session, apply=True, client=client, allow_trigger=True,
            confirm_trigger=True,
        )
        assert client.trigger_calls == 1
        assert "trigger_compute" in r3["applied_actions"]

    def test_termine_sans_xlsx_fetch(self, tmp_path):
        from repair import diagnose_and_repair_session
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav"], flags={"uploaded": True}, xlsx=False,
        )
        client = _FakeRepairClient(etat="TERMINE", files=["a.wav"])

        report = diagnose_and_repair_session(
            session, apply=True, client=client, allow_fetch=True,
        )

        assert "fetch_xlsx" in report["suggested_actions"]
        assert "fetch_xlsx" in report["applied_actions"]
        assert client.fetch_calls == 1
        assert report["has_xlsx"] is True
        xlsx = session / "participation-pid123-observations.xlsx"
        assert xlsx.is_file()

    def test_fetch_disabled(self, tmp_path):
        from repair import diagnose_and_repair_session
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav"], flags={"uploaded": True},
        )
        client = _FakeRepairClient(etat="TERMINE", files=["a.wav"])

        report = diagnose_and_repair_session(
            session, apply=True, client=client, allow_fetch=False,
        )
        assert client.fetch_calls == 0
        assert any(s["action"] == "fetch_xlsx" for s in report["skipped_actions"])

    def test_missing_participation_id(self, tmp_path):
        from repair import diagnose_and_repair_session
        from manifest import Manifest
        session = tmp_path / "sess"
        session.mkdir()
        (session / "Data_k").mkdir()
        Manifest.load_or_create(session).save(session)

        report = diagnose_and_repair_session(session, apply=False, client=_FakeRepairClient())
        assert report["participation_id"] is None
        assert any("participation" in e.lower() for e in report["errors"])
        assert report["suggested_actions"] == ["noop"]

    def test_registry_update_targeted(self, tmp_path):
        from repair import diagnose_and_repair_session
        from registry import Registry

        workspace = tmp_path / "ws"
        workspace.mkdir()
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav"], flags={"uploaded": False},
        )
        reg = Registry(workspace)
        reg.upsert_session({
            "id": session.name,
            "path": str(session),
            "canonical_name": session.name,
            "uploaded": 0,
            "analyzed": 0,
        })
        client = _FakeRepairClient(etat="TERMINE", files=["a.wav"])

        report = diagnose_and_repair_session(
            session, apply=True, client=client, registry=reg,
            allow_fetch=True,
        )
        assert "set_uploaded_true" in report["applied_actions"]
        row = reg.get_session(session.name)
        assert row is not None
        assert int(row["uploaded"]) == 1
        assert int(row["analyzed"]) == 1  # xlsx fetché
        assert row.get("api_etat") == "TERMINE"
        reg.close()

    def test_never_trigger_if_missing_even_with_confirm(self, tmp_path):
        from repair import diagnose_and_repair_session
        session = _session_with_manifest(
            tmp_path, wavs=["a.wav", "b.wav"], flags={"uploaded": False},
        )
        client = _FakeRepairClient(etat="", files=["a.wav"])

        diagnose_and_repair_session(
            session, apply=True, client=client,
            allow_trigger=True, confirm_trigger=True,
        )
        assert client.trigger_calls == 0


# =========================================================================
# point_selection + chirosurf_nights (SPEC v0.6 / issue #3)
# =========================================================================

class TestPointSelection:
    def test_haversine_and_radius(self):
        from point_selection import haversine_km, filter_points_within_radius
        # ~111 km per degree latitude
        d = haversine_km(45.0, 5.0, 45.01, 5.0)
        assert 0.5 < d < 2.0
        pts = [{"lat": 45.0, "lon": 5.0}, {"lat": 46.0, "lon": 5.0}]
        near = filter_points_within_radius(
            pts, center_lat=45.0, center_lon=5.0, radius_km=5)
        assert len(near) == 1
        assert near[0]["_distance_km"] < 1

    def test_resolve_focus_manifest_first(self):
        from point_selection import resolve_focus_coords, PointSelection
        meta = {
            "n_site_tadarida": "381009",
            "n_point_fixe": "Z1",
            "point_lat": 45.123,
            "point_lon": 5.678,
        }
        r = resolve_focus_coords(manifest_meta=meta)
        assert r is not None
        lat, lon, label = r
        assert abs(lat - 45.123) < 1e-9
        assert "Z1" in label

        ps = PointSelection(site_numero="1", point_code="z2",
                            lat=1.0, lon=2.0, provenance="other")
        m = ps.to_manifest_meta()
        assert m["n_site_tadarida"] == "000001"
        assert m["n_point_fixe"] == "Z2"
        assert m["point_lat"] == 1.0
        assert "autre" in ps.label_humain.lower() or "obs" in ps.label_humain.lower()

    def test_format_label_human_first(self):
        from point_selection import format_point_label
        lbl = format_point_label("381009", "Z1", commune="Vif")
        assert lbl.startswith("Vif")
        assert "Z1" in lbl
        assert "381009" in lbl


class TestChiroSurfNights:
    def test_split_biological_nights_benjamin(self):
        """Fixtures issue #3 : 8000 + 7839 = multi."""
        from pathlib import Path
        from chirosurf_nights import (
            read_csv, split_rows_by_biological_night, biological_night_key,
        )
        sample = Path(__file__).resolve().parent.parent / (
            "samples/issue3_benjamin/multi_nuits-observations.csv")
        if not sample.is_file():
            pytest.skip("samples issue #3 absents")
        headers, rows = read_csv(sample)
        slices = split_rows_by_biological_night(headers, rows)
        assert len(slices) == 2
        assert slices[0].n_contacts == 8000
        assert slices[1].n_contacts == 7839
        assert slices[0].night_date.isoformat() == "2026-07-28"
        assert slices[1].night_date.isoformat() == "2026-07-29"

    def test_naming_d11(self):
        from chirosurf_nights import raw_csv_name, vu_csv_name, origin_stem_from_xlsx_name
        stem = origin_stem_from_xlsx_name(
            "444976eda-participation-6a70-observations.xlsx")
        assert "observations" in stem
        assert raw_csv_name(1, stem) == f"Nuit1_{stem}.csv"
        assert vu_csv_name(1, stem) == f"Nuit1_{stem}_Vu.csv"

    def test_prepare_writes_lazy(self, tmp_path):
        from pathlib import Path
        from chirosurf_nights import (
            prepare_chirosurf_nights, list_chirosurf_nights,
        )
        # build mini multi-night xlsx-like via csv then openpyxl path needs xlsx
        # use write_csv + rows_from path via prepare expecting xlsx — write xlsx
        import openpyxl
        session = tmp_path / "sess"
        session.mkdir()
        xlsx = session / "participation-abc-observations.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["nom du fichier", "tadarida_taxon", "tadarida_probabilite",
                    "observateur_taxon", "observateur_probabilite",
                    "temps_debut", "temps_fin", "frequence_mediane",
                    "tadarida_taxon_autre", "validateur_taxon",
                    "validateur_probabilite"])
        # night 1 evening + night 1 morning after midnight
        ws.append(["Car381009-2026-Pass1-Z1-SMU_20260728_220000_000",
                    "Pippip", 0.9, "", "", 0, 1, 40, "", "", ""])
        ws.append(["Car381009-2026-Pass1-Z1-SMU_20260729_020000_000",
                    "Pippip", 0.8, "", "", 0, 1, 40, "", "", ""])
        # night 2
        ws.append(["Car381009-2026-Pass1-Z1-SMU_20260729_220000_000",
                    "Barbar", 0.95, "", "", 0, 1, 30, "", "", ""])
        wb.save(xlsx)

        nights = prepare_chirosurf_nights(session, xlsx)
        assert len(nights) == 2
        assert nights[0].has_raw and nights[0].raw_path.is_file()
        assert nights[0].n_contacts == 2
        assert nights[1].n_contacts == 1
        assert (session / "chirosurf").is_dir()
        # second call does not require force
        nights2 = list_chirosurf_nights(session)
        assert len(nights2) == 2

    def test_biological_night_cuts_at_noon(self):
        from chirosurf_nights import biological_night_key
        assert biological_night_key(
            "Car381009-2026-Pass1-Z1-X_20260717_015900_000.wav") == "2026-07-16"
        assert biological_night_key(
            "Car381009-2026-Pass1-Z1-X_20260717_115959_000.wav") == "2026-07-16"
        assert biological_night_key(
            "Car381009-2026-Pass1-Z1-X_20260717_120000_000.wav") == "2026-07-17"
        assert biological_night_key(
            "Car381009-2026-Pass1-Z1-X_20260716_210000_000.wav") == "2026-07-16"

    def test_synthesis_night_menu_and_resolve(self):
        from pathlib import Path
        from chirosurf_nights import (
            read_csv, split_rows_by_biological_night,
            synthesis_night_menu, resolve_synthesis_table,
        )
        sample = Path(__file__).resolve().parent.parent / (
            "samples/issue3_benjamin/multi_nuits-observations.csv")
        if not sample.is_file():
            pytest.skip("samples issue #3 absents")
        headers, rows = read_csv(sample)
        slices = split_rows_by_biological_night(headers, rows)
        menu = synthesis_night_menu(slices, vu_indexes={1})
        keys = [k for k, _l in menu]
        assert keys[0] == 0
        assert 1 in keys and 2 in keys
        assert any("_Vu" in lab for k, lab in menu if k == 1)
        h1, r1, src1, mixed1 = resolve_synthesis_table(
            headers, rows, night_index=1)
        assert not mixed1
        assert len(r1) == 8000
        assert "Nuit 1" in src1
        h0, r0, src0, mixed0 = resolve_synthesis_table(
            headers, rows, night_index=0)
        assert mixed0
        assert len(r0) == 15839


class TestSynthesisMinProba:
    def test_min_proba_keeps_validated(self):
        from synthesis import compute_night_synthesis
        headers = ["nom du fichier", "tadarida_taxon", "tadarida_probabilite",
                    "observateur_taxon"]
        rows = [
            ["a", "Pippip", "0.9", ""],
            ["b", "Pippip", "0.2", ""],
            ["c", "Barbar", "0.1", "Barbar"],  # validé malgré faible proba
        ]
        res = compute_night_synthesis(headers, rows, min_tadarida_proba=0.5)
        by = {s["taxon"]: s for s in res["species"]}
        assert by["Pippip"]["n_contacts"] == 1
        assert by["Barbar"]["n_contacts"] == 1
        assert res["total_contacts"] == 2


if __name__ == "__main__":
    # Permet de lancer directement : python tests/test_core.py
    import sys
    sys.exit(pytest.main([__file__, "-v"]))
