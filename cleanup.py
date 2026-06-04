"""
cleanup.py — nettoyage post-Tadarida.

Lit le fichier d'observations ``participation-<id>-observations*.xlsx`` (ou
``.csv``) déposé dans une session, applique des seuils **par groupe
taxonomique**, et :

  1. copie l'xlsx vers ``<nom>_cleanup.xlsx`` en y ajoutant 3 colonnes qui
     explicitent la décision par contact :
        - ``cleanup_group``       : 'chiros'|'orthos'|'micromam'|'oiseaux'|'noise'|'unknown'
        - ``cleanup_decision``    : 'kept'|'deleted_noise'|'deleted_low_confidence'|
                                    'deleted_group_disabled'|'deleted_no_contact'
        - ``cleanup_reason``      : texte court lisible
  2. écrit ``_stats_before_cleanup.json`` (distribution, comptages)
  3. supprime les WAV dans ``Data_k/`` (ou dossier équivalent) qui ne sont
     gardés par aucun contact
  4. enregistre l'action ``cleanup`` dans le manifest

Règle de décision par WAV (demande utilisateur) :
  → un WAV est CONSERVÉ dès lors qu'**au moins un** de ses contacts est conservé.
  → sinon il est supprimé.

Un WAV présent sur disque mais sans contact dans le xlsx est traité selon
``--silent-policy`` : ``delete`` (défaut, le fichier est considéré silencieux)
ou ``keep``.

L'xlsx original n'est **jamais modifié**.

Exemple d'usage :

    python cleanup.py "E:/.../20250903_site212097_Z3_Pass2_enr07" \\
        --threshold-chiros 0.5 \\
        --threshold-orthos 0.5 \\
        --threshold-micromam 0.5 \\
        --threshold-oiseaux 0.5 \\
        --dry-run

Options clés :
  --threshold-<groupe>  seuil minimal pour conserver (défaut 0.5)
  --disable-<groupe>    désactive complètement le groupe (supprime tout)
  --silent-policy       delete|keep (défaut delete)
  --dry-run             ne supprime rien, affiche le plan
  --force               relance même si le manifest dit déjà nettoyé
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
import warnings
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# PyInstaller --windowed : sys.stdout peut être None (pas de console).
# On guard pour éviter AttributeError au démarrage du .exe.
if sys.stdout is not None and getattr(sys.stdout, "encoding", None):
    if sys.stdout.encoding.lower() != "utf-8":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            if sys.stderr is not None:
                sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

import openpyxl
from openpyxl.styles import PatternFill

from chiro_core import _dir_has_wavs, _find_raw_wav_subdir, classify_wav_name
from manifest import Manifest
from taxons import GROUPS, classify_taxon


try:
    from version import __version__ as TOOL_VERSION
except Exception:
    TOOL_VERSION = "0.1-dev"


# ---------------------------------------------------------------------------
# Lecture de l'observations xlsx/csv
# ---------------------------------------------------------------------------

EXPECTED_HEADERS = {
    "nom du fichier": "filename",
    "temps_debut": "t_start",
    "temps_fin": "t_end",
    "frequence_mediane": "freq_med",
    "tadarida_taxon": "tadarida_taxon",
    "tadarida_probabilite": "tadarida_prob",
    "tadarida_taxon_autre": "tadarida_other",
    "observateur_taxon": "obs_taxon",
    "observateur_probabilite": "obs_prob",
    "validateur_taxon": "val_taxon",
    "validateur_probabilite": "val_prob",
}


def _find_observations_file(session: Path) -> Path | None:
    """Cherche l'xlsx ou csv d'observations à la racine de la session."""
    candidates: list[Path] = []
    for p in session.iterdir():
        if not p.is_file():
            continue
        n = p.name.lower()
        if n.startswith("participation-") and "observations" in n and (
            n.endswith(".xlsx") or n.endswith(".csv")
        ):
            # On exclut nos copies _cleanup pour ne pas boucler
            if "_cleanup" in n:
                continue
            candidates.append(p)
    if not candidates:
        return None
    # Plus récent en priorité
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _read_observations(path: Path) -> tuple[list[str], list[list]]:
    """Retourne (headers_originaux, lignes_brutes)."""
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return [], []
    data = [list(r) for r in rows_iter if any(c is not None for c in r)]
    return [str(h) if h is not None else "" for h in headers], data


def _col_map(headers: list[str]) -> dict[str, int]:
    """{logical_key: col_index} depuis l'en-tête."""
    m: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = (h or "").strip().lower()
        if key in EXPECTED_HEADERS:
            m[EXPECTED_HEADERS[key]] = i
    return m


# ---------------------------------------------------------------------------
# Décision par contact puis par fichier
# ---------------------------------------------------------------------------

DECISIONS = (
    "kept",
    "deleted_noise",
    "deleted_low_confidence",
    "deleted_group_disabled",
    "deleted_unknown_group",   # par défaut gardé, mais paramétrable
)


def decide_contact(
    taxon: str | None,
    proba: float | None,
    thresholds: dict[str, float],
    disabled: set[str],
    unknown_action: str = "keep",
) -> tuple[str, str, str]:
    """Retourne (group, decision, reason) pour une ligne du xlsx."""
    group = classify_taxon(taxon)

    if group == "noise":
        return group, "deleted_noise", f"tadarida_taxon={taxon!r} → noise"

    if group in disabled:
        return group, "deleted_group_disabled", f"groupe {group} désactivé"

    if group == "unknown":
        if unknown_action == "delete":
            return group, "deleted_unknown_group", f"taxon {taxon!r} non classé, action=delete"
        return group, "kept", f"taxon {taxon!r} non classé, action=keep"

    threshold = thresholds.get(group, 0.5)
    if proba is None:
        return group, "kept", f"{group} sans proba → conservé par défaut"
    if proba < threshold:
        return group, "deleted_low_confidence", (
            f"tadarida_probabilite={proba:.2f} < {threshold:.2f} (seuil {group})"
        )
    return group, "kept", f"{group} : proba {proba:.2f} ≥ {threshold:.2f}"


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _norm_stem(name: str) -> str:
    """Normalise un nom pour comparaison : sans extension, casse respectée."""
    p = Path(str(name))
    return p.stem if p.suffix else p.name


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def cleanup_session(
    session: Path,
    thresholds: dict[str, float],
    disabled: set[str],
    silent_policy: str = "delete",
    unknown_action: str = "keep",
    dry_run: bool = False,
    force: bool = False,
    mass_delete_ratio: float = 0.8,
    allow_mass_delete: bool = False,
    progress=None,
) -> dict:
    """Pipeline complet de nettoyage.

    Paramètres de sécurité :
      - ``mass_delete_ratio`` : ratio au-delà duquel on considère que le cleanup
        "va trop loin" (seuil mal configuré, xlsx incomplet, etc.). Défaut 0.8
        = 80 % des WAV.
      - ``allow_mass_delete`` : si False (défaut), le cleanup refuse
        d'exécuter la suppression si elle dépasse ``mass_delete_ratio``. Le
        xlsx annoté + stats sont quand même écrits pour que l'utilisateur
        puisse inspecter la décision avant de relancer avec le flag ``True``.
    """
    session = session.resolve()
    out: dict = {
        "session": str(session),
        "dry_run": dry_run,
        "errors": [],
        "warnings": [],
    }

    # --- manifest
    m = Manifest.load_or_create(session)
    if m.is_done("cleanup") and not force:
        out["warnings"].append("déjà nettoyé (cf manifest). Utilise --force pour relancer.")
        return out

    # --- repérage de l'xlsx
    obs_path = _find_observations_file(session)
    if obs_path is None:
        out["errors"].append(
            "pas de fichier 'participation-*-observations*.xlsx/csv' trouvé à la racine"
        )
        return out
    out["observations_input"] = str(obs_path)

    # --- repérage du dossier WAV à nettoyer (priorité : Data_k/)
    candidates = [session / "Data_k", session / "1-K"]
    data_k = next((p for p in candidates if p.is_dir() and _dir_has_wavs(p)), None)
    if data_k is None:
        # fallback : sous-dossier Data/ (cas d'un upload pré-TE) ou la session elle-même
        sub = _find_raw_wav_subdir(session)
        data_k = sub if sub is not None else (session if _dir_has_wavs(session) else None)

    if data_k is None:
        out["errors"].append("aucun dossier WAV trouvé dans la session")
        return out
    out["wav_dir"] = str(data_k)

    # --- lecture du xlsx
    headers, rows = _read_observations(obs_path)
    col = _col_map(headers)
    if "filename" not in col or "tadarida_taxon" not in col:
        out["errors"].append(
            f"colonnes attendues absentes dans {obs_path.name} : "
            f"trouvé={list(col.keys())}"
        )
        return out

    out["n_rows_xlsx"] = len(rows)

    # --- décisions par contact
    # contacts_by_file : {stem_normalisé: [decision_info, ...]}
    contacts_by_file: dict[str, list[dict]] = defaultdict(list)
    enriched_rows: list[list] = []

    stats_contact = Counter()
    stats_group = Counter()
    prob_dist: dict[str, list[float]] = defaultdict(list)

    for r in rows:
        # on copie la ligne pour y ajouter 3 colonnes
        row = list(r)
        filename = row[col["filename"]] if col["filename"] < len(row) else None

        # Priorité aux validations humaines (validateur > observateur > tadarida).
        # Si un humain a corrigé un faux noise en "Pippip" via ChiroSurf, on
        # respecte SA décision et on n'applique PAS le seuil de probabilité
        # (qui n'a plus de sens — c'est validé manuellement).
        val_taxon = None
        if "val_taxon" in col and col["val_taxon"] < len(row):
            v = row[col["val_taxon"]]
            if v is not None and str(v).strip():
                val_taxon = str(v).strip()
        obs_taxon = None
        if "obs_taxon" in col and col["obs_taxon"] < len(row):
            v = row[col["obs_taxon"]]
            if v is not None and str(v).strip():
                obs_taxon = str(v).strip()

        tadarida_taxon = row[col["tadarida_taxon"]] if col["tadarida_taxon"] < len(row) else None
        proba = _to_float(row[col["tadarida_prob"]]) if "tadarida_prob" in col and col["tadarida_prob"] < len(row) else None

        # Choix du taxon de référence
        if val_taxon:
            taxon = val_taxon
            # Forçage proba à 1.0 : validation humaine = certitude
            effective_proba = 1.0
        elif obs_taxon:
            taxon = obs_taxon
            effective_proba = 1.0
        else:
            taxon = tadarida_taxon
            effective_proba = proba

        group, decision, reason = decide_contact(
            taxon, effective_proba, thresholds, disabled, unknown_action
        )
        stats_contact[decision] += 1
        stats_group[group] += 1
        if effective_proba is not None:
            prob_dist[group].append(effective_proba)

        stem = _norm_stem(filename) if filename else ""
        contacts_by_file[stem].append({
            "group": group, "decision": decision, "reason": reason,
            "taxon": taxon, "proba": effective_proba,
            "validated_by_human": bool(val_taxon or obs_taxon),
        })
        enriched_rows.append(row + [group, decision, reason])

    # --- décision par fichier (OR : au moins 1 "kept" → on garde)
    file_kept: set[str] = set()
    for stem, contacts in contacts_by_file.items():
        if any(c["decision"] == "kept" for c in contacts):
            file_kept.add(stem)

    # --- balayage des WAV physiques
    wavs = [p for p in data_k.iterdir() if p.is_file() and p.suffix.lower() == ".wav"]
    total_wav = len(wavs)
    to_delete: list[Path] = []
    silent_files: list[Path] = []
    kept_files: list[Path] = []

    for w in wavs:
        stem = w.stem
        if stem in contacts_by_file:
            if stem in file_kept:
                kept_files.append(w)
            else:
                to_delete.append(w)
        else:
            silent_files.append(w)

    if silent_policy == "delete":
        to_delete.extend(silent_files)
    else:
        kept_files.extend(silent_files)

    # --- stats avant suppression
    total_bytes_to_delete = 0
    for p in to_delete:
        try:
            total_bytes_to_delete += p.stat().st_size
        except OSError:
            pass
    total_bytes_all = 0
    for p in wavs:
        try:
            total_bytes_all += p.stat().st_size
        except OSError:
            pass

    def _pdist(lst: list[float]) -> dict | None:
        if not lst:
            return None
        s = sorted(lst)
        n = len(s)
        return {
            "n": n,
            "min": s[0], "q1": s[n // 4], "median": s[n // 2],
            "q3": s[3 * n // 4], "max": s[-1],
        }

    stats: dict = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "observations_file": obs_path.name,
        "thresholds": thresholds,
        "disabled_groups": sorted(disabled),
        "silent_policy": silent_policy,
        "unknown_action": unknown_action,
        "contacts": {
            "total": len(rows),
            "by_decision": dict(stats_contact),
            "by_group": dict(stats_group),
            "prob_dist_by_group": {g: _pdist(v) for g, v in prob_dist.items()},
        },
        "files": {
            "total_wav_on_disk": total_wav,
            "in_xlsx": len(contacts_by_file),
            "silent_on_disk": len(silent_files),
            "planned_kept": len(kept_files),
            "planned_deleted": len(to_delete),
            "bytes_total": total_bytes_all,
            "bytes_to_delete": total_bytes_to_delete,
        },
    }
    out["stats"] = stats

    # --- écriture de l'xlsx enrichi
    new_headers = list(headers) + ["cleanup_group", "cleanup_decision", "cleanup_reason"]
    out_xlsx = obs_path.with_name(obs_path.stem + "_cleanup.xlsx")
    out["observations_output"] = str(out_xlsx)

    if not dry_run:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(new_headers)
        # Coloration légère des lignes supprimées
        fill_del = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        fill_keep = PatternFill(start_color="EFFFE6", end_color="EFFFE6", fill_type="solid")
        for row in enriched_rows:
            ws.append(row)
            last_row = ws.max_row
            dec = row[-2]
            if dec.startswith("deleted"):
                for c in ws[last_row]:
                    c.fill = fill_del
            elif dec == "kept":
                for c in ws[last_row]:
                    c.fill = fill_keep
        wb.save(out_xlsx)

        # stats json — écrit AVANT la suppression pour garder une trace même
        # si la suppression est bloquée par le garde-fou mass-delete.
        stats_path = session / "_stats_before_cleanup.json"
        with stats_path.open("w", encoding="utf-8") as f:
            json.dump(stats, f, indent=2, ensure_ascii=False)
        out["stats_file"] = str(stats_path)

        # GARDE-FOU : si on s'apprête à supprimer > ``mass_delete_ratio`` des
        # WAV, on bloque sauf si ``allow_mass_delete=True``. Protège contre un
        # seuil mal configuré (ex: --threshold-chiros=0.99) qui détruirait
        # la nuit entière. L'xlsx annoté + stats sont déjà écrits pour revue.
        ratio = (len(to_delete) / total_wav) if total_wav else 0.0
        if ratio > mass_delete_ratio and not allow_mass_delete:
            out["errors"].append(
                f"Suppression bloquée : {len(to_delete)}/{total_wav} WAV "
                f"({ratio:.0%}) seraient supprimés (seuil sécurité "
                f"{mass_delete_ratio:.0%}). Inspecte {out_xlsx.name} puis "
                f"relance avec allow_mass_delete=True si c'est intentionnel."
            )
            out["mass_delete_ratio"] = ratio
            out["n_deleted"] = 0
            # On ne record PAS l'action cleanup : le manifest reste en état
            # non-nettoyé, l'utilisateur peut corriger et retenter.
            return out

        # suppression effective (avec progress callback toutes les ~10 unlinks)
        errors = 0
        total_to_delete = len(to_delete)
        for i, p in enumerate(to_delete, 1):
            try:
                p.unlink()
            except OSError as e:
                errors += 1
                out["warnings"].append(f"échec suppression {p.name} : {e}")
            if progress is not None and (i % 10 == 0 or i == total_to_delete):
                try:
                    progress(i, total_to_delete, "cleanup WAV")
                except Exception:
                    pass
        out["n_deleted"] = len(to_delete) - errors

        # manifest — status=ok par défaut, patché en verify_failed par
        # pipeline.run_phase_cleanup si la vérification indépendante échoue.
        m.record_action(
            "cleanup",
            status="ok",
            params={
                "thresholds": thresholds,
                "disabled_groups": sorted(disabled),
                "silent_policy": silent_policy,
                "unknown_action": unknown_action,
                "observations_file": obs_path.name,
            },
            stats={
                "wav_total": total_wav,
                "wav_deleted": out["n_deleted"],
                "wav_kept": total_wav - out["n_deleted"],
                "bytes_deleted": total_bytes_to_delete,
                "bytes_kept": total_bytes_all - total_bytes_to_delete,
                "contacts_by_decision": dict(stats_contact),
                "contacts_by_group": dict(stats_group),
            },
            tool_version=TOOL_VERSION,
        )
        m.stats_before_cleanup = stats
        m.save(session)

    return out


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _print_plan(out: dict) -> None:
    print(f"\nSession : {out['session']}")
    if out.get("observations_input"):
        print(f"Obs (entrée)    : {Path(out['observations_input']).name}")
    if out.get("wav_dir"):
        print(f"WAV à balayer   : {out['wav_dir']}")
    if out.get("stats"):
        s = out["stats"]
        c = s["contacts"]
        f = s["files"]
        print(f"\nContacts xlsx : {c['total']}")
        for dec in ("kept", "deleted_noise", "deleted_low_confidence",
                    "deleted_group_disabled", "deleted_unknown_group"):
            n = c["by_decision"].get(dec, 0)
            if n:
                print(f"   {dec:<28s} {n:6d}")
        print("\nPar groupe :")
        for g in GROUPS:
            n = c["by_group"].get(g, 0)
            if n:
                print(f"   {g:<10s} {n:6d}")
        print(f"\nFichiers WAV sur disque : {f['total_wav_on_disk']}")
        print(f"   dans xlsx            : {f['in_xlsx']}")
        print(f"   silencieux           : {f['silent_on_disk']}")
        print(f"   → conservés (plan)   : {f['planned_kept']}")
        print(f"   → supprimés (plan)   : {f['planned_deleted']}")
        mb = f['bytes_to_delete'] / (1024 * 1024)
        total_mb = f['bytes_total'] / (1024 * 1024)
        print(f"   → espace libéré      : {mb:.1f} MB  / {total_mb:.1f} MB total")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("session", type=Path)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--silent-policy", choices=("delete", "keep"), default="delete",
                    help="sort des WAV présents mais absents du xlsx (défaut: delete)")
    ap.add_argument("--unknown-action", choices=("keep", "delete"), default="keep",
                    help="sort des contacts non classés (défaut: keep = conservés)")
    for g in ("chiros", "orthos", "micromam", "oiseaux"):
        ap.add_argument(f"--threshold-{g}", type=float, default=0.5,
                        help=f"seuil minimum de proba pour {g} (défaut 0.5)")
        ap.add_argument(f"--disable-{g}", action="store_true",
                        help=f"désactive le groupe {g} (= supprime tout)")
    args = ap.parse_args()

    thresholds = {
        "chiros": args.threshold_chiros,
        "orthos": args.threshold_orthos,
        "micromam": args.threshold_micromam,
        "oiseaux": args.threshold_oiseaux,
    }
    disabled: set[str] = {
        g for g in ("chiros", "orthos", "micromam", "oiseaux")
        if getattr(args, f"disable_{g}")
    }

    if not args.session.is_dir():
        print(f"❌ session introuvable : {args.session}", file=sys.stderr)
        return 2

    print(f"Seuils : {thresholds}")
    if disabled:
        print(f"Groupes désactivés : {sorted(disabled)}")
    print(f"Politique silencieux : {args.silent_policy}")
    print(f"Politique taxon non classé : {args.unknown_action}")

    out = cleanup_session(
        session=args.session,
        thresholds=thresholds,
        disabled=disabled,
        silent_policy=args.silent_policy,
        unknown_action=args.unknown_action,
        dry_run=args.dry_run,
        force=args.force,
    )

    _print_plan(out)
    if out["warnings"]:
        print("\n⚠  Avertissements :")
        for w in out["warnings"]:
            print(f"   - {w}")
    if out["errors"]:
        print("\n❌ Erreurs :", file=sys.stderr)
        for e in out["errors"]:
            print(f"   - {e}", file=sys.stderr)
        return 1

    if args.dry_run:
        print("\n(dry-run, rien n'a été touché)")
    else:
        print(f"\n✓ xlsx annoté   : {Path(out['observations_output']).name}")
        print(f"✓ stats         : _stats_before_cleanup.json")
        print(f"✓ WAV supprimés : {out.get('n_deleted', 0)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
