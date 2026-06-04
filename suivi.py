"""
suivi.py — lecture du fichier "Suivi analyse Chiros YYYY.xlsx"

Expose une API simple pour :
  - lister les contrats (nom + nb de lignes)
  - retrouver les lignes d'un contrat donné
  - obtenir la table des enregistreurs (Feuil2 : n° 1→28 ↔ n° série ↔ n° micro)
  - résoudre un enregistreur depuis son n° court (11) OU son n° de série (SMU03252)
  - proposer les (site, point, passage) connus pour un contrat (autocomplete GUI)

Gère les 2 schémas observés :
  - 2025 : 24 colonnes (pas de CP)
  - 2026 : 25 colonnes (+ colonne CP en position 1)

Usage CLI :
    python suivi.py                                # résumé du 2026
    python suivi.py --year 2025                    # résumé du 2025
    python suivi.py --contrat "MonContrat"         # détail contrat (match substring)
    python suivi.py --enregistreur 11              # détail d'un enregistreur
    python suivi.py --serial SMU03252              # résolution par n° série
"""

from __future__ import annotations

import argparse
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

# openpyxl râle sur les Data Validations du fichier, on coupe le warning
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# PyInstaller --windowed : sys.stdout peut être None (pas de console).
# On guard pour éviter AttributeError au démarrage du .exe.
if sys.stdout is not None and getattr(sys.stdout, "encoding", None):
    if sys.stdout.encoding.lower() != "utf-8":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            if sys.stderr is not None:
                sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

import openpyxl


# ---------------------------------------------------------------------------
# Modèles
# ---------------------------------------------------------------------------

@dataclass
class Enregistreur:
    num: int
    marque: str | None = None
    modele: str | None = None
    serie_enr: str | None = None
    serie_micro: str | None = None

    def __str__(self) -> str:
        return (
            f"#{self.num:<2}  {self.marque or '?':<20}  {self.modele or '?':<15}  "
            f"série={self.serie_enr!r:<15}  micro={self.serie_micro!r}"
        )


@dataclass
class SuiviRow:
    row_index: int                  # ligne Excel (1-indexed, header = 1)
    nom_contrat: str | None = None
    cp: str | None = None           # présent en 2026+
    date_debut: datetime | None = None
    date_fin: datetime | None = None
    emplacement: str | None = None
    couverture: str | None = None
    vent: str | None = None
    temperatures: str | None = None  # plage "28-19" ou similaire
    n_site_tadarida: str | None = None   # 6 chiffres ou plage
    n_point_fixe: str | None = None       # A1, C2, Z4...
    n_passage: int | None = None          # 1, 2, 3 (normalisé depuis " Passage 1")
    n_enregistreur: int | None = None     # 1-28
    n_serie: str | None = None
    n_micro: str | None = None
    etat: str | None = None
    flag_lupas_kaleidoscope: bool | None = None
    flag_tadarida: bool | None = None
    traitement_fichier: str | None = None
    verification: str | None = None
    verificateur: str | None = None
    commentaires: str | None = None
    etat_avancement: str | None = None
    echeance: str | None = None
    sous_traitance: str | None = None
    envoi_st_cl: str | None = None

    def key(self) -> tuple:
        """Clé d'identification "métier" d'une ligne."""
        return (
            (self.nom_contrat or "").strip(),
            self.date_debut.date() if self.date_debut else None,
            self.n_site_tadarida,
            self.n_point_fixe,
            self.n_passage,
            self.n_enregistreur,
        )

    def summary(self) -> str:
        date = self.date_debut.strftime("%Y-%m-%d") if self.date_debut else "??"
        return (
            f"L{self.row_index:3d}  {date}  {self.nom_contrat or '?':<40}  "
            f"site={self.n_site_tadarida!r:<10} "
            f"pt={self.n_point_fixe!r:<6} "
            f"pass={self.n_passage} "
            f"enr=#{self.n_enregistreur} ({self.n_serie})"
        )


# ---------------------------------------------------------------------------
# Normalisation de cellules
# ---------------------------------------------------------------------------

def _to_str(v) -> str | None:
    if v is None or v == "":
        return None
    return str(v).strip() or None


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    try:
        return int(s)
    except ValueError:
        return None


def _to_bool(v) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "vrai", "oui", "ok", "x", "1"):
        return True
    if s in ("false", "faux", "non", "0"):
        return False
    return None


def _to_dt(v) -> datetime | None:
    if isinstance(v, datetime):
        return v
    return None


def _normalize_passage(v) -> int | None:
    """
    ' Passage 1' -> 1, 'Passage 2' -> 2, 3 -> 3, 'P1' -> 1.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            n = int(v)
            return n if 1 <= n <= 9 else None
        except (ValueError, OverflowError):
            return None
    s = str(v).strip().lower().replace("passage", "").replace("pass", "").replace("p", "").strip()
    try:
        return int(s) if s else None
    except ValueError:
        return None


def _normalize_site_tadarida(v) -> str | None:
    """
    Préserve les 6 chiffres en zéro-paddé (ex : 45945 -> '045945').
    Pour les plages non normalisées ('13-15'), retourne la chaîne telle quelle.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return f"{int(v):06d}"
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    if s.isdigit():
        return f"{int(s):06d}"
    return s


# ---------------------------------------------------------------------------
# Colonnes attendues — détection automatique 2025 vs 2026
# ---------------------------------------------------------------------------

# Synonymes tolérés (header → clé logique)
HEADER_MAP = {
    "nom contrat": "nom_contrat",
    "cp": "cp",
    "date début (jj/mm/aaaa)": "date_debut",
    "date de fin (jj/mm/aaaa)": "date_fin",
    "emplacement": "emplacement",
    "couverture nuageuse": "couverture",
    "vent (km/h)": "vent",
    "n° site tadarida": "n_site_tadarida",
    "n° point fixe": "n_point_fixe",
    "n° passage": "n_passage",
    "n° enregistreur": "n_enregistreur",
    "n° série": "n_serie",
    "n° micro": "n_micro",
    "etat": "etat",
    "lupas + kaleidoscope": "flag_lupas_kaleidoscope",
    "tadarida": "flag_tadarida",
    "tcgmt fichier": "traitement_fichier",
    "vérification": "verification",
    "vérificateur": "verificateur",
    "commentaires": "commentaires",
    "etat d'avancement": "etat_avancement",
    "échéance": "echeance",
    "sous-traitance": "sous_traitance",
    "envoi st à cl": "envoi_st_cl",
}


def _map_headers(header_row: tuple) -> dict[int, str]:
    """Retourne {col_index: logical_key} pour les colonnes reconnues."""
    m: dict[int, str] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower().replace("\n", " ").replace("  ", " ")
        # Températures : match large (le nom a un \n)
        if "températures" in key:
            m[i] = "temperatures"
            continue
        if key in HEADER_MAP:
            m[i] = HEADER_MAP[key]
    return m


# ---------------------------------------------------------------------------
# Suivi
# ---------------------------------------------------------------------------

class Suivi:
    """Lecture du fichier Suivi analyse Chiros (lecture seule)."""

    def __init__(self, path: Path):
        self.path = Path(path)
        self.year: int | None = None
        self.rows: list[SuiviRow] = []
        self.enregistreurs: dict[int, Enregistreur] = {}
        self._load()

    # -- Chargement ---------------------------------------------------------

    def _load(self) -> None:
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        try:
            self._load_from_wb(wb)
        finally:
            # CRITIQUE : sans close(), le handle Windows reste actif et bloque
            # les backups/renames ultérieurs (suivi_write.check_not_locked).
            try:
                wb.close()
            except Exception:
                pass

    def _load_from_wb(self, wb) -> None:
        # Feuille principale : celle dont le nom est l'année, ou sinon la 1ère
        main_sheet = None
        for sn in wb.sheetnames:
            if sn.isdigit() and len(sn) == 4:
                main_sheet = sn
                self.year = int(sn)
                break
        if main_sheet is None:
            main_sheet = wb.sheetnames[0]
        ws = wb[main_sheet]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return
        col_map = _map_headers(header_row)

        for idx, row in enumerate(rows_iter, start=2):
            # Ligne totalement vide → skip
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue

            r = SuiviRow(row_index=idx)
            for col_i, key in col_map.items():
                if col_i >= len(row):
                    continue
                raw = row[col_i]
                if key == "nom_contrat":
                    r.nom_contrat = _to_str(raw)
                elif key == "cp":
                    r.cp = _to_str(raw)
                elif key == "date_debut":
                    r.date_debut = _to_dt(raw)
                elif key == "date_fin":
                    r.date_fin = _to_dt(raw)
                elif key == "emplacement":
                    r.emplacement = _to_str(raw)
                elif key == "couverture":
                    r.couverture = _to_str(raw)
                elif key == "vent":
                    r.vent = _to_str(raw)
                elif key == "temperatures":
                    r.temperatures = _to_str(raw)
                elif key == "n_site_tadarida":
                    r.n_site_tadarida = _normalize_site_tadarida(raw)
                elif key == "n_point_fixe":
                    r.n_point_fixe = _to_str(raw)
                elif key == "n_passage":
                    r.n_passage = _normalize_passage(raw)
                elif key == "n_enregistreur":
                    r.n_enregistreur = _to_int(raw)
                elif key == "n_serie":
                    r.n_serie = _to_str(raw)
                elif key == "n_micro":
                    r.n_micro = _to_str(raw)
                elif key == "etat":
                    r.etat = _to_str(raw)
                elif key == "flag_lupas_kaleidoscope":
                    r.flag_lupas_kaleidoscope = _to_bool(raw)
                elif key == "flag_tadarida":
                    r.flag_tadarida = _to_bool(raw)
                elif key == "traitement_fichier":
                    r.traitement_fichier = _to_str(raw)
                elif key == "verification":
                    r.verification = _to_str(raw)
                elif key == "verificateur":
                    r.verificateur = _to_str(raw)
                elif key == "commentaires":
                    r.commentaires = _to_str(raw)
                elif key == "etat_avancement":
                    r.etat_avancement = _to_str(raw)
                elif key == "echeance":
                    r.echeance = _to_str(raw)
                elif key == "sous_traitance":
                    r.sous_traitance = _to_str(raw)
                elif key == "envoi_st_cl":
                    r.envoi_st_cl = _to_str(raw)
            self.rows.append(r)

        # Feuille Feuil2 : table des enregistreurs
        if "Feuil2" in wb.sheetnames:
            f2 = wb["Feuil2"]
            rows2 = list(f2.iter_rows(values_only=True))
            for row in rows2[1:]:
                num = _to_int(row[0]) if row else None
                if num is None:
                    continue
                self.enregistreurs[num] = Enregistreur(
                    num=num,
                    marque=_to_str(row[1]) if len(row) > 1 else None,
                    modele=_to_str(row[2]) if len(row) > 2 else None,
                    serie_enr=_to_str(row[3]) if len(row) > 3 else None,
                    serie_micro=_to_str(row[4]) if len(row) > 4 else None,
                )

    # -- API ----------------------------------------------------------------

    def contrats(self) -> list[str]:
        seen: list[str] = []
        for r in self.rows:
            if r.nom_contrat and r.nom_contrat not in seen:
                seen.append(r.nom_contrat)
        return seen

    def by_contrat(self, name_substring: str) -> list[SuiviRow]:
        s = name_substring.strip().lower()
        return [r for r in self.rows if r.nom_contrat and s in r.nom_contrat.lower()]

    def points_for_contrat(self, name_substring: str) -> list[tuple[str | None, str | None, int | None]]:
        """Retourne la liste unique (site, point, passage) pour un contrat."""
        out: list[tuple] = []
        seen: set[tuple] = set()
        for r in self.by_contrat(name_substring):
            k = (r.n_site_tadarida, r.n_point_fixe, r.n_passage)
            if k not in seen:
                seen.add(k)
                out.append(k)
        return out

    def enregistreur(self, num: int) -> Enregistreur | None:
        return self.enregistreurs.get(num)

    def enregistreur_by_serial(self, serial: str) -> Enregistreur | None:
        s = serial.strip().lower()
        for e in self.enregistreurs.values():
            if e.serie_enr and e.serie_enr.strip().lower() == s:
                return e
        return None

    def resolve_enregistreur(self, value) -> Enregistreur | None:
        """Accepte un n° (int ou str numérique) ou un numéro de série."""
        as_int = _to_int(value)
        if as_int is not None and as_int in self.enregistreurs:
            return self.enregistreurs[as_int]
        if isinstance(value, str):
            return self.enregistreur_by_serial(value)
        return None

    def find_row_for_session(
        self,
        *,
        nom_contrat: str | None = None,
        date_debut: datetime | None = None,
        n_enregistreur: int | None = None,
        serial: str | None = None,
        n_site_tadarida: str | None = None,
        n_point_fixe: str | None = None,
        n_passage: int | None = None,
    ) -> list[SuiviRow]:
        """Cherche les lignes qui matchent un critère multi-champ.

        Plus il y a de critères, plus c'est spécifique. On retourne toutes les
        lignes compatibles avec les critères non-None fournis (AND).
        """
        matches = list(self.rows)

        if nom_contrat:
            ns = nom_contrat.strip().lower()
            matches = [r for r in matches if r.nom_contrat and ns in r.nom_contrat.lower()]
        if date_debut is not None:
            d = date_debut.date() if hasattr(date_debut, "date") else date_debut
            matches = [r for r in matches if r.date_debut and r.date_debut.date() == d]
        if n_enregistreur is not None:
            matches = [r for r in matches if r.n_enregistreur == n_enregistreur]
        if serial:
            ns = serial.strip().lower()
            matches = [r for r in matches if r.n_serie and r.n_serie.strip().lower() == ns]
        if n_site_tadarida:
            matches = [r for r in matches if r.n_site_tadarida == _normalize_site_tadarida(n_site_tadarida)]
        if n_point_fixe:
            matches = [r for r in matches if r.n_point_fixe == n_point_fixe]
        if n_passage is not None:
            matches = [r for r in matches if r.n_passage == n_passage]

        return matches


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _resolve_workspace_root() -> Path:
    """Détermine la racine workspace où chercher le Suivi.

    Ordre de priorité :
      1. Workspace courant des préférences GUI (settings.last_workspace) s'il
         existe. C'est le cas nominal en mode exe PyInstaller.
      2. Dossier courant Python (cwd) en dernier recours.
      3. Dossier parent du module (dev CLI).

    BUG historique corrigé : avant, on utilisait `Path(__file__).parent.parent`
    toujours, ce qui pointait sur le dossier d'extraction PyInstaller
    (AppData/Local/Temp/_MEIxxxx/) en mode exe. Le Suivi n'était alors
    jamais trouvé.
    """
    # 1. Workspace GUI via gui_config (import lazy pour éviter cycle)
    try:
        from gui_config import load_settings
        s = load_settings()
        if s.last_workspace:
            p = Path(s.last_workspace)
            if p.is_dir():
                return p
    except Exception:
        pass

    # 2. cwd si le module a été lancé depuis un dossier plausible
    cwd = Path.cwd()
    if any(cwd.glob("Suivi analyse Chiros *.xlsx")):
        return cwd

    # 3. Fallback dev CLI : dossier parent du module
    try:
        dev = Path(__file__).resolve().parent.parent
        if dev.is_dir() and any(dev.glob("Suivi analyse Chiros *.xlsx")):
            return dev
    except Exception:
        pass

    # Dernier recours : cwd même si pas de Suivi trouvé (on continuera en
    # erreur "fichier introuvable" explicite, pas un chemin Temp absurde)
    return cwd


def _default_path(year: int | None = None) -> Path:
    root = _resolve_workspace_root()
    if year:
        return root / f"Suivi analyse Chiros {year}.xlsx"
    # Par défaut : on prend l'année la plus récente trouvée
    candidates = sorted(root.glob("Suivi analyse Chiros *.xlsx"))
    if not candidates:
        return root / "Suivi analyse Chiros 2026.xlsx"
    return candidates[-1]


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", type=Path, default=None, help="chemin du .xlsx (défaut : auto)")
    ap.add_argument("--year", type=int, default=None)
    ap.add_argument("--contrat", default=None, help="filtre sur nom de contrat (substring)")
    ap.add_argument("--enregistreur", type=int, default=None, help="n° court 1-28")
    ap.add_argument("--serial", default=None, help="n° de série enregistreur")
    args = ap.parse_args()

    path = args.file if args.file else _default_path(args.year)
    if not path.is_file():
        print(f"❌ introuvable : {path}", file=sys.stderr)
        return 2

    print(f"📂 {path}")
    s = Suivi(path)
    print(f"   {len(s.rows)} lignes · {len(s.contrats())} contrats · "
          f"{len(s.enregistreurs)} enregistreurs (Feuil2)")
    if s.year:
        print(f"   Année feuille : {s.year}")

    if args.enregistreur is not None:
        e = s.enregistreur(args.enregistreur)
        if e:
            print(f"\n{e}")
        else:
            print(f"\n❌ pas d'enregistreur #{args.enregistreur}")
        return 0

    if args.serial:
        e = s.enregistreur_by_serial(args.serial)
        if e:
            print(f"\n{e}")
        else:
            print(f"\n❌ pas d'enregistreur avec série {args.serial!r}")
        return 0

    if args.contrat:
        rows = s.by_contrat(args.contrat)
        print(f"\n=== Contrat '{args.contrat}' : {len(rows)} ligne(s) ===")
        for r in rows:
            print(" ", r.summary())
        pts = s.points_for_contrat(args.contrat)
        print(f"\nCouples (site, point, passage) uniques : {len(pts)}")
        for site, pt, pas in pts:
            print(f"   site={site!r:<10} point={pt!r:<5} passage={pas}")
        return 0

    # Aperçu général
    print("\n=== Contrats (top 15 par nb de lignes) ===")
    from collections import Counter
    c = Counter(r.nom_contrat for r in s.rows if r.nom_contrat)
    for name, n in c.most_common(15):
        print(f"  {n:3d}  {name}")

    print("\n=== Enregistreurs (Feuil2) ===")
    for num in sorted(s.enregistreurs):
        print(" ", s.enregistreurs[num])

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
