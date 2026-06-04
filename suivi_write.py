"""
suivi_write.py — écriture dans le fichier ``Suivi analyse Chiros YYYY.xlsx``.

Garanties :

  1. **Jamais d'écriture sur un fichier ouvert dans Excel** — on détecte le
     fichier verrou ``~$<nom>.xlsx`` avant toute action.
  2. **Backup automatique** horodaté dans ``_backups/`` avant chaque écriture,
     rotation des N derniers (défaut 20).
  3. **Écriture atomique** : on écrit dans un .tmp puis ``.replace()``.
  4. **Préserve la structure** : openpyxl conserve formules et mise en forme.
     Une validation de donnée (Data Validation) complexe peut être perdue
     dans l'écriture — c'est pour ça qu'on garde les backups.
  5. **Idempotent** : si la ligne recherchée existe et a déjà la valeur
     demandée, aucune écriture.

API principale :
  * ``SuiviWriter(path).update_row(key, updates)``
  * ``SuiviWriter(path).mark_step_done(key, step, done=True)``
  * ``SuiviWriter(path).set_state(key, etat, commentaire=None)``

Usage CLI (démonstration / dépannage) :

    python suivi_write.py --contrat "MonContrat" --date 2025-09-03 \\
        --site 212097 --point Z3 --pass 2 --enr 7 \\
        --set-lupas-kaleidoscope --state "ok"
"""

from __future__ import annotations

import argparse
import shutil
import sys
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# PyInstaller --windowed : sys.stdout peut être None (pas de console).
# On guard pour éviter AttributeError au démarrage du .exe.
if sys.stdout is not None and getattr(sys.stdout, "encoding", None):
    if sys.stdout.encoding.lower() != "utf-8":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            if sys.stderr is not None:
                sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

import openpyxl

from suivi import HEADER_MAP, _normalize_passage, _normalize_site_tadarida


# ---------------------------------------------------------------------------
# Clé d'identification d'une ligne
# ---------------------------------------------------------------------------

@dataclass
class RowKey:
    """Clé métier d'une ligne du Suivi. Tous les champs sauf
    ``nom_contrat`` peuvent être None pour une recherche approximative."""
    nom_contrat: str | None = None
    date_debut: datetime | None = None
    n_site_tadarida: str | None = None
    n_point_fixe: str | None = None
    n_passage: int | None = None
    n_enregistreur: int | None = None


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class SuiviFileLockedError(RuntimeError):
    """Le fichier est ouvert dans Excel (ou verrouillé)."""


class SuiviRowNotFoundError(RuntimeError):
    pass


class SuiviAmbiguousKeyError(RuntimeError):
    pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _excel_lock_path(xlsx_path: Path) -> Path:
    """Fichier verrou créé par Excel quand le fichier est ouvert : '~$<name>.xlsx'."""
    return xlsx_path.with_name("~$" + xlsx_path.name)


def check_not_locked(xlsx_path: Path) -> None:
    lock = _excel_lock_path(xlsx_path)
    if lock.exists():
        raise SuiviFileLockedError(
            f"le fichier est ouvert dans Excel (verrou : {lock.name}). "
            "Ferme-le puis réessaie."
        )


def _cell_to_native(v: Any) -> Any:
    """Normalise une valeur pour comparaison (evite d'écrire si déjà égale)."""
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def _equal(a: Any, b: Any) -> bool:
    a = _cell_to_native(a)
    b = _cell_to_native(b)
    if isinstance(a, datetime) and isinstance(b, datetime):
        return a.replace(microsecond=0) == b.replace(microsecond=0)
    return a == b


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------

class SuiviWriter:
    """Gestionnaire d'écriture pour un fichier Suivi."""

    def __init__(self, path: Path, backup_dir: Path | None = None, keep_backups: int = 20):
        self.path = Path(path).resolve()
        if not self.path.is_file():
            raise FileNotFoundError(self.path)
        self.backup_dir = backup_dir or (self.path.parent / "_backups")
        self.keep_backups = keep_backups
        self.year: int | None = None
        self._header_to_col: dict[str, int] = {}
        self._main_sheet: str | None = None
        self._load_headers()

    # -- introspection du schéma --------------------------------------------

    def _load_headers(self) -> None:
        check_not_locked(self.path)
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        # Feuille = année si possible
        for sn in wb.sheetnames:
            if sn.isdigit() and len(sn) == 4:
                self._main_sheet = sn
                self.year = int(sn)
                break
        if self._main_sheet is None:
            self._main_sheet = wb.sheetnames[0]
        ws = wb[self._main_sheet]
        header_row = next(ws.iter_rows(values_only=True), None)
        if header_row is None:
            raise RuntimeError("feuille vide")
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace("\n", " ").replace("  ", " ")
            if "températures" in key:
                self._header_to_col["temperatures"] = i
            elif key in HEADER_MAP:
                self._header_to_col[HEADER_MAP[key]] = i
        wb.close()

    def has_column(self, logical_key: str) -> bool:
        return logical_key in self._header_to_col

    # -- backup --------------------------------------------------------------

    def _backup(self) -> Path:
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = self.backup_dir / f"{self.path.stem}_backup_{ts}{self.path.suffix}"
        shutil.copy2(self.path, dst)
        self._rotate_backups()
        return dst

    def _rotate_backups(self) -> None:
        pattern = f"{self.path.stem}_backup_*{self.path.suffix}"
        backups = sorted(self.backup_dir.glob(pattern), key=lambda p: p.stat().st_mtime)
        excess = len(backups) - self.keep_backups
        for p in backups[:max(0, excess)]:
            try:
                p.unlink()
            except OSError:
                pass

    # -- écriture atomique ---------------------------------------------------

    def _save_atomic(self, wb) -> None:
        """Écrit dans un .tmp puis replace() atomique.

        Sous Windows, replace() peut échouer par PermissionError si un antivirus
        (Sophos, Defender, CrowdStrike…) ou un client de synchro (OneDrive)
        tient un handle transitoire sur le fichier cible. On retry 3× avec
        backoff court pour passer ces fenêtres de blocage sans faire planter
        l'utilisateur.
        """
        import time
        tmp = self.path.with_suffix(self.path.suffix + ".tmp")
        wb.save(str(tmp))
        last_exc: Exception | None = None
        for attempt in range(3):
            try:
                tmp.replace(self.path)
                return
            except PermissionError as e:
                last_exc = e
                time.sleep(0.2 * (attempt + 1))
        # Dernier recours : nettoyer le .tmp puis propager pour que l'utilisateur
        # voie un message actionable (plutôt que laisser un .tmp orphelin).
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        if last_exc is not None:
            raise last_exc

    # -- recherche d'une ligne -----------------------------------------------

    def _find_row(self, wb, key: RowKey) -> int | None:
        """Retourne l'index de ligne (1-indexé) matchant la clé, None sinon.
        Lève SuiviAmbiguousKeyError si plusieurs matchs."""
        ws = wb[self._main_sheet]
        matches: list[int] = []

        def _cell(row_idx: int, logical: str) -> Any:
            col = self._header_to_col.get(logical)
            if col is None:
                return None
            return ws.cell(row=row_idx, column=col + 1).value

        for r in range(2, ws.max_row + 1):
            if key.nom_contrat is not None:
                v = _cell(r, "nom_contrat")
                if not v or key.nom_contrat.strip().lower() not in str(v).strip().lower():
                    continue
            if key.date_debut is not None:
                v = _cell(r, "date_debut")
                if not isinstance(v, datetime):
                    continue
                if v.date() != key.date_debut.date():
                    continue
            if key.n_site_tadarida is not None:
                v = _cell(r, "n_site_tadarida")
                if _normalize_site_tadarida(v) != key.n_site_tadarida:
                    continue
            if key.n_point_fixe is not None:
                v = _cell(r, "n_point_fixe")
                if not v or str(v).strip().upper() != key.n_point_fixe.upper():
                    continue
            if key.n_passage is not None:
                v = _cell(r, "n_passage")
                if _normalize_passage(v) != key.n_passage:
                    continue
            if key.n_enregistreur is not None:
                v = _cell(r, "n_enregistreur")
                try:
                    if int(v) != int(key.n_enregistreur):
                        continue
                except (TypeError, ValueError):
                    continue
            matches.append(r)

        if not matches:
            return None
        if len(matches) > 1:
            raise SuiviAmbiguousKeyError(
                f"clé ambiguë : {len(matches)} lignes matchent "
                f"(lignes {matches})"
            )
        return matches[0]

    # -- API haut niveau -----------------------------------------------------

    def find_row(self, key: RowKey) -> int | None:
        """Cherche (lecture seule), retourne l'index de ligne ou None."""
        check_not_locked(self.path)
        wb = openpyxl.load_workbook(self.path)
        try:
            return self._find_row(wb, key)
        finally:
            wb.close()

    def update_row(
        self,
        key: RowKey,
        updates: dict[str, Any],
        *,
        create_if_missing: bool = False,
        backup: bool = True,
    ) -> dict:
        """
        Met à jour une ligne en place. ``updates`` est un dict
        {logical_key: valeur} (ex : {"etat": "ok", "flag_lupas_kaleidoscope": True}).

        Si ``create_if_missing`` et aucune ligne ne matche : une ligne est
        ajoutée à la fin, en remplissant à la fois la clé et les updates.
        """
        check_not_locked(self.path)

        # vérif que les colonnes demandées existent
        unknown = [k for k in updates if k not in self._header_to_col]
        if unknown:
            raise KeyError(f"colonnes inconnues dans le Suivi : {unknown}")

        wb = openpyxl.load_workbook(self.path)
        try:
            ws = wb[self._main_sheet]
            row_idx = self._find_row(wb, key)

            if row_idx is None and not create_if_missing:
                raise SuiviRowNotFoundError(f"aucune ligne ne matche {key}")

            if row_idx is None:
                # Création d'une nouvelle ligne
                row_idx = ws.max_row + 1
                # Remplir la clé (seulement les champs non-None)
                key_fields = {
                    "nom_contrat": key.nom_contrat,
                    "date_debut": key.date_debut,
                    "n_site_tadarida": key.n_site_tadarida,
                    "n_point_fixe": key.n_point_fixe,
                    "n_passage": key.n_passage,
                    "n_enregistreur": key.n_enregistreur,
                }
                for logical, v in key_fields.items():
                    if v is None:
                        continue
                    col = self._header_to_col.get(logical)
                    if col is not None:
                        ws.cell(row=row_idx, column=col + 1).value = v
                created = True
            else:
                created = False

            # Appliquer les updates (seulement ceux qui changent)
            changed: list[str] = []
            for logical, new_val in updates.items():
                col = self._header_to_col[logical]
                cell = ws.cell(row=row_idx, column=col + 1)
                if _equal(cell.value, new_val):
                    continue
                cell.value = new_val
                changed.append(logical)

            if not changed and not created:
                wb.close()
                return {
                    "action": "noop",
                    "row_index": row_idx,
                    "changed": [],
                    "created": False,
                    "backup": None,
                }

            backup_path = str(self._backup()) if backup else None
            self._save_atomic(wb)

            return {
                "action": "created" if created else "updated",
                "row_index": row_idx,
                "changed": changed,
                "created": created,
                "backup": backup_path,
            }
        finally:
            wb.close()

    # -- raccourcis sémantiques ---------------------------------------------

    # Mapping : nom logique d'étape → colonne dans le Suivi
    STEP_COL = {
        "lupas_kaleidoscope": "flag_lupas_kaleidoscope",
        "tadarida": "flag_tadarida",
    }

    def mark_step_done(self, key: RowKey, step: str, done: bool = True, **kw) -> dict:
        col = self.STEP_COL.get(step)
        if col is None:
            raise KeyError(f"étape inconnue : {step!r} (attendu : {list(self.STEP_COL)})")
        return self.update_row(key, {col: bool(done)}, **kw)

    def set_state(self, key: RowKey, etat: str | None, commentaire: str | None = None, **kw) -> dict:
        updates: dict[str, Any] = {}
        if etat is not None:
            updates["etat"] = etat
        if commentaire is not None and "commentaires" in self._header_to_col:
            updates["commentaires"] = commentaire
        if not updates:
            return {"action": "noop", "changed": []}
        return self.update_row(key, updates, **kw)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", type=Path, default=None, help="chemin xlsx (défaut : auto)")
    ap.add_argument("--year", type=int, default=None)
    ap.add_argument("--no-backup", action="store_true")
    ap.add_argument("--create-if-missing", action="store_true")

    ap.add_argument("--contrat", required=True)
    ap.add_argument("--date", required=True, help="YYYY-MM-DD")
    ap.add_argument("--site", required=True)
    ap.add_argument("--point", required=True)
    ap.add_argument("--pass", dest="pass_num", type=int, required=True)
    ap.add_argument("--enr", type=int, required=True)

    g = ap.add_argument_group("updates")
    g.add_argument("--set-lupas-kaleidoscope", action="store_true")
    g.add_argument("--set-tadarida", action="store_true")
    g.add_argument("--state", help="valeur colonne Etat")
    g.add_argument("--comment", help="valeur colonne Commentaires")
    g.add_argument("--serie", help="valeur colonne N° Série")
    g.add_argument("--micro", help="valeur colonne N° Micro")
    g.add_argument("--temperatures", help="'début-fin', ex: '28-19'")
    g.add_argument("--couverture", help="'0-25' / '25-50' / '50-75' / '75-100'")
    g.add_argument("--vent", help="'nul' / 'faible' / 'moyen' / 'fort'")

    args = ap.parse_args()

    # résolution chemin xlsx
    if args.file:
        path = args.file
    else:
        from suivi import _default_path
        path = _default_path(args.year)
    if not path.is_file():
        print(f"❌ introuvable : {path}", file=sys.stderr)
        return 2

    date = datetime.strptime(args.date, "%Y-%m-%d")
    key = RowKey(
        nom_contrat=args.contrat,
        date_debut=date,
        n_site_tadarida=_normalize_site_tadarida(args.site),
        n_point_fixe=args.point,
        n_passage=args.pass_num,
        n_enregistreur=args.enr,
    )

    updates: dict[str, Any] = {}
    if args.set_lupas_kaleidoscope:
        updates["flag_lupas_kaleidoscope"] = True
    if args.set_tadarida:
        updates["flag_tadarida"] = True
    if args.state:
        updates["etat"] = args.state
    if args.comment:
        updates["commentaires"] = args.comment
    if args.serie:
        updates["n_serie"] = args.serie
    if args.micro:
        updates["n_micro"] = args.micro
    if args.temperatures:
        updates["temperatures"] = args.temperatures
    if args.couverture:
        updates["couverture"] = args.couverture
    if args.vent:
        updates["vent"] = args.vent

    try:
        w = SuiviWriter(path)
    except SuiviFileLockedError as e:
        print(f"❌ {e}", file=sys.stderr)
        return 1

    print(f"📂 {path}   (année {w.year})")

    try:
        res = w.update_row(
            key,
            updates,
            create_if_missing=args.create_if_missing,
            backup=not args.no_backup,
        )
    except (SuiviRowNotFoundError, SuiviAmbiguousKeyError, KeyError) as e:
        print(f"❌ {e}", file=sys.stderr)
        return 1

    print(f"\nAction   : {res['action']}")
    print(f"Ligne    : {res.get('row_index')}")
    print(f"Changées : {res.get('changed')}")
    if res.get("backup"):
        print(f"Backup   : {Path(res['backup']).name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
