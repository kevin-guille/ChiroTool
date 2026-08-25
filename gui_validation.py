"""
gui_validation.py — Vue de validation d'une nuit d'enregistrement.

Ouvre une fenêtre modale dédiée pour vérifier les contacts détectés par
Tadarida sur une session : tableau filtrable, édition inline, ouverture
directe dans ChiroSurf pour validation acoustique.

Fonctionnalités :
  - Chargement de ``participation-*-observations*.xlsx`` de la session
  - Tableau (ttk.Treeview) avec colonnes : heure, durée, fréq, taxon Tadarida,
    proba, taxon observateur, confiance
  - Filtres : par taxon/groupe, par proba min, seulement non-validés,
    taxon observateur renseigné, seulement patrimoniaux
  - Tri au clic sur les en-têtes (asc / desc / ordre d'origine)
  - Édition inline du contact sélectionné (taxon + confiance)
  - Bouton "▶ ChiroSurf" par ligne (double-clic aussi)
  - Raccourcis clavier :
      ↓/↑   : contact suivant / précédent
      O     : confiance POSSIBLE + taxon = tadarida → suivant
      P     : confiance PROBABLE + taxon = tadarida → suivant
      S     : confiance SUR + taxon = tadarida → suivant
      Espace: relance ChiroSurf sur le contact courant
      Suppr : efface la validation du contact courant
      Ctrl+S: enregistre
  - Sauvegarde dans une copie ``*_<initiales>.xlsx`` pour préserver l'original
"""

from __future__ import annotations

import os
import subprocess
import sys
import threading
from pathlib import Path
from tkinter import messagebox, ttk
from typing import Any

import customtkinter as ctk

from credentials import load_token
from gui_config import Settings, save_settings
from manifest import Manifest
from gui_autocomplete import AutocompleteEntry
from gui_tooltip import TreeCellTooltip
from sync_state import (
    build_row_key_map, is_sendable, next_sync_state, sync_label,
    SYNC_ERROR, SYNC_MODIFIED, SYNC_PENDING, SYNC_SYNCED, SYNC_TO_RETRACT,
)
from taxon_index import genus_code_for, is_known, load_taxon_index, suggest
from taxons import classify_taxon
from vigiechiro_api import (
    ApiError, VigieChiroClient, entries_from_donnees,
    load_observation_sidecar, save_observation_sidecar,
)


# Espèces patrimoniales (à étendre selon besoin client / région)
PATRIMONIAL_CODES = {
    "barbar",   # Barbastelle
    "rhifer", "rhihip", "rhieur", "rhimeh",  # Rhinolophes
    "minsch",   # Minioptère
    "myogt", "myocap", "myobec", "myobly", "myoalc",  # Myotis patrimoniaux
    "nyclas",   # Noctule géante
    "plemac",   # Oreillard montagnard
    "tadten",   # Molosse de Cestoni
}

CONFIDENCE_VALUES = ("POSSIBLE", "PROBABLE", "SUR")


# ---------------------------------------------------------------------------
# Logique de filtrage PURE (sans GUI) — testable unitairement
# ---------------------------------------------------------------------------

def row_passes_filters(row: list, col_idx: dict, *, taxon_filter: str = "Tous",
                       proba_min: float = 0.0, hide_cleaned: bool = False,
                       wav_present: bool = True, only_unvalidated: bool = False,
                       only_observer_taxon: bool = False,
                       only_patrimonial: bool = False,
                       patrimonial_codes=frozenset()) -> bool:
    """La ligne d'observations passe-t-elle les filtres actifs ?

    Fonction PURE (aucune dépendance GUI) → testable. ``wav_present`` indique si
    le WAV existe encore (pour l'option « masquer les sons supprimés »).
    ``only_observer_taxon`` : ne garder que les lignes avec taxon observateur
    renseigné (issue #4). Incompatible avec ``only_unvalidated`` : si les deux
    sont True, rien ne passe.
    """
    ci = col_idx

    def _get(key):
        j = ci.get(key)
        return row[j] if (j is not None and j < len(row)) else None

    tad = _get("tadarida_taxon")
    proba = _get("tadarida_probabilite")
    obs = _get("observateur_taxon")

    if hide_cleaned and not wav_present:
        return False
    if taxon_filter and taxon_filter != "Tous" and str(tad) != taxon_filter:
        return False
    if proba is not None:
        try:
            if float(proba) < float(proba_min):
                return False
        except (TypeError, ValueError):
            pass
    if only_unvalidated and obs:
        return False
    if only_observer_taxon and not obs:
        return False
    if only_patrimonial and (str(tad) or "").lower() not in patrimonial_codes:
        return False
    return True


def eligible_taxons(rows: list, col_idx: dict, *, proba_min: float = 0.0,
                    hide_cleaned: bool = False, wav_present: list | None = None,
                    only_unvalidated: bool = False,
                    only_observer_taxon: bool = False,
                    only_patrimonial: bool = False,
                    patrimonial_codes=frozenset()) -> list[str]:
    """Liste triée des taxons Tadarida présents parmi les lignes qui passent les
    filtres AUTRES que le taxon (pour peupler dynamiquement le menu). Pur/testable.

    ``wav_present`` : liste de bool par ligne (None → toutes présentes).
    """
    ci = col_idx
    if "tadarida_taxon" not in ci:
        return []
    taxons: set[str] = set()
    for i, r in enumerate(rows):
        wp = True if wav_present is None else (
            wav_present[i] if i < len(wav_present) else True)
        if not row_passes_filters(
            r, ci, taxon_filter="Tous", proba_min=proba_min,
            hide_cleaned=hide_cleaned, wav_present=wp,
            only_unvalidated=only_unvalidated,
            only_observer_taxon=only_observer_taxon,
            only_patrimonial=only_patrimonial,
            patrimonial_codes=patrimonial_codes,
        ):
            continue
        j = ci["tadarida_taxon"]
        t = r[j] if j < len(r) else None
        if t:
            taxons.add(str(t))
    return sorted(taxons, key=str.lower)


HEADING_LABELS = {
    "heure": "Heure",
    "duree": "Durée (s)",
    "freq": "Fréq. méd. (kHz)",
    "tad_taxon": "Taxon Tadarida",
    "tad_proba": "Proba",
    "obs_taxon": "Taxon observateur",
    "obs_conf": "Confiance",
    "groupe": "Groupe",
}


def heure_from_filename(filename: str) -> str | None:
    """HH:MM:SS extrait du bloc ``_HHMMSS`` du nom de fichier, ou None."""
    try:
        for p in str(filename or "").split("_"):
            if len(p) == 6 and p.isdigit():
                return f"{p[:2]}:{p[2:4]}:{p[4:]}"
    except Exception:
        return None
    return None


def contact_sort_value(row: list, col_idx: dict, column: str):
    """Valeur comparable pour une colonne d'affichage. ``None`` si vide/illisible.

    PURE → testable. Types : str (heure, taxons, groupe, confiance) ou float
    (durée, freq, proba).
    """
    ci = col_idx

    def _get(key):
        j = ci.get(key)
        return row[j] if (j is not None and j < len(row)) else None

    def _f(v):
        if v in (None, ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    if column == "heure":
        return heure_from_filename(_get("nom du fichier") or "")
    if column == "duree":
        a, b = _f(_get("temps_debut")), _f(_get("temps_fin"))
        if a is None or b is None:
            return None
        return b - a
    if column == "freq":
        return _f(_get("frequence_mediane"))
    if column == "tad_taxon":
        v = _get("tadarida_taxon")
        return str(v).strip().lower() if v not in (None, "") else None
    if column == "tad_proba":
        return _f(_get("tadarida_probabilite"))
    if column == "obs_taxon":
        v = _get("observateur_taxon")
        return str(v).strip().lower() if v not in (None, "") else None
    if column == "obs_conf":
        v = _get("observateur_probabilite")
        return str(v).strip().lower() if v not in (None, "") else None
    if column == "groupe":
        tad = _get("tadarida_taxon")
        return classify_taxon(str(tad)) if tad else None
    return None


def sort_filtered_indexes(indexes: list[int], rows: list, col_idx: dict,
                          column: str, *, descending: bool = False) -> list[int]:
    """Trie les index filtrés. Les valeurs manquantes restent en fin. PURE."""
    present: list[tuple[int, object]] = []
    missing: list[int] = []
    for i in indexes:
        if i < 0 or i >= len(rows):
            missing.append(i)
            continue
        v = contact_sort_value(rows[i], col_idx, column)
        if v is None or v == "":
            missing.append(i)
        else:
            present.append((i, v))
    present.sort(key=lambda t: t[1], reverse=descending)
    return [i for i, _ in present] + missing


def count_observer_progress(headers: list, rows: list) -> dict:
    """``n_validated`` / ``n_total`` d'après ``observateur_taxon``. PURE.

    ``n_total`` = lignes non vides. Une ligne est validée si la colonne
    observateur est renseignée (même logique que le pied de Valider).
    """
    lower = [str(h).lower().strip() for h in (headers or [])]
    i_obs = lower.index("observateur_taxon") if "observateur_taxon" in lower else None
    n_total = 0
    n_val = 0
    for r in rows or []:
        if not r or all(c in (None, "") for c in r):
            continue
        n_total += 1
        if i_obs is not None and i_obs < len(r) and r[i_obs] not in (None, ""):
            n_val += 1
    return {"n_validated": n_val, "n_total": n_total}


class ValidationView(ctk.CTkToplevel):
    """Fenêtre de validation d'une nuit."""

    def __init__(self, master, *, session_path: Path,
                 xlsx_path: Path,
                 chirosurf_path: str | None = None,
                 initiales: str = "XX"):
        super().__init__(master)
        self.session_path = Path(session_path)
        self.xlsx_path = Path(xlsx_path)
        self.chirosurf_path = chirosurf_path
        self.initiales = initiales

        self.title(f"Validation — {self.session_path.name}")
        self.geometry("1280x780")
        self.minsize(1100, 640)

        self.transient(master)
        self.after(50, self.grab_set)
        self.focus()

        # Data
        self.headers: list[str] = []
        self.rows: list[list] = []
        self.filtered_indexes: list[int] = []
        self._wav_present: list[bool] = []   # par ligne : le WAV existe-t-il encore ?
        self._dirty = False

        # Index des taxons acceptés par le serveur (autocomplétion + validation
        # de la saisie ; {} si snapshot absent → saisie libre comme avant).
        self._taxon_index = load_taxon_index()

        # Synchro serveur (envoi des identifications) — cf. sync_state.py
        self._sidecar_entries: list[dict] = []      # mapping natif (donnee_id, index)
        self.row_keys: dict[int, str] = {}          # row_idx → "donnee_id#index"
        self.sync_state: dict[str, dict] = {}       # key → {state, pushed_taxon, ...}
        self.mappable = False                       # au moins une ligne mappée serveur ?
        self._participation_id = None               # pour reconstruire le mapping si besoin
        self._dot_imgs: dict[str, Any] = {}         # PhotoImage par état (anti-GC)
        self._pushing = False

        # Mapping index colonne -> clé logique
        self.col_idx: dict[str, int] = {}
        self._sort_col: str | None = None
        self._sort_desc = False

        self._build_ui()
        self.after(100, self._load_xlsx)

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # -- UI -----------------------------------------------------------------

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # --- Header ---
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
        header.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(
            header, text=f"📋  {self.session_path.name}",
            font=ctk.CTkFont(size=16, weight="bold"), anchor="w",
        )
        title.grid(row=0, column=0, sticky="w")

        self.header_info = ctk.CTkLabel(
            header, text="chargement…", anchor="e",
            text_color=("gray40", "gray70"),
            font=ctk.CTkFont(size=11),
        )
        self.header_info.grid(row=0, column=1, sticky="e")

        # --- Filtres ---
        filters = ctk.CTkFrame(self, fg_color=("gray92", "gray20"),
                                 corner_radius=8, border_width=1,
                                 border_color=("gray80", "gray25"))
        filters.grid(row=1, column=0, sticky="ew", padx=12, pady=4)
        for c in range(8):
            filters.grid_columnconfigure(c, weight=0)

        ctk.CTkLabel(filters, text="Taxon :",
                      font=ctk.CTkFont(size=11)).grid(row=0, column=0, padx=(10, 4), pady=8)
        self.taxon_filter_var = ctk.StringVar(value="Tous")
        self.taxon_menu = ctk.CTkOptionMenu(
            filters, variable=self.taxon_filter_var, width=180, height=28,
            values=["Tous"], command=lambda _: self._apply_filters(),
        )
        self.taxon_menu.grid(row=0, column=1, padx=(0, 10), pady=8)

        ctk.CTkLabel(filters, text="Proba ≥",
                      font=ctk.CTkFont(size=11)).grid(row=0, column=2, padx=(0, 4), pady=8)
        self.proba_var = ctk.DoubleVar(value=0.0)
        self.proba_slider = ctk.CTkSlider(
            filters, from_=0, to=1, width=160, number_of_steps=20,
            command=self._on_proba_change,
        )
        self.proba_slider.set(0.0)
        self.proba_slider.grid(row=0, column=3, padx=(0, 4), pady=8)
        self.proba_lbl = ctk.CTkLabel(
            filters, text="0.00", width=40,
            font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
        )
        self.proba_lbl.grid(row=0, column=4, padx=(0, 16), pady=8)

        self.only_unvalidated_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            filters, text="Non validés seulement",
            variable=self.only_unvalidated_var,
            command=self._on_unvalidated_toggle,
        ).grid(row=0, column=5, padx=(0, 12), pady=8)

        self.only_patrimonial_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            filters, text="Patrimoniaux",
            variable=self.only_patrimonial_var,
            command=self._on_filters_changed,
        ).grid(row=0, column=6, padx=(0, 12), pady=8)

        ctk.CTkButton(
            filters, text="Réinitialiser", width=110, height=28,
            fg_color=("gray85", "gray25"),
            text_color=("gray15", "gray90"),
            hover_color=("gray75", "gray35"),
            command=self._reset_filters,
        ).grid(row=0, column=7, padx=(0, 10), pady=8)

        # 2e ligne de filtres : masquer les contacts dont le WAV a été supprimé
        # au nettoyage. Activée seulement si un nettoyage a manifestement eu lieu
        # (voir _compute_wav_presence).
        self.only_observer_taxon_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            filters, text="Taxon observateur renseigné uniquement",
            variable=self.only_observer_taxon_var,
            command=self._on_observer_filter_toggle,
        ).grid(row=1, column=0, columnspan=3,
               padx=(10, 12), pady=(0, 8), sticky="w")

        self.hide_cleaned_var = ctk.BooleanVar(value=False)
        self.hide_cleaned_chk = ctk.CTkCheckBox(
            filters, text="Masquer les sons supprimés au nettoyage",
            variable=self.hide_cleaned_var, command=self._apply_filters,
        )
        self.hide_cleaned_chk.grid(row=1, column=3, columnspan=5,
                                    padx=(10, 12), pady=(0, 8), sticky="w")
        self.hide_cleaned_chk.configure(state="disabled")

        # --- Tableau (ttk.Treeview) ---
        table_frame = ctk.CTkFrame(self)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=4)
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        columns = ("heure", "duree", "freq", "tad_taxon", "tad_proba",
                   "obs_taxon", "obs_conf", "groupe")
        # ``tree headings`` : on affiche la colonne d'arbre #0 comme une fine
        # gouttière de statut de synchro (une petite pastille par ligne), sans
        # toucher aux colonnes de données.
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="tree headings", selectmode="extended",
        )
        # #0 = gouttière de statut. Tkinter réserve un retrait interne (zone
        # d'indicateur d'arbre) AVANT l'image : à 22 px la pastille de 14 px était
        # rognée. On élargit pour la dégager entièrement, colonne non étirable.
        self.tree.column("#0", width=42, minwidth=42, stretch=False, anchor="center")
        self.tree.heading("#0", text="⬆")
        self._build_dot_images()
        for col, label in HEADING_LABELS.items():
            self.tree.heading(
                col, text=label,
                command=lambda c=col: self._on_sort_heading(c),
            )
        self.tree.column("heure", width=80, anchor="center")
        self.tree.column("duree", width=70, anchor="center")
        self.tree.column("freq", width=120, anchor="center")
        self.tree.column("tad_taxon", width=140, anchor="w")
        self.tree.column("tad_proba", width=60, anchor="center")
        self.tree.column("obs_taxon", width=160, anchor="w")
        self.tree.column("obs_conf", width=100, anchor="center")
        self.tree.column("groupe", width=100, anchor="w")

        # Couleurs lignes (tags)
        self.tree.tag_configure("validated", background="#e8f5ec")
        self.tree.tag_configure("patrimonial", background="#fdf3e6")
        self.tree.tag_configure("noise", foreground="#999")

        self.tree.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.bind("<<TreeviewSelect>>", lambda e: self._on_row_select())
        self.tree.bind("<Double-1>", lambda e: self._open_in_chirosurf())
        # Infobulle au survol de la gouttière de synchro (#0)
        TreeCellTooltip(self.tree, self._sync_tooltip)

        # --- Édition ---
        edit = ctk.CTkFrame(self, fg_color=("gray92", "gray20"),
                              corner_radius=8, border_width=1,
                              border_color=("gray80", "gray25"))
        edit.grid(row=3, column=0, sticky="ew", padx=12, pady=(4, 4))
        edit.grid_columnconfigure(7, weight=1)

        ctk.CTkLabel(edit, text="Sélection :",
                      font=ctk.CTkFont(size=11, weight="bold")).grid(
            row=0, column=0, padx=(10, 6), pady=8)

        self.sel_info_lbl = ctk.CTkLabel(
            edit, text="(aucune ligne sélectionnée)",
            font=ctk.CTkFont(size=11),
            text_color=("gray40", "gray70"), anchor="w",
        )
        self.sel_info_lbl.grid(row=0, column=1, columnspan=2, padx=(0, 16),
                                 pady=8, sticky="w")

        ctk.CTkLabel(edit, text="Taxon :",
                      font=ctk.CTkFont(size=11)).grid(row=1, column=0, padx=(10, 6), pady=8)
        self.edit_taxon_var = ctk.StringVar(value="")
        # Autocomplétion sur les taxons acceptés par le serveur (si snapshot chargé),
        # sinon champ libre classique. La sélection insère le code.
        self.edit_taxon_entry = AutocompleteEntry(
            edit, textvariable=self.edit_taxon_var, width=160,
            font=ctk.CTkFont(family="Consolas", size=12),
            suggest_fn=(lambda q: suggest(q, self._taxon_index)
                        if self._taxon_index else []),
            on_pick=lambda code: self._update_taxon_indicator(),
        )
        self.edit_taxon_entry.grid(row=1, column=1, padx=(0, 6), pady=8)
        self.edit_taxon_entry.bind("<Return>", lambda e: self._apply_edit_and_next())
        # Indicateur de validité serveur (✓ connu / ⚠ inconnu → ne partira pas)
        self.taxon_ind = ctk.CTkLabel(edit, text="", width=90, anchor="w",
                                      font=ctk.CTkFont(size=11))
        self.taxon_ind.grid(row=2, column=1, sticky="w", padx=(0, 6))
        self.edit_taxon_var.trace_add("write", lambda *a: self._update_taxon_indicator())

        ctk.CTkLabel(edit, text="Confiance :",
                      font=ctk.CTkFont(size=11)).grid(row=1, column=2, padx=(16, 6), pady=8)
        self.edit_conf_var = ctk.StringVar(value="")
        for i, val in enumerate(CONFIDENCE_VALUES):
            ctk.CTkRadioButton(
                edit, text=val.capitalize(), variable=self.edit_conf_var,
                value=val,
            ).grid(row=1, column=3 + i, padx=(0, 8), pady=8)

        ctk.CTkButton(
            edit, text="Appliquer + suivant", width=150, height=30,
            font=ctk.CTkFont(weight="bold"),
            command=self._apply_edit_and_next,
        ).grid(row=1, column=6, padx=(16, 8), pady=8)

        ctk.CTkButton(
            edit, text="▶ ChiroSurf", width=120, height=30,
            command=self._open_in_chirosurf,
        ).grid(row=1, column=7, padx=(0, 10), pady=8, sticky="e")

        # « Monter au genre » : son incertain entre 2 espèces d'un même genre
        # (ex. Pleaur/Pleaus → Plesp « Oreillard sp. »). Confiance courante conservée.
        if self._taxon_index:
            ctk.CTkButton(
                edit, text="↑ Monter au genre (incertain)", width=220, height=28,
                fg_color=("gray85", "gray25"), text_color=("gray15", "gray90"),
                hover_color=("gray75", "gray35"), command=self._raise_to_genus,
            ).grid(row=2, column=3, columnspan=4, sticky="w", padx=(0, 8), pady=(0, 6))

        # --- Footer ---
        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.grid(row=4, column=0, sticky="ew", padx=12, pady=(4, 12))
        footer.grid_columnconfigure(0, weight=1)

        self.stats_lbl = ctk.CTkLabel(
            footer, text="", anchor="w",
            font=ctk.CTkFont(size=11),
        )
        self.stats_lbl.grid(row=0, column=0, sticky="w")

        # Bouton d'envoi des identifications au serveur. Contextuel : selon la
        # sélection, il envoie « la sélection » ou « tout » (lignes envoyables).
        self.send_btn = ctk.CTkButton(
            footer, text="⬆ Envoyer", width=190, height=32,
            font=ctk.CTkFont(weight="bold"),
            command=self._on_send_click,
        )
        self.send_btn.grid(row=0, column=1, padx=(6, 6))

        ctk.CTkButton(
            footer, text="🔄 Rafraîchir", width=110, height=32,
            fg_color=("gray85", "gray25"),
            text_color=("gray15", "gray90"),
            hover_color=("gray75", "gray35"),
            command=self._on_filters_changed,
        ).grid(row=0, column=2, padx=(6, 6))

        self.save_btn = ctk.CTkButton(
            footer, text="Enregistrer", width=130, height=32,
            font=ctk.CTkFont(weight="bold"),
            command=self._save,
        )
        self.save_btn.grid(row=0, column=3, padx=(6, 6))
        self._save_btn_fg = self.save_btn.cget("fg_color")   # couleur par défaut

        ctk.CTkButton(
            footer, text="Fermer", width=100, height=32,
            fg_color=("gray85", "gray25"),
            text_color=("gray15", "gray90"),
            hover_color=("gray75", "gray35"),
            command=self._on_close,
        ).grid(row=0, column=4)

        # --- Raccourcis clavier ---
        # CRITIQUE : les binds niveau Toplevel captureraient la frappe aussi
        # quand le focus est dans un CTkEntry (saisie taxon libre), ce qui
        # validerait accidentellement des lignes en tapant "obs" dans le champ.
        # On filtre via _is_typing_in_entry() en tête de chaque handler.
        self.bind("<Down>", self._on_key_down)
        self.bind("<Up>", self._on_key_up)
        self.bind("<space>", self._on_key_space)
        self.bind("<Delete>", self._on_key_delete)
        self.bind("<Control-s>", lambda e: self._save())
        self.bind("<KeyPress-o>", self._on_key_opsr("POSSIBLE"))
        self.bind("<KeyPress-p>", self._on_key_opsr("PROBABLE"))
        self.bind("<KeyPress-s>", self._on_key_opsr("SUR"))

    def _is_typing_in_entry(self) -> bool:
        """True si le focus est dans un CTkEntry / Entry / Text / Textbox."""
        try:
            w = self.focus_get()
        except Exception:
            return False
        if w is None:
            return False
        cls = w.winfo_class()
        # Tk classes : Entry, Text, TEntry ; CTk wrappe mais expose des Entry internes
        if cls in ("Entry", "TEntry", "Text"):
            return True
        # CTkEntry : classe CTkEntry mais le vrai focus peut être sur le Entry interne
        name = type(w).__name__
        if "Entry" in name or "Text" in name:
            return True
        return False

    def _on_key_down(self, _e):
        if self._is_typing_in_entry():
            return
        self._move_selection(1)

    def _on_key_up(self, _e):
        if self._is_typing_in_entry():
            return
        self._move_selection(-1)

    def _on_key_space(self, _e):
        if self._is_typing_in_entry():
            return
        self._open_in_chirosurf()

    def _on_key_delete(self, _e):
        if self._is_typing_in_entry():
            return
        self._clear_validation()

    def _on_key_opsr(self, confidence: str):
        def handler(_e):
            if self._is_typing_in_entry():
                return  # laisse la frappe passer au champ
            self._quick_validate(confidence)
        return handler

    # -- Chargement xlsx ----------------------------------------------------

    # -- pastilles de synchro serveur --------------------------------------

    def _build_dot_images(self):
        """Construit une petite pastille (~14 px) par état de synchro. Gardées
        dans ``self._dot_imgs`` pour éviter la collecte GC de Tk (sinon l'image
        devient invisible). Dégradation propre si Pillow est absent."""
        self._dot_imgs = {}
        try:
            from PIL import Image, ImageDraw, ImageTk
        except Exception:
            return
        palette = {
            SYNC_PENDING: ("#b7bdc7", False, False),    # cercle creux gris
            SYNC_SYNCED: ("#16a34a", True, False),      # plein vert
            SYNC_MODIFIED: ("#e08a1e", True, False),    # plein orange
            SYNC_TO_RETRACT: ("#e08a1e", True, True),   # orange barré
            SYNC_ERROR: ("#dc2626", True, False),       # plein rouge
        }
        size, r = 14, 4
        c = size // 2
        for state, (col, fill, barred) in palette.items():
            img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
            d = ImageDraw.Draw(img)
            box = [c - r, c - r, c + r, c + r]
            if fill:
                d.ellipse(box, fill=col)
            else:
                d.ellipse(box, outline=col, width=1)
            if barred:
                d.line([c - r, c + r, c + r, c - r], fill="#ffffff", width=1)
            self._dot_imgs[state] = ImageTk.PhotoImage(img)

    def _row_state(self, idx: int) -> str | None:
        key = self.row_keys.get(idx)
        rec = self.sync_state.get(key) if key else None
        return rec.get("state") if rec else None

    def _sync_tooltip(self, row_id, col_name):
        """Infobulle de la gouttière de synchro (colonne #0)."""
        if col_name != "#0":
            return None
        try:
            idx = int(row_id)
        except (TypeError, ValueError):
            return None
        return sync_label(self._row_state(idx))

    def _row_dot_image(self, idx: int):
        """PhotoImage de la pastille d'une ligne (ou "" si aucune à afficher)."""
        state = self._row_state(idx)
        if not state:
            return ""
        return self._dot_imgs.get(state, "")

    def _update_row_dot(self, idx: int):
        iid = str(idx)
        if self.tree.exists(iid):
            self.tree.item(iid, image=self._row_dot_image(idx))

    def _row_taxon_conf(self, idx: int) -> tuple[str, str]:
        r = self.rows[idx]
        ci = self.col_idx
        taxon = r[ci["observateur_taxon"]] if "observateur_taxon" in ci else ""
        conf = r[ci["observateur_probabilite"]] if "observateur_probabilite" in ci else ""
        return (str(taxon or ""), str(conf or ""))

    def _refresh_sync_state(self, idx: int):
        """Recalcule l'état de synchro d'une ligne après édition + MAJ pastille."""
        key = self.row_keys.get(idx)
        if key is None:
            return
        taxon, conf = self._row_taxon_conf(idx)
        rec = self.sync_state.get(key)
        new = next_sync_state(taxon, conf, rec)
        if new is None:
            self.sync_state.pop(key, None)
        else:
            rec = dict(rec or {})
            rec["state"] = new
            self.sync_state[key] = rec
        self._update_row_dot(idx)

    def _is_sendable(self, idx: int) -> bool:
        """Ligne envoyable = taxon + confiance renseignés. Le mapping serveur est
        résolu au moment de l'envoi (sidecar ou re-fetch de la participation)."""
        taxon, conf = self._row_taxon_conf(idx)
        return is_sendable(taxon, conf)

    def _load_xlsx(self):
        try:
            import warnings
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            import openpyxl
            wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
        except Exception as e:
            messagebox.showerror("Lecture impossible",
                                 f"Impossible d'ouvrir {self.xlsx_path.name} :\n{e}")
            self.destroy()
            return

        if not data:
            messagebox.showwarning("Fichier vide",
                                   "Le fichier xlsx est vide.")
            self.destroy()
            return

        self.headers = [str(h) if h is not None else "" for h in data[0]]
        self.rows = [list(r) for r in data[1:] if any(c is not None for c in r)]

        # Index des colonnes métier
        lower = [h.lower().strip() for h in self.headers]
        for wanted in ("nom du fichier", "temps_debut", "temps_fin",
                        "frequence_mediane", "tadarida_taxon",
                        "tadarida_probabilite", "observateur_taxon",
                        "observateur_probabilite"):
            if wanted in lower:
                self.col_idx[wanted] = lower.index(wanted)

        # Mapping serveur (sidecar) : associe chaque ligne à (donnee_id, index)
        # natif, et charge l'état de synchro persisté. Sans sidecar (xlsx reçu par
        # mail ou téléchargé avant cette version), la vue reste « non mappable » →
        # l'envoi sera désactivé proprement.
        side = load_observation_sidecar(self.xlsx_path)
        self._sidecar_entries = side.get("entries", [])
        self.sync_state = dict(side.get("sync", {}))
        self.row_keys, n_unmapped = build_row_key_map(
            self.rows, self.col_idx, self._sidecar_entries)
        self.mappable = bool(self.row_keys)

        # participation_id (manifest) : permet de reconstruire le mapping serveur
        # si le sidecar est absent (xlsx téléchargé avant cette version).
        try:
            m = Manifest.load(self.session_path)
            self._participation_id = (m.meta.get("vigiechiro_participation_id")
                                      if m else None)
        except Exception:
            self._participation_id = None

        header_txt = f"{len(self.rows)} contacts chargés"
        if self._sidecar_entries and n_unmapped and self.mappable:
            # Mapping partiel : quelques lignes non appariées (fichier vide, trou…)
            header_txt = f"{len(self.rows)} contacts · {n_unmapped} non mappés serveur"

        self._compute_wav_presence()
        self._on_filters_changed()   # peuple le menu taxons (filtré) + rafraîchit
        self.header_info.configure(text=header_txt)

    def _compute_wav_presence(self):
        """Détermine, par ligne, si le WAV correspondant existe encore sur disque.

        Un « son supprimé au nettoyage » = WAV absent de Data_k/ (ou Data/).
        La case « Masquer les sons supprimés » n'est activée QUE si un nettoyage a
        manifestement eu lieu (certains WAV présents, d'autres absents) — évite de
        tout masquer avant nettoyage, ou si le dossier a été entièrement effacé.
        """
        self._wav_present = [True] * len(self.rows)
        present_stems = set()
        for sub in ("Data_k", "Data"):
            d = self.session_path / sub
            if d.is_dir():
                for p in d.glob("*.wav"):
                    present_stems.add(p.stem.lower())

        n_present = n_absent = 0
        fi = self.col_idx.get("nom du fichier", 0)
        if present_stems:
            for i, r in enumerate(self.rows):
                name = str(r[fi] or "").strip()
                if name.lower().endswith(".wav"):
                    name = name[:-4]
                ok = bool(name) and name.lower() in present_stems
                self._wav_present[i] = ok
                n_present += ok
                n_absent += (not ok)

        try:
            if present_stems and n_present > 0 and n_absent > 0:
                self.hide_cleaned_chk.configure(state="normal")
            else:
                self.hide_cleaned_var.set(False)
                self.hide_cleaned_chk.configure(state="disabled")
        except Exception:
            pass

    # -- Filtres ------------------------------------------------------------

    def _on_proba_change(self, v):
        self.proba_lbl.configure(text=f"{float(v):.2f}")
        self._on_filters_changed()

    def _reset_filters(self):
        self.taxon_filter_var.set("Tous")
        self.proba_slider.set(0.0)
        self.proba_lbl.configure(text="0.00")
        self.only_unvalidated_var.set(False)
        self.only_observer_taxon_var.set(False)
        self.only_patrimonial_var.set(False)
        self.hide_cleaned_var.set(False)
        self._sort_col = None
        self._sort_desc = False
        self._refresh_heading_labels()
        self._on_filters_changed()

    def _on_unvalidated_toggle(self):
        if self.only_unvalidated_var.get():
            self.only_observer_taxon_var.set(False)
        self._on_filters_changed()

    def _on_observer_filter_toggle(self):
        if self.only_observer_taxon_var.get():
            self.only_unvalidated_var.set(False)
        self._on_filters_changed()

    def _on_sort_heading(self, col: str):
        """Clic en-tête : asc → desc → ordre d'origine."""
        if self._sort_col == col:
            if not self._sort_desc:
                self._sort_desc = True
            else:
                self._sort_col = None
                self._sort_desc = False
        else:
            self._sort_col = col
            self._sort_desc = False
        self._refresh_heading_labels()
        self._apply_filters()

    def _refresh_heading_labels(self):
        for col, label in HEADING_LABELS.items():
            mark = ""
            if col == self._sort_col:
                mark = " ▼" if self._sort_desc else " ▲"
            self.tree.heading(col, text=label + mark)

    def _filter_kwargs(self) -> dict:
        """Paramètres de filtrage courants (lus des widgets) pour les fonctions
        pures de filtrage."""
        return dict(
            proba_min=float(self.proba_slider.get()),
            hide_cleaned=self.hide_cleaned_var.get(),
            only_unvalidated=self.only_unvalidated_var.get(),
            only_observer_taxon=self.only_observer_taxon_var.get(),
            only_patrimonial=self.only_patrimonial_var.get(),
            patrimonial_codes=PATRIMONIAL_CODES,
        )

    def _passes_filters(self, i: int, r: list, *, ignore_taxon: bool = False) -> bool:
        wp = self._wav_present[i] if i < len(self._wav_present) else True
        return row_passes_filters(
            r, self.col_idx, wav_present=wp,
            taxon_filter="Tous" if ignore_taxon else self.taxon_filter_var.get(),
            **self._filter_kwargs(),
        )

    def _rebuild_taxon_menu(self):
        """Recalcule les taxons proposés : uniquement ceux présents parmi les
        lignes qui passent les AUTRES filtres (proba, masquage nettoyage, non
        validés, patrimoniaux). La liste reflète l'état courant des filtres."""
        taxons = eligible_taxons(
            self.rows, self.col_idx, wav_present=self._wav_present,
            **self._filter_kwargs(),
        )
        self.taxon_menu.configure(values=["Tous"] + taxons)
        # Conserve la sélection si toujours valide, sinon repli sur "Tous".
        cur = self.taxon_filter_var.get()
        if cur != "Tous" and cur not in taxons:
            self.taxon_filter_var.set("Tous")

    def _on_filters_changed(self):
        """Un filtre AUTRE que le taxon a changé : recalcule la liste des taxons
        proposés puis ré-applique les filtres."""
        self._rebuild_taxon_menu()
        self._apply_filters()

    def _apply_filters(self):
        self.filtered_indexes = [
            i for i, r in enumerate(self.rows) if self._passes_filters(i, r)
        ]
        if self._sort_col:
            self.filtered_indexes = sort_filtered_indexes(
                self.filtered_indexes, self.rows, self.col_idx,
                self._sort_col, descending=self._sort_desc,
            )
        self._refresh_table()

    def _refresh_table(self):
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        ci = self.col_idx
        for idx in self.filtered_indexes:
            r = self.rows[idx]
            filename = r[ci.get("nom du fichier", 0)] or ""
            tad = r[ci["tadarida_taxon"]] if "tadarida_taxon" in ci else ""
            tad_lower = (str(tad) or "").lower()

            t_deb = r[ci.get("temps_debut", 1)] if "temps_debut" in ci else ""
            t_fin = r[ci.get("temps_fin", 2)] if "temps_fin" in ci else ""
            freq = r[ci.get("frequence_mediane", 3)] if "frequence_mediane" in ci else ""
            tad_proba = r[ci["tadarida_probabilite"]] if "tadarida_probabilite" in ci else ""
            obs = r[ci["observateur_taxon"]] if "observateur_taxon" in ci else ""
            obs_conf = r[ci["observateur_probabilite"]] if "observateur_probabilite" in ci else ""

            # Heure extraite depuis le nom de fichier : ..._YYYYMMDD_HHMMSS_NNN
            heure = heure_from_filename(str(filename)) or "?"

            duree = ""
            try:
                if t_deb is not None and t_fin is not None:
                    duree = f"{float(t_fin) - float(t_deb):.2f}"
            except (TypeError, ValueError):
                pass

            freq_str = ""
            try:
                if freq is not None:
                    freq_str = f"{float(freq):.1f}"
            except (TypeError, ValueError):
                pass

            proba_str = ""
            try:
                if tad_proba is not None:
                    proba_str = f"{float(tad_proba):.2f}"
            except (TypeError, ValueError):
                pass

            groupe = classify_taxon(str(tad)) if tad else ""

            tags = []
            if obs:
                tags.append("validated")
            if tad_lower in PATRIMONIAL_CODES:
                tags.append("patrimonial")
            if groupe == "noise":
                tags.append("noise")

            self.tree.insert(
                "", "end", iid=str(idx),
                values=(heure, duree, freq_str, tad, proba_str,
                        obs or "", obs_conf or "", groupe),
                tags=tags,
                image=self._row_dot_image(idx),
            )

        self._refresh_stats()

    def _refresh_stats(self):
        ci = self.col_idx
        total = len(self.rows)
        validated = sum(
            1 for r in self.rows
            if "observateur_taxon" in ci and r[ci["observateur_taxon"]]
        )
        patri = sum(
            1 for r in self.rows
            if "tadarida_taxon" in ci
            and str(r[ci["tadarida_taxon"]] or "").lower() in PATRIMONIAL_CODES
        )
        patri_val = sum(
            1 for r in self.rows
            if "tadarida_taxon" in ci
            and str(r[ci["tadarida_taxon"]] or "").lower() in PATRIMONIAL_CODES
            and "observateur_taxon" in ci and r[ci["observateur_taxon"]]
        )
        # Garde contre total==0 (xlsx avec en-tête mais zéro contact)
        pct = (validated / total * 100) if total else 0.0
        # Compteur de synchro serveur : X envoyées / Y envoyables (mappées + taxon
        # + confiance). Y ne dépend PAS de la présence locale du WAV (une identif
        # sur un son purgé au nettoyage reste valide côté serveur).
        sync_txt = ""
        if self.mappable:
            sendable = [i for i in range(total) if self._is_sendable(i)]
            n_synced = sum(1 for i in sendable if self._row_state(i) == SYNC_SYNCED)
            sync_txt = f"  ·  ⬆ {n_synced}/{len(sendable)} identifications envoyées"
        self.stats_lbl.configure(text=(
            f"{len(self.filtered_indexes)}/{total} affichés  ·  "
            f"{validated} validés ({pct:.1f}%)  ·  "
            f"{patri} patrimoniaux ({patri_val} validés){sync_txt}"
        ))
        self._update_send_button()

    # -- Sélection / édition ------------------------------------------------

    def _on_row_select(self):
        self._update_send_button()       # « Envoyer la sélection (N) » / « tout »
        sel = self.tree.selection()
        if not sel:
            self.sel_info_lbl.configure(text="(aucune ligne sélectionnée)")
            return
        if len(sel) > 1:
            # Multi-sélection : on n'écrase PAS les champs d'édition (l'utilisateur
            # saisit UN taxon à appliquer à tout le lot) ; O/P/S reprend le taxon
            # Tadarida propre à chaque ligne.
            self.sel_info_lbl.configure(
                text=f"{len(sel)} lignes sélectionnées  ·  saisis un taxon puis "
                     f"« Appliquer », ou O/P/S (reprend le Tadarida de chaque ligne)"
            )
            return
        idx = int(sel[0])
        r = self.rows[idx]
        ci = self.col_idx
        filename = r[ci.get("nom du fichier", 0)] or ""
        tad = r[ci["tadarida_taxon"]] if "tadarida_taxon" in ci else ""
        tad_proba = r[ci["tadarida_probabilite"]] if "tadarida_probabilite" in ci else ""
        obs = r[ci["observateur_taxon"]] if "observateur_taxon" in ci else ""
        obs_conf = r[ci["observateur_probabilite"]] if "observateur_probabilite" in ci else ""

        self.sel_info_lbl.configure(
            text=f"{filename}   ·   Tadarida = {tad} ({tad_proba})"
        )
        self.edit_taxon_var.set(str(obs) if obs else "")
        self.edit_conf_var.set(str(obs_conf).upper() if obs_conf else "")

    def _mark_dirty(self):
        """Signale des modifications non enregistrées (titre ●, bouton coloré,
        header). Réinitialisé après _save. Évite le « ✓ enregistré » trompeur qui
        persistait après de nouvelles éditions."""
        self._dirty = True
        try:
            self.title(f"● Validation — {self.session_path.name}")
            self.save_btn.configure(text="● Enregistrer", fg_color="#c9700f")
            self.header_info.configure(text="modifications non enregistrées")
        except Exception:
            pass

    def _apply_to_row(self, idx: int, taxon: str, conf: str):
        """Applique (taxon observateur, confiance) à une ligne + MAJ visuelle."""
        if self._pushing:
            return   # édition gelée pendant un envoi (évite qu'une ligne diverge
                     # de la valeur réellement postée au serveur — cf. worker)
        r = self.rows[idx]
        ci = self.col_idx
        if "observateur_taxon" in ci:
            r[ci["observateur_taxon"]] = taxon or None
        if "observateur_probabilite" in ci:
            r[ci["observateur_probabilite"]] = conf or None
        self._mark_dirty()
        iid = str(idx)
        if self.tree.exists(iid):
            tad = r[ci["tadarida_taxon"]] if "tadarida_taxon" in ci else ""
            tad_lower = (str(tad) or "").lower()
            tags = []
            if taxon:
                tags.append("validated")
            if tad_lower in PATRIMONIAL_CODES:
                tags.append("patrimonial")
            if classify_taxon(str(tad)) == "noise":
                tags.append("noise")
            current = list(self.tree.item(iid, "values"))
            current[5] = taxon
            current[6] = conf
            self.tree.item(iid, values=current, tags=tags)
        # État de synchro (recalculé même si la ligne est filtrée/masquée).
        self._refresh_sync_state(idx)

    def _apply_edit_and_next(self):
        self._apply_edit()
        self._move_to_next_after_selection()

    def _apply_edit(self):
        """Applique le taxon + la confiance saisis à TOUTES les lignes
        sélectionnées (mono ou multi-sélection)."""
        sel = self.tree.selection()
        if not sel:
            return
        new_taxon = self.edit_taxon_var.get().strip()
        new_conf = self.edit_conf_var.get().strip().upper()
        for iid in sel:
            self._apply_to_row(int(iid), new_taxon, new_conf)
        self._refresh_stats()

    def _update_taxon_indicator(self):
        """Signale si le code saisi est accepté par le serveur (✓) ou non (⚠ :
        il ne pourra pas être envoyé). Silencieux sans snapshot taxons."""
        ind = getattr(self, "taxon_ind", None)
        if ind is None:
            return
        code = self.edit_taxon_var.get().strip()
        if not self._taxon_index or not code:
            ind.configure(text="")
        elif is_known(code, self._taxon_index):
            ind.configure(text="✓ connu", text_color="#16a34a")
        else:
            ind.configure(text="⚠ inconnu", text_color="#e08a1e")

    def _raise_to_genus(self):
        """Remplace, pour les lignes sélectionnées, l'espèce par son code « genre »
        (Pleaur/Pleaus → Plesp « Oreillard sp. ») quand il existe côté serveur —
        cas des sons incertains entre deux espèces d'un genre. Confiance conservée."""
        sel = self.tree.selection()
        if not sel:
            # Aucune ligne sélectionnée : si un taxon est saisi, propose son genre
            # dans le champ ; sinon explique quoi faire.
            typed = self.edit_taxon_var.get().strip()
            g = genus_code_for(typed, self._taxon_index) if typed else None
            if g:
                self.edit_taxon_var.set(g)
            else:
                messagebox.showinfo(
                    "Monter au genre",
                    "Sélectionnez d'abord une ou plusieurs lignes, ou saisissez une "
                    "espèce ayant un code « genre » (ex. Pleaur → Plesp).",
                    parent=self)
            return
        ci = self.col_idx
        conf = self.edit_conf_var.get().strip().upper()
        n_ok = 0
        for iid in sel:
            idx = int(iid)
            cur = None
            if "observateur_taxon" in ci and self.rows[idx][ci["observateur_taxon"]]:
                cur = self.rows[idx][ci["observateur_taxon"]]
            elif "tadarida_taxon" in ci:
                cur = self.rows[idx][ci["tadarida_taxon"]]
            g = genus_code_for(str(cur or ""), self._taxon_index)
            if g:
                c = conf or (str(self.rows[idx][ci["observateur_probabilite"]]).upper()
                             if "observateur_probabilite" in ci
                             and self.rows[idx][ci["observateur_probabilite"]] else "")
                self._apply_to_row(idx, g, c)
                n_ok += 1
        self._refresh_stats()
        if n_ok == 0:
            messagebox.showinfo(
                "Monter au genre",
                "Aucun code « genre » disponible pour la sélection "
                "(genre sans code groupe, ou déjà au genre).", parent=self)

    def _quick_validate(self, confidence: str):
        """Valide les lignes sélectionnées : chacune reçoit SON taxon Tadarida
        avec la confiance donnée (O/P/S). Mono ou multi-sélection."""
        sel = self.tree.selection()
        if not sel:
            return
        ci = self.col_idx
        for iid in sel:
            idx = int(iid)
            tad = self.rows[idx][ci["tadarida_taxon"]] if "tadarida_taxon" in ci else ""
            self._apply_to_row(idx, str(tad) if tad else "", confidence)
        self._refresh_stats()
        self._move_to_next_after_selection()

    def _clear_validation(self):
        """Efface la validation observateur des lignes sélectionnées."""
        sel = self.tree.selection()
        if not sel:
            return
        for iid in sel:
            self._apply_to_row(int(iid), "", "")
        self._refresh_stats()

    def _move_to_next_after_selection(self):
        """Sélectionne la ligne juste après la DERNIÈRE ligne sélectionnée
        (poursuite du flux de validation après un lot)."""
        children = self.tree.get_children()
        if not children:
            return
        sel = self.tree.selection()
        positions = [children.index(i) for i in sel if i in children]
        pos = max(positions) if positions else -1
        target = children[min(len(children) - 1, pos + 1)]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self.tree.see(target)

    def _move_selection(self, delta: int):
        sel = self.tree.selection()
        children = self.tree.get_children()
        if not children:
            return
        if not sel:
            target = children[0]
        else:
            try:
                pos = children.index(sel[0])
            except ValueError:
                target = children[0]
            else:
                new_pos = max(0, min(len(children) - 1, pos + delta))
                target = children[new_pos]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self.tree.see(target)

    # -- ChiroSurf ----------------------------------------------------------

    def _open_in_chirosurf(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        r = self.rows[idx]
        ci = self.col_idx
        filename = r[ci.get("nom du fichier", 0)] or ""

        wav_path = self._find_wav(str(filename))
        if wav_path is None:
            messagebox.showwarning(
                "WAV introuvable",
                f"Impossible de trouver le fichier WAV correspondant à :\n{filename}\n\n"
                f"Cherché dans {self.session_path / 'Data_k'} et {self.session_path / 'Data'}."
            )
            return

        if not self.chirosurf_path or not Path(self.chirosurf_path).is_file():
            if messagebox.askyesno(
                "ChiroSurf non configuré",
                "Le chemin vers ChiroSurf.exe n'est pas configuré.\n"
                "Ouvrir les préférences ?"):
                # Ping parent pour ouvrir ses préférences
                self.master.event_generate("<<OpenPreferences>>")
            return

        try:
            subprocess.Popen(
                [self.chirosurf_path, str(wav_path)],
                shell=False, close_fds=True,
            )
        except Exception as e:
            messagebox.showerror("Échec ouverture ChiroSurf", str(e))

    def _find_wav(self, filename_stem: str) -> Path | None:
        """Trouve le WAV correspondant au stem (avec ou sans .wav)."""
        stem = filename_stem.strip()
        if stem.lower().endswith(".wav"):
            stem = stem[:-4]
        # Le nom provient du xlsx d'observations (contenu non maîtrisé) : on ne
        # garde que le dernier composant pour empêcher tout path traversal
        # (../, chemin absolu) lors de la construction du chemin WAV.
        stem = Path(stem).name
        if not stem:
            return None

        for subdir in ("Data_k", "Data"):
            d = self.session_path / subdir
            if d.is_dir():
                candidate = d / f"{stem}.wav"
                if candidate.is_file():
                    return candidate
                # Au cas où : recherche par préfixe
                for p in d.glob(f"{stem}*.wav"):
                    return p
        # Racine de session
        candidate = self.session_path / f"{stem}.wav"
        if candidate.is_file():
            return candidate
        return None

    # -- envoi des identifications au serveur -------------------------------

    def _update_send_button(self):
        """Bouton contextuel : « Envoyer la sélection (N) » si des lignes sont
        sélectionnées, sinon « Envoyer tout (M) ». Désactivé si la vue n'est pas
        mappable (xlsx sans sidecar) ou si rien n'est envoyable."""
        if not hasattr(self, "send_btn") or self._pushing:
            return
        if not (self.mappable or self._participation_id):
            # ni sidecar, ni participation connue → rien à quoi rattacher l'envoi
            self.send_btn.configure(state="disabled", text="⬆ Envoyer (indispo.)")
            return
        def _to_send(i):
            # aligné sur _rows_to_push : envoyable ET pas déjà synchronisé
            return self._is_sendable(i) and self._row_state(i) != SYNC_SYNCED
        sel = [int(i) for i in self.tree.selection()]
        if sel:
            n = sum(1 for i in sel if _to_send(i))
            txt = f"⬆ Envoyer la sélection ({n})"
        else:
            n = sum(1 for i in range(len(self.rows)) if _to_send(i))
            txt = f"⬆ Envoyer tout ({n})"
        self.send_btn.configure(text=txt, state="normal" if n else "disabled")

    def _rows_to_push(self) -> list[int]:
        """Lignes à envoyer : la sélection si non vide, sinon toutes ; on ne garde
        que les envoyables (taxon + confiance + mappées) et pas déjà « synced »."""
        sel = [int(i) for i in self.tree.selection()]
        base = sel if sel else list(range(len(self.rows)))
        return [i for i in base
                if self._is_sendable(i) and self._row_state(i) != SYNC_SYNCED]

    def _on_send_click(self):
        if self._pushing or not (self.mappable or self._participation_id):
            return
        token = load_token()
        if not token:
            messagebox.showwarning(
                "Token requis",
                "Aucun token Vigie-Chiro enregistré.\n"
                "Renseignez-le dans ⚙ Préférences → API Vigie-Chiro.")
            return
        targets = self._rows_to_push()
        if not targets:
            messagebox.showinfo(
                "Rien à envoyer",
                "Aucune identification envoyable (taxon + confiance) en attente.")
            return
        n_retract = sum(1 for i in range(len(self.rows))
                        if self._row_state(i) == SYNC_TO_RETRACT)
        msg = f"Envoyer {len(targets)} identification(s) vers Vigie-Chiro ?"
        if n_retract:
            msg += (f"\n\nNote : {n_retract} identification(s) effacée(s) localement "
                    "resteront sur le serveur (une donnée ne peut pas être "
                    "supprimée côté observateur).")
        if not messagebox.askyesno("Envoyer les identifications", msg):
            return
        self._pushing = True
        self.send_btn.configure(state="disabled", text="⬆ Envoi en cours…")
        threading.Thread(target=self._push_worker, args=(token, targets),
                         daemon=True).start()

    def _push_worker(self, token: str, targets: list[int]):
        """Thread : reconstruit le mapping serveur si le sidecar manque, puis
        PATCH chaque identification (no_bilan=True) et relance le bilan une fois
        à la fin sur un envoi réussi."""
        results: dict[int, tuple] = {}
        last_ok = None
        new_keys = None
        new_entries = None
        try:
            client = VigieChiroClient(token)
        except Exception as e:
            self.after(0, self._on_push_done, results, f"Client API : {e}")
            return
        keymap = dict(self.row_keys)
        # Fallback : xlsx sans sidecar → re-fetch la participation pour mapper les
        # lignes à leurs observations serveur (par nom + temps_debut/fin).
        if not keymap and self._participation_id:
            try:
                donnees = list(client.iter_donnees(self._participation_id))
                new_entries = entries_from_donnees(donnees)
                keymap, _ = build_row_key_map(self.rows, self.col_idx, new_entries)
                new_keys = keymap
            except Exception as e:
                self.after(0, self._on_push_done, results,
                           f"Association au serveur impossible : {e}")
                return
        for idx in targets:
            key = keymap.get(idx)
            if key is None:
                results[idx] = (SYNC_ERROR, "contact introuvable côté serveur",
                                None, None)
                continue
            donnee_id, _, obs_index = key.partition("#")
            taxon, conf = self._row_taxon_conf(idx)
            try:
                taxon_id = client.resolve_taxon_id(taxon)
                client.push_observation(donnee_id, int(obs_index), taxon_id, conf,
                                        no_bilan=True)
                # valeur RÉELLEMENT postée (pas la ligne courante qui a pu bouger)
                results[idx] = (SYNC_SYNCED, "", taxon, conf)
                last_ok = (donnee_id, int(obs_index), taxon_id, conf)
            except Exception as e:
                results[idx] = (SYNC_ERROR, str(e), None, None)
        bilan_err = None
        if last_ok is not None:
            try:
                d_id, o_idx, t_id, cf = last_ok
                client.push_observation(d_id, o_idx, t_id, cf, no_bilan=False)
            except Exception as e:
                bilan_err = str(e)
        self.after(0, self._on_push_done, results, None, bilan_err, new_keys, new_entries)

    def _on_push_done(self, results, fatal=None, bilan_err=None,
                      new_keys=None, new_entries=None):
        """Thread UI : applique les résultats, MAJ pastilles + compteur, persiste."""
        self._pushing = False
        if fatal:
            messagebox.showerror("Envoi échoué", fatal, parent=self)
            self._update_send_button()
            return
        # Mapping reconstruit à la volée (fallback) : on l'adopte + le persistera _save.
        if new_keys is not None:
            self.row_keys = new_keys
            if new_entries is not None:
                self._sidecar_entries = new_entries
            self.mappable = bool(new_keys)
        n_ok = n_err = 0
        for idx, (state, message, sent_taxon, sent_conf) in results.items():
            if state == SYNC_SYNCED:
                key = self.row_keys.get(idx)
                if key is None:
                    continue
                rec = dict(self.sync_state.get(key) or {})
                # pushed_* = valeur postée ; l'état est recalculé depuis la ligne
                # courante (si elle a divergé → 'modified' au lieu d'un faux 'synced').
                rec["pushed_taxon"], rec["pushed_conf"] = sent_taxon, sent_conf
                rec.pop("error", None)
                cur_t, cur_c = self._row_taxon_conf(idx)
                rec["state"] = next_sync_state(cur_t, cur_c, rec) or SYNC_SYNCED
                self.sync_state[key] = rec
                self._update_row_dot(idx)
                n_ok += 1
            else:
                n_err += 1
                key = self.row_keys.get(idx)
                if key is not None:
                    rec = dict(self.sync_state.get(key) or {})
                    rec["state"] = state
                    rec["error"] = message
                    self.sync_state[key] = rec
                    self._update_row_dot(idx)
        self._save(silent=True)          # persiste xlsx + sidecar (état de synchro)
        self._update_registry_rollup()   # best-effort : rollup vue transverse
        self._refresh_stats()
        recap = f"{n_ok} identification(s) envoyée(s)"
        if n_err:
            recap += f", {n_err} en échec (pastilles rouges)"
        if bilan_err:
            recap += f"\n\nBilan non relancé : {bilan_err}"
        (messagebox.showwarning if (n_err or bilan_err) else messagebox.showinfo)(
            "Envoi terminé", recap)

    def _update_registry_rollup(self):
        """Best-effort : reporte X/Y (envoyées / envoyables) dans le registre
        central pour la vue transverse « Registre ». Localise ``registry.db`` en
        remontant depuis la session. Totalement silencieux en cas d'échec (le
        registre peut être absent, verrouillé, ou la session non référencée)."""
        try:
            total = len(self.rows)
            sendable = [i for i in range(total) if self._is_sendable(i)]
            n_synced = sum(1 for i in sendable if self._row_state(i) == SYNC_SYNCED)
            workspace = None
            for anc in [self.session_path, *self.session_path.parents]:
                if (anc / "registry.db").is_file():
                    workspace = anc
                    break
            if workspace is None:
                return
            from registry import Registry
            Registry(workspace).update_fields(
                self.session_path.name,
                {"ident_pushed": n_synced, "ident_total": len(sendable)})
        except Exception:
            pass

    # -- Sauvegarde ---------------------------------------------------------

    def _save(self, silent: bool = False):
        """Enregistre dans un fichier <nom>_<initiales>.xlsx pour préserver
        l'original. Persiste aussi le sidecar de synchro sous le nouveau stem
        (sinon le mapping serveur serait perdu au 1er enregistrement).

        ``silent`` : appel interne (après un envoi) — ne pas afficher le message
        « ✓ enregistré » (le récapitulatif d'envoi s'en charge)."""
        if not self.rows:
            return

        stem = self.xlsx_path.stem
        # Si le fichier d'origine se termine déjà par _KG ou similaire, on
        # écrase tel quel ; sinon on ajoute le suffixe initiales.
        if stem.endswith(f"_{self.initiales}"):
            out_path = self.xlsx_path
        else:
            out_path = self.xlsx_path.with_name(f"{stem}_{self.initiales}.xlsx")

        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "validation"
            ws.append(self.headers)
            for r in self.rows:
                ws.append(r)
            tmp = out_path.with_suffix(out_path.suffix + ".tmp")
            wb.save(tmp)
            tmp.replace(out_path)
        except Exception as e:
            messagebox.showerror("Enregistrement échoué", str(e))
            return

        # Sidecar de synchro : recopie le mapping natif + l'état de synchro sous
        # le stem sauvegardé (survit au renommage _<initiales>). Best-effort.
        if self._sidecar_entries or self.sync_state:
            try:
                save_observation_sidecar(out_path, self._sidecar_entries, self.sync_state)
            except Exception:
                pass

        self._dirty = False
        try:
            self.title(f"Validation — {self.session_path.name}")
            self.save_btn.configure(text="Enregistrer", fg_color=self._save_btn_fg)
        except Exception:
            pass
        if not silent:
            self.header_info.configure(text=f"✓ enregistré : {out_path.name}")

    def _on_close(self):
        if self._dirty:
            ans = messagebox.askyesnocancel(
                "Modifications non enregistrées",
                "Des modifications n'ont pas été enregistrées.\n"
                "Enregistrer avant de fermer ?"
            )
            if ans is None:  # cancel
                return
            if ans:
                self._save()
        self.destroy()


# ---------------------------------------------------------------------------
# Helper : trouver l'xlsx d'observations dans une session
# ---------------------------------------------------------------------------

def find_observations_xlsx(session_path: Path) -> Path | None:
    """Cherche un fichier observations compatible (sans _cleanup)."""
    candidates: list[Path] = []
    # Priorité : version _<initiales> si existe (validation en cours)
    # Sinon l'original téléchargé
    for p in session_path.iterdir():
        if not p.is_file():
            continue
        n = p.name.lower()
        if not n.endswith(".xlsx"):
            continue
        if "_cleanup" in n:
            continue
        if n.startswith("participation-") and "observations" in n:
            candidates.append(p)
    if not candidates:
        return None
    # Plus récent en priorité
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]
