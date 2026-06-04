"""
registry.py — registre centralisé SQLite des sessions.

Source de vérité persistante par workspace. Remplace la dépendance au
fichier ``Suivi analyse Chiros YYYY.xlsx`` tout en permettant l'import/export
vers celui-ci.

Fonctionnalités :
  - CRUD sessions (upsert idempotent depuis les manifestes scannés)
  - Colonnes custom utilisateur (ajout/suppression dynamique)
  - Import depuis Suivi xlsx
  - Export vers Suivi xlsx
  - Fusion de registres (merge .db ↔ .db avec last-write-wins)
  - Log de fusion (merge_log)
  - Calcul d'état global (pastilles couleur)
"""

from __future__ import annotations

import json
import os
import shutil
import sqlite3
import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

REGISTRY_FILENAME = "registry.db"
# Sous-dossier interne au workspace : isole les fichiers techniques
# (DB SQLite + WAL/SHM + backups) pour qu'ils ne pollent pas la racine
# du workspace utilisateur. Préfixe "_" pour rester visible mais distinct.
INTERNAL_DIR_NAME = "_chirotool"
INTERNAL_BACKUPS_SUBDIR = "backups"

INTERNAL_README = """\
Dossier interne à ChiroTool — NE PAS MODIFIER MANUELLEMENT
============================================================

Contenu :

  registry.db
      Base SQLite contenant l'index de toutes tes sessions, leur état
      (renommé, TE×10 fait, uploadé, nettoyé), les commentaires, les
      participations Vigie-Chiro synchronisées, etc. C'est le coeur de
      ChiroTool — sa perte impose un re-scan complet du workspace.

  registry.db-shm, registry.db-wal
      Fichiers techniques SQLite (mode WAL = Write-Ahead Logging). Ils
      contiennent les écritures en attente de commit et de la mémoire
      partagée pour les accès concurrents. Inutile d'y toucher — SQLite
      les gère automatiquement. Si tu les supprimes alors que ChiroTool
      est fermé, ils seront re-créés (avec perte potentielle des dernières
      écritures non commitées).

  backups/
      Copies horodatées de registry.db créées automatiquement avant les
      opérations destructrices (déduplication, fusion, etc.) et avant
      chaque mise à jour majeure. Tu peux supprimer manuellement les plus
      anciennes si l'espace disque devient critique.

Pour repartir de zéro : supprime ce dossier entier (ChiroTool en recréera
un nouveau au prochain démarrage par re-scan du workspace).
"""


def _internal_dir(workspace: Path) -> Path:
    """Retourne le chemin du dossier interne, le crée si nécessaire."""
    p = Path(workspace) / INTERNAL_DIR_NAME
    if not p.exists():
        try:
            p.mkdir(parents=True, exist_ok=True)
            (p / "README.txt").write_text(INTERNAL_README, encoding="utf-8")
        except OSError:
            pass
    return p


def _migrate_to_internal(workspace: Path) -> dict:
    """Si registry.db (et fichiers associés) sont à la racine du workspace,
    les déplace dans _chirotool/. Idempotent.

    Retourne un dict {moved: [...], skipped: [...]} pour traçabilité.
    """
    import shutil
    workspace = Path(workspace)
    report = {"moved": [], "skipped": []}
    target_dir = _internal_dir(workspace)

    # Fichiers à migrer : registry.db + ses fichiers techniques SQLite
    main_targets = ["registry.db", "registry.db-shm", "registry.db-wal"]
    for name in main_targets:
        src = workspace / name
        dst = target_dir / name
        if src.exists() and not dst.exists():
            try:
                shutil.move(str(src), str(dst))
                report["moved"].append(name)
            except OSError as e:
                report["skipped"].append(f"{name}: {e}")
        elif src.exists() and dst.exists():
            # Conflit : un registry.db existe à la fois à la racine et
            # dans _chirotool/. On ne touche pas (risque de perte de données).
            report["skipped"].append(
                f"{name}: existe aux deux emplacements, intervention manuelle requise"
            )

    # Migration des backups .backup_* éparpillés à la racine
    backups_dir = target_dir / INTERNAL_BACKUPS_SUBDIR
    has_legacy_backups = list(workspace.glob("registry.db*.backup_*"))
    if has_legacy_backups:
        backups_dir.mkdir(parents=True, exist_ok=True)
        for src in has_legacy_backups:
            dst = backups_dir / src.name
            if dst.exists():
                continue  # déjà copié
            try:
                shutil.move(str(src), str(dst))
                report["moved"].append(src.name)
            except OSError as e:
                report["skipped"].append(f"{src.name}: {e}")

    return report
SCHEMA_VERSION = 2   # v2 : ajout colonnes vigiechiro_participation_id + source


# ---------------------------------------------------------------------------
# État global calculé
# ---------------------------------------------------------------------------

ETAT_LABELS = {
    "not_started": ("🔴", "Brut"),
    "prepared":    ("🟠", "Préparé"),
    "in_progress": ("🟡", "En traitement"),
    "processed":   ("🟢", "Traité"),
    "archived":    ("⚫", "Archivé"),
}


def compute_etat(flags: dict) -> str:
    if flags.get("archived"):
        return "archived"
    done = sum(1 for k in ("renamed", "te10_done", "uploaded", "analyzed", "cleaned")
               if flags.get(k))
    if done == 0:
        return "not_started"
    if done <= 2:
        return "prepared"
    if done < 5:
        return "in_progress"
    return "processed"


# ---------------------------------------------------------------------------
# Schema SQL
# ---------------------------------------------------------------------------

_CREATE_TABLES = """
CREATE TABLE IF NOT EXISTS meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS sessions (
    id                TEXT PRIMARY KEY,
    workspace_path    TEXT,
    session_path      TEXT,
    campaign          TEXT,
    canonical_name    TEXT,

    date_debut        TEXT,
    date_fin          TEXT,
    n_site_tadarida   TEXT,
    n_point_fixe      TEXT,
    n_passage         INTEGER,
    n_enregistreur    INTEGER,
    n_serie           TEXT,
    n_micro           TEXT,

    nom_contrat       TEXT,
    cp                TEXT,
    emplacement       TEXT,

    meteo_vent        TEXT,
    meteo_couverture  TEXT,
    temperature_debut REAL,
    temperature_fin   REAL,

    renamed           INTEGER DEFAULT 0,
    te10_done         INTEGER DEFAULT 0,
    uploaded          INTEGER DEFAULT 0,
    analyzed          INTEGER DEFAULT 0,
    cleaned           INTEGER DEFAULT 0,
    archived          INTEGER DEFAULT 0,

    etat_global       TEXT DEFAULT 'not_started',
    commentaires      TEXT,
    custom_fields     TEXT DEFAULT '{}',

    n_wav             INTEGER DEFAULT 0,
    total_bytes       INTEGER DEFAULT 0,

    -- Couplage API Vigie-Chiro
    vigiechiro_participation_id TEXT,    -- ObjectId de la participation si connue
    vigiechiro_site_id          TEXT,
    source            TEXT DEFAULT 'local',   -- 'local' | 'server' | 'synced'
    api_etat          TEXT,                    -- état traitement côté API
    last_api_sync_at  TEXT,                    -- dernier pull

    first_scanned_at  TEXT,
    last_updated_at   TEXT
);

CREATE TABLE IF NOT EXISTS custom_columns (
    name         TEXT PRIMARY KEY,
    display_name TEXT,
    data_type    TEXT DEFAULT 'text',
    options      TEXT DEFAULT '[]',
    position     INTEGER DEFAULT 0,
    active       INTEGER DEFAULT 0,
    suggestion   INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS saved_views (
    name      TEXT PRIMARY KEY,
    filters   TEXT DEFAULT '{}',
    sort_by   TEXT DEFAULT 'date_debut',
    sort_asc  INTEGER DEFAULT 0,
    view_mode TEXT DEFAULT 'grouped'
);

CREATE TABLE IF NOT EXISTS merge_log (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT,
    source    TEXT,
    added     INTEGER DEFAULT 0,
    updated   INTEGER DEFAULT 0,
    conflicts INTEGER DEFAULT 0,
    notes     TEXT
);

"""

# Index créés APRÈS migration (évite d'utiliser des colonnes v2 sur une base v1)
_CREATE_INDEXES = """
CREATE INDEX IF NOT EXISTS idx_sessions_contrat ON sessions(nom_contrat);
CREATE INDEX IF NOT EXISTS idx_sessions_date    ON sessions(date_debut);
CREATE INDEX IF NOT EXISTS idx_sessions_etat    ON sessions(etat_global);
CREATE INDEX IF NOT EXISTS idx_sessions_campaign ON sessions(campaign);
"""

_DEFAULT_SUGGESTIONS = [
    ("facture",           "Facture",           "text",   "[]", 100, 0, 1),
    ("chef_projet",       "Chef de projet",    "text",   "[]", 101, 0, 1),
    ("prix_ht",           "Prix HT",           "number", "[]", 102, 0, 1),
    ("date_rendu_client", "Date rendu client", "date",   "[]", 103, 0, 1),
]

_DEFAULT_VIEWS = [
    ("Année en cours",  '{"year_current": true}',              "date_debut", 0, "grouped"),
    ("À traiter",       '{"etat_global": ["not_started", "prepared"]}', "date_debut", 0, "flat"),
    ("En traitement",   '{"etat_global": ["in_progress"]}',    "date_debut", 0, "flat"),
    ("Terminées",       '{"etat_global": ["processed"]}',      "date_debut", 0, "grouped"),
]


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------

class Registry:
    """Registre SQLite d'un workspace."""

    def __init__(self, workspace: Path):
        self.workspace = Path(workspace).resolve()
        # Migration auto : déplace registry.db (et SHM/WAL/backups) de la
        # racine vers _chirotool/ si nécessaire. Idempotent — n'a aucun
        # effet sur un workspace déjà migré ou nouveau.
        try:
            _migrate_to_internal(self.workspace)
        except Exception:
            pass
        self._internal_dir = _internal_dir(self.workspace)
        self.db_path = self._internal_dir / REGISTRY_FILENAME
        # Compat : si pour une raison X la migration n'a pas marché et que
        # registry.db reste à la racine, on l'utilise tel quel (pas de
        # création d'un duplicate dans _chirotool/).
        legacy = self.workspace / REGISTRY_FILENAME
        if legacy.exists() and not self.db_path.exists():
            self.db_path = legacy
        self._conn: sqlite3.Connection | None = None
        # Verrou applicatif : sqlite3 en mode check_same_thread=False autorise
        # les appels cross-thread, mais la connexion reste non thread-safe.
        # Les threads GUI (sync API, feed_from_scan background) sérialisent via ce lock.
        self._lock = threading.RLock()
        self._ensure_schema()

    # -- Connexion ---------------------------------------------------------

    def _get_conn(self) -> sqlite3.Connection:
        # Init paresseuse protégée : deux threads pourraient sinon créer deux
        # connexions concurrentes. Le RLock rend l'appel ré-entrant sûr depuis
        # les méthodes déjà lockées.
        if self._conn is None:
            with self._lock:
                if self._conn is None:  # double-check après acquisition
                    # check_same_thread=False + lock applicatif : accès depuis
                    # les threads GUI sans sqlite3.ProgrammingError.
                    conn = sqlite3.connect(
                        str(self.db_path), timeout=10, check_same_thread=False)
                    conn.row_factory = sqlite3.Row
                    conn.execute("PRAGMA journal_mode=WAL")
                    conn.execute("PRAGMA foreign_keys=ON")
                    conn.execute("PRAGMA busy_timeout=5000")
                    self._conn = conn
        return self._conn

    def close(self):
        with self._lock:
            if self._conn:
                try:
                    self._conn.close()
                except sqlite3.ProgrammingError:
                    pass
                self._conn = None

    def _migrate_schema(self, conn):
        """Ajoute les nouvelles colonnes aux bases existantes (v1 → v2)."""
        existing = {r[1] for r in conn.execute(
            "PRAGMA table_info(sessions)").fetchall()}
        to_add = [
            ("vigiechiro_participation_id", "TEXT"),
            ("vigiechiro_site_id", "TEXT"),
            ("source", "TEXT DEFAULT 'local'"),
            ("api_etat", "TEXT"),
            ("last_api_sync_at", "TEXT"),
        ]
        for col, typ in to_add:
            if col not in existing:
                conn.execute(f"ALTER TABLE sessions ADD COLUMN {col} {typ}")
        conn.commit()

    def _ensure_schema(self):
        conn = self._get_conn()
        # 1. Tables (CREATE IF NOT EXISTS → no-op si déjà là)
        conn.executescript(_CREATE_TABLES)
        # 2. Migrations (ALTER TABLE pour ajouter colonnes v2 sur base v1)
        self._migrate_schema(conn)
        # 3. Index (après migration pour pouvoir utiliser les colonnes ajoutées)
        conn.executescript(_CREATE_INDEXES)
        # 4. Meta
        conn.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES (?, ?)",
            ("schema_version", str(SCHEMA_VERSION)),
        )
        # Suggestions colonnes
        for row in _DEFAULT_SUGGESTIONS:
            conn.execute(
                "INSERT OR IGNORE INTO custom_columns "
                "(name, display_name, data_type, options, position, active, suggestion) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)", row,
            )
        # Vues par défaut
        for row in _DEFAULT_VIEWS:
            conn.execute(
                "INSERT OR IGNORE INTO saved_views "
                "(name, filters, sort_by, sort_asc, view_mode) "
                "VALUES (?, ?, ?, ?, ?)", row,
            )
        conn.commit()

    # -- CRUD sessions -----------------------------------------------------

    def upsert_session(self, data: dict, *, _commit: bool = True) -> str:
        """Insert ou update une session. Retourne 'inserted' ou 'updated'.

        Thread-safe : sérialisé par self._lock (appelé potentiellement depuis
        sync_from_api en thread background et feed_from_scan en parallèle).

        ``_commit=False`` : mode batch — la transaction n'est pas commitée
        après chaque upsert. L'appelant DOIT commit manuellement en fin de
        batch (et rollback en cas d'erreur). Utilisé par ``feed_from_scan``
        pour éviter ~N fsync (gain ×5-20 sur HDD).
        """
        with self._lock:
            return self._upsert_session_locked(data, _commit=_commit)

    def _upsert_session_locked(self, data: dict, *, _commit: bool = True) -> str:
        sid = data.get("id") or data.get("canonical_name") or data.get("name")
        if not sid:
            raise ValueError("session sans identifiant")

        now = datetime.now().isoformat(timespec="seconds")
        flags = {
            "renamed": int(bool(data.get("renamed") or data.get("flag_renamed"))),
            "te10_done": int(bool(data.get("te10_done") or data.get("flag_te10_done"))),
            "uploaded": int(bool(data.get("uploaded") or data.get("flag_uploaded_hint"))),
            "analyzed": int(bool(data.get("analyzed") or data.get("flag_analyzed"))),
            "cleaned": int(bool(data.get("cleaned") or data.get("flag_cleaned"))),
            "archived": int(bool(data.get("archived", 0))),
        }
        etat = compute_etat(flags)

        conn = self._get_conn()
        existing = conn.execute("SELECT id FROM sessions WHERE id=?", (sid,)).fetchone()

        values = {
            "id": sid,
            "workspace_path": str(data.get("workspace_path") or self.workspace),
            "session_path": str(data.get("path") or data.get("session_path") or ""),
            "campaign": data.get("campaign") or "",
            "canonical_name": data.get("canonical_name") or sid,
            "date_debut": data.get("date_debut") or "",
            "date_fin": data.get("date_fin") or "",
            "n_site_tadarida": data.get("n_site_tadarida") or "",
            "n_point_fixe": data.get("n_point_fixe") or "",
            "n_passage": data.get("n_passage"),
            "n_enregistreur": data.get("n_enregistreur"),
            "n_serie": data.get("n_serie") or "",
            "n_micro": data.get("n_micro") or "",
            "nom_contrat": data.get("nom_contrat") or "",
            "cp": data.get("cp") or "",
            "emplacement": data.get("emplacement") or "",
            "meteo_vent": data.get("meteo_vent") or "",
            "meteo_couverture": data.get("meteo_couverture") or "",
            "temperature_debut": data.get("temperature_debut"),
            "temperature_fin": data.get("temperature_fin"),
            **flags,
            "etat_global": etat,
            "commentaires": data.get("commentaires") or "",
            "custom_fields": json.dumps(data.get("custom_fields") or {},
                                          ensure_ascii=False),
            "n_wav": data.get("n_wav") or 0,
            "total_bytes": data.get("total_bytes") or 0,
            "last_updated_at": now,
        }

        if existing:
            cols = ", ".join(f"{k}=?" for k in values if k != "id")
            vals = [v for k, v in values.items() if k != "id"] + [sid]
            conn.execute(f"UPDATE sessions SET {cols} WHERE id=?", vals)
            if _commit:
                conn.commit()
            return "updated"
        else:
            values["first_scanned_at"] = now
            cols = ", ".join(values.keys())
            phs = ", ".join("?" * len(values))
            conn.execute(f"INSERT INTO sessions ({cols}) VALUES ({phs})",
                         list(values.values()))
            if _commit:
                conn.commit()
            return "inserted"

    def get_session(self, sid: str) -> dict | None:
        with self._lock:
            row = self._get_conn().execute(
                "SELECT * FROM sessions WHERE id=?", (sid,)).fetchone()
            return dict(row) if row else None

    def list_sessions(self, *, include_archived: bool = False) -> list[dict]:
        """Alias de all_sessions sans filtre WHERE custom (API publique stable)."""
        return self.all_sessions(include_archived=include_archived)

    def all_sessions(self, *, include_archived: bool = False,
                      where: str = "", params: tuple = ()) -> list[dict]:
        sql = "SELECT * FROM sessions"
        conditions = []
        p = list(params)
        if not include_archived:
            conditions.append("archived=0")
        if where:
            conditions.append(where)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY date_debut DESC"
        with self._lock:
            return [dict(r) for r in self._get_conn().execute(sql, p).fetchall()]

    def search(self, query: str, *, include_archived: bool = False) -> list[dict]:
        """Recherche floue sur tous les champs texte principaux."""
        like = f"%{query}%"
        sql = """SELECT * FROM sessions WHERE (
            nom_contrat LIKE ? OR campaign LIKE ? OR canonical_name LIKE ?
            OR n_site_tadarida LIKE ? OR n_point_fixe LIKE ? OR n_serie LIKE ?
            OR commentaires LIKE ? OR etat_global LIKE ?
        )"""
        if not include_archived:
            sql += " AND archived=0"
        sql += " ORDER BY date_debut DESC"
        with self._lock:
            return [dict(r) for r in self._get_conn().execute(
                sql, (like,) * 8).fetchall()]

    # Colonnes autorisées en écriture via update_fields (whitelist anti-injection
    # sur les noms de colonnes interpolés en f-string).
    _WRITABLE_COLUMNS = frozenset({
        "workspace_path", "session_path", "campaign", "canonical_name",
        "date_debut", "date_fin", "n_site_tadarida", "n_point_fixe",
        "n_passage", "n_enregistreur", "n_serie", "n_micro", "nom_contrat",
        "cp", "emplacement", "meteo_vent", "meteo_couverture",
        "temperature_debut", "temperature_fin", "renamed", "te10_done",
        "uploaded", "analyzed", "cleaned", "archived", "etat_global",
        "commentaires", "custom_fields", "n_wav", "total_bytes",
        "vigiechiro_participation_id", "vigiechiro_site_id", "api_etat",
        "source", "last_api_sync_at", "last_updated_at", "first_scanned_at",
    })

    def update_fields(self, sid: str, fields: dict) -> None:
        """Mise à jour partielle de champs spécifiques.

        Thread-safe + whitelist des noms de colonnes (les noms sont interpolés
        en f-string, on rejette toute clé inconnue pour éviter une requête
        cassée ou une surface d'injection via édition inline / custom_fields).
        """
        if not fields:
            return
        with self._lock:
            self._update_fields_locked(sid, fields, commit=True)

    def _update_fields_locked(self, sid: str, fields: dict, *,
                                commit: bool = True) -> None:
        """Variante interne (lock déjà détenu). commit=False pour batcher."""
        now = datetime.now().isoformat(timespec="seconds")
        fields = dict(fields)  # copie défensive
        fields["last_updated_at"] = now
        # Recalcul etat si flags modifiés
        if any(k in fields for k in ("renamed", "te10_done", "uploaded",
                                       "analyzed", "cleaned", "archived")):
            row = self._get_conn().execute(
                "SELECT * FROM sessions WHERE id=?", (sid,)).fetchone()
            current = dict(row) if row else {}
            merged = {**current, **fields}
            fields["etat_global"] = compute_etat(merged)
        # Whitelist : on ne garde que les colonnes connues
        safe = {k: v for k, v in fields.items()
                if k in self._WRITABLE_COLUMNS}
        unknown = set(fields) - set(safe)
        if unknown:
            import logging as _log
            _log.warning("update_fields : colonnes ignorées (inconnues) : %s",
                          ", ".join(sorted(unknown)))
        if not safe:
            return
        cols = ", ".join(f"{k}=?" for k in safe)
        vals = list(safe.values()) + [sid]
        conn = self._get_conn()
        conn.execute(f"UPDATE sessions SET {cols} WHERE id=?", vals)
        if commit:
            conn.commit()

    def archive_session(self, sid: str) -> None:
        self.update_fields(sid, {"archived": 1})

    def unarchive_session(self, sid: str) -> None:
        self.update_fields(sid, {"archived": 0})

    def delete_session(self, sid: str) -> None:
        with self._lock:
            conn = self._get_conn()
            conn.execute("DELETE FROM sessions WHERE id=?", (sid,))
            conn.commit()

    # -- Déduplication -----------------------------------------------------

    def find_duplicates(self) -> list[list[dict]]:
        """Retourne les groupes de sessions ayant la même clé fonctionnelle.

        Clé : (n_site_tadarida, n_point_fixe.upper, n_passage, date_debut[:10]).
        Un groupe avec ≥ 2 sessions = doublon à fusionner.

        Causes typiques :
          - Rename de dossier qui produit un nouvel id (ancien id resté)
          - Déplacement de workspace inter-PC qui re-scanne avec le même contenu
          - Ghost API (source=server) qui matche maintenant une session locale
        """
        with self._lock:
            conn = self._get_conn()
            rows = conn.execute(
                "SELECT * FROM sessions WHERE archived=0 "
                "AND n_site_tadarida IS NOT NULL AND n_site_tadarida <> '' "
                "AND n_point_fixe IS NOT NULL AND n_point_fixe <> '' "
                "AND date_debut IS NOT NULL AND date_debut <> ''"
            ).fetchall()
        groups: dict[tuple, list[dict]] = {}
        for r in rows:
            d = dict(r)
            key = (
                (d.get("n_site_tadarida") or "").strip(),
                (d.get("n_point_fixe") or "").upper().strip(),
                d.get("n_passage"),
                str(d.get("date_debut") or "")[:10],
            )
            if not key[0] or not key[1] or key[2] is None or not key[3]:
                continue
            groups.setdefault(key, []).append(d)
        return [g for g in groups.values() if len(g) > 1]

    def find_empty_entries(self) -> list[dict]:
        """Retourne les entrées sans métadonnées exploitables (site/point/date).

        Typiquement des artefacts d'anciens scans sur des noms de dossier
        non-canoniques qui ne contenaient pas assez d'info.
        """
        with self._lock:
            conn = self._get_conn()
            rows = conn.execute(
                "SELECT * FROM sessions WHERE archived=0 AND ("
                "n_site_tadarida IS NULL OR n_site_tadarida = '' "
                "OR n_point_fixe IS NULL OR n_point_fixe = '' "
                "OR date_debut IS NULL OR date_debut = '')"
            ).fetchall()
        return [dict(r) for r in rows]

    def _score_session(self, s: dict) -> tuple:
        """Score pour choisir la "meilleure" entrée à conserver dans un groupe.

        Critères (du plus important au moins important) :
          1. Session avec un path qui EXISTE encore sur disque
          2. Flags les plus avancés (cleaned > analyzed > uploaded > te10 > renamed)
          3. participation_id présent (synchronisé avec API)
          4. canonical_name au format moderne (commence par YYYYMMDD_site)
          5. last_updated_at le plus récent
        """
        path = s.get("session_path") or ""
        path_exists = bool(path) and Path(path).is_dir()
        flag_score = (
            int(bool(s.get("cleaned"))) * 16
            + int(bool(s.get("analyzed"))) * 8
            + int(bool(s.get("uploaded"))) * 4
            + int(bool(s.get("te10_done"))) * 2
            + int(bool(s.get("renamed"))) * 1
        )
        has_pid = bool(s.get("vigiechiro_participation_id"))
        canonical_modern = (s.get("canonical_name") or "").startswith(
            tuple(f"{y}" for y in range(2020, 2100)))
        last_upd = s.get("last_updated_at") or ""
        return (
            int(path_exists), flag_score, int(has_pid),
            int(canonical_modern), last_upd,
        )

    def deduplicate(self, *, dry_run: bool = False) -> dict:
        """Fusionne les entrées doublonnées par (site, point, passage, jour).

        Pour chaque groupe :
          - Sélectionne la "meilleure" entrée via ``_score_session``
          - Fusionne les flags des autres dans la winner (max booléen)
          - Préserve participation_id, vigiechiro_site_id, commentaires non vides
            si la winner ne les a pas
          - Supprime les autres entrées

        Nettoie aussi les entrées orphelines (sans site/point/date).

        Retourne ``{groups, merged, deleted_orphans, dry_run}``.
        """
        report = {"groups": 0, "merged": 0, "deleted_orphans": 0,
                  "dry_run": dry_run, "details": []}

        # 1. Doublons fonctionnels
        groups = self.find_duplicates()
        report["groups"] = len(groups)

        with self._lock:
            conn = self._get_conn()
            for group in groups:
                # Trie : meilleur en premier (score décroissant)
                sorted_group = sorted(group, key=self._score_session,
                                        reverse=True)
                winner = sorted_group[0]
                losers = sorted_group[1:]

                # Fusionne les flags (max booléen) et champs API
                merged_fields = {}
                for f in ("renamed", "te10_done", "uploaded",
                          "analyzed", "cleaned"):
                    cur = bool(winner.get(f))
                    for l in losers:
                        cur = cur or bool(l.get(f))
                    if cur and not winner.get(f):
                        merged_fields[f] = 1
                # participation_id : prend le 1er non-null si winner manquant
                if not winner.get("vigiechiro_participation_id"):
                    for l in losers:
                        if l.get("vigiechiro_participation_id"):
                            merged_fields["vigiechiro_participation_id"] = \
                                l["vigiechiro_participation_id"]
                            break
                if not winner.get("vigiechiro_site_id"):
                    for l in losers:
                        if l.get("vigiechiro_site_id"):
                            merged_fields["vigiechiro_site_id"] = \
                                l["vigiechiro_site_id"]
                            break
                # Commentaires : concat non-vides en évitant doublons
                comments = [winner.get("commentaires") or ""]
                for l in losers:
                    c = l.get("commentaires")
                    if c and c not in comments:
                        comments.append(c)
                non_empty = [c for c in comments if c]
                if non_empty and len(non_empty) > 1:
                    merged_fields["commentaires"] = "\n---\n".join(non_empty)

                # Détail pour reporting
                report["details"].append({
                    "key": (winner.get("n_site_tadarida"),
                            winner.get("n_point_fixe"),
                            winner.get("n_passage"),
                            str(winner.get("date_debut") or "")[:10]),
                    "winner_id": winner["id"],
                    "losers_ids": [l["id"] for l in losers],
                    "merged_fields": list(merged_fields.keys()),
                })

                if dry_run:
                    continue

                # Applique le merge sur la winner
                if merged_fields:
                    cols = ", ".join(f"{k}=?" for k in merged_fields)
                    vals = list(merged_fields.values()) + [winner["id"]]
                    conn.execute(
                        f"UPDATE sessions SET {cols}, last_updated_at=? "
                        f"WHERE id=?",
                        vals[:-1] + [datetime.now().isoformat(timespec="seconds"),
                                      winner["id"]],
                    )

                # Supprime les losers
                for l in losers:
                    conn.execute("DELETE FROM sessions WHERE id=?",
                                  (l["id"],))
                    report["merged"] += 1

            # 2. Orphelines (entrées sans site/point/date)
            orphans = self.find_empty_entries()
            for o in orphans:
                if not dry_run:
                    conn.execute("DELETE FROM sessions WHERE id=?",
                                  (o["id"],))
                report["deleted_orphans"] += 1

            if not dry_run:
                conn.commit()

        return report

    # -- Stats rapides -----------------------------------------------------

    def stats(self) -> dict:
        with self._lock:
            conn = self._get_conn()
            total = conn.execute("SELECT COUNT(*) FROM sessions WHERE archived=0").fetchone()[0]
            by_etat = {}
            for row in conn.execute(
                "SELECT etat_global, COUNT(*) FROM sessions WHERE archived=0 "
                "GROUP BY etat_global"):
                by_etat[row[0]] = row[1]
            contracts = conn.execute(
                "SELECT COUNT(DISTINCT nom_contrat) FROM sessions WHERE archived=0"
            ).fetchone()[0]
        return {"total": total, "by_etat": by_etat, "contracts": contracts}

    # -- Colonnes custom ---------------------------------------------------

    def custom_columns(self, *, active_only: bool = False) -> list[dict]:
        sql = "SELECT * FROM custom_columns"
        if active_only:
            sql += " WHERE active=1"
        sql += " ORDER BY position"
        with self._lock:
            return [dict(r) for r in self._get_conn().execute(sql).fetchall()]

    def add_custom_column(self, name: str, display_name: str,
                           data_type: str = "text") -> None:
        with self._lock:
            conn = self._get_conn()
            conn.execute(
                "INSERT OR REPLACE INTO custom_columns "
                "(name, display_name, data_type, active, suggestion) "
                "VALUES (?, ?, ?, 1, 0)",
                (name, display_name, data_type),
            )
            conn.commit()

    def toggle_custom_column(self, name: str, active: bool) -> None:
        with self._lock:
            conn = self._get_conn()
            conn.execute(
                "UPDATE custom_columns SET active=? WHERE name=?",
                (int(active), name),
            )
            conn.commit()

    # -- Vues sauvegardées -------------------------------------------------

    def saved_views(self) -> list[dict]:
        with self._lock:
            return [dict(r) for r in self._get_conn().execute(
                "SELECT * FROM saved_views ORDER BY name").fetchall()]

    # -- Merge (fusion de registres) ---------------------------------------

    def merge_from_db(self, source_path: Path) -> dict:
        """Fusionne un autre registry.db dans celui-ci.

        Retourne {added, updated, conflicts, details: [...]}.
        """
        source = sqlite3.connect(str(source_path), timeout=10)
        source.row_factory = sqlite3.Row
        result = {"added": 0, "updated": 0, "conflicts": 0, "details": []}

        for row in source.execute("SELECT * FROM sessions").fetchall():
            src = dict(row)
            sid = src["id"]
            existing = self.get_session(sid)

            if existing is None:
                self.upsert_session(src)
                result["added"] += 1
                result["details"].append(("add", sid, None))
            else:
                src_ts = src.get("last_updated_at") or ""
                dst_ts = existing.get("last_updated_at") or ""
                if src_ts > dst_ts:
                    # Source plus récente → update
                    # Merge commentaires et custom_fields
                    merged_cf = json.loads(existing.get("custom_fields") or "{}")
                    src_cf = json.loads(src.get("custom_fields") or "{}")
                    for k, v in src_cf.items():
                        if v:
                            merged_cf[k] = v
                    src["custom_fields"] = merged_cf
                    if not src.get("commentaires") and existing.get("commentaires"):
                        src["commentaires"] = existing["commentaires"]
                    self.upsert_session(src)
                    result["updated"] += 1
                    result["details"].append(("update", sid,
                                               f"{dst_ts} → {src_ts}"))
                else:
                    # Cible plus récente → rien à faire
                    pass

        source.close()

        # Log
        now = datetime.now().isoformat(timespec="seconds")
        with self._lock:
            conn = self._get_conn()
            conn.execute(
                "INSERT INTO merge_log (timestamp, source, added, updated, conflicts) "
                "VALUES (?, ?, ?, ?, ?)",
                (now, str(source_path), result["added"], result["updated"],
                 result["conflicts"]),
            )
            conn.commit()
        return result

    def export_db(self, dest: Path) -> Path:
        """Copie le registry.db vers un fichier (pour échange).

        Checkpoint WAL sous lock avant la copie pour que le fichier exporté
        contienne bien les dernières écritures (sinon elles restent dans le
        -wal et la copie du .db seul serait incomplète).
        """
        dest = Path(dest)
        with self._lock:
            self._get_conn().execute("PRAGMA wal_checkpoint(FULL)")
            shutil.copy2(self.db_path, dest)
        return dest

    def export_to_csv(self, dest: Path, *,
                       include_archived: bool = False) -> dict:
        """Export complet du registre en CSV (toutes colonnes, UTF-8 BOM).

        Format conçu pour faire le suivi d'avancement d'une équipe : 1 ligne
        par session, toutes les colonnes intéressantes (méta, flags, API,
        matériel, météo, volumétrie).

        UTF-8 BOM pour qu'Excel ouvre proprement les accents en double-clic.
        Séparateur ``;`` pour respecter la locale française par défaut
        d'Excel (",") qui s'embrouille avec les chiffres décimaux.

        Fusion entre BE : 3 BE exportent leur CSV, on les concatène avec
        ``head -1 fichier1.csv > total.csv ; tail -n +2 fichier*.csv >> total.csv``
        ou via pandas pour de la dédup propre par ``session_id``.

        Retourne {"path", "n_rows", "size_bytes"}.
        """
        import csv
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)

        # Colonnes SQL réelles (cf schéma _CREATE_TABLES). Le header affiché
        # dans le CSV reprend des noms plus parlants pour l'utilisateur final.
        sql_columns = [
            # Identification
            "id", "canonical_name", "nom_contrat", "campaign",
            "session_path", "workspace_path",
            # Métadonnées protocole
            "date_debut", "date_fin", "n_site_tadarida", "n_point_fixe",
            "n_passage", "emplacement", "cp",
            # Flags pipeline
            "etat_global", "renamed", "te10_done", "uploaded",
            "analyzed", "cleaned", "archived",
            # API Vigie-Chiro
            "vigiechiro_participation_id", "vigiechiro_site_id",
            "api_etat", "source", "last_api_sync_at",
            # Matériel
            "n_enregistreur", "n_serie", "n_micro",
            # Météo
            "meteo_vent", "meteo_couverture",
            "temperature_debut", "temperature_fin",
            # Volumétrie
            "n_wav", "total_bytes",
            # Suivi
            "commentaires", "first_scanned_at", "last_updated_at",
        ]
        # En-tête CSV plus lisible que le nom SQL pour la colonne id
        header_columns = ["session_id" if c == "id" else c
                           for c in sql_columns]
        # Colonne calculée additionnelle : taille en Go (lisible humain).
        # Sera ajoutée APRÈS total_bytes. On la marque pour traitement spécial.
        DATE_COLUMNS = {
            "date_debut", "date_fin", "last_api_sync_at",
            "first_scanned_at", "last_updated_at",
        }
        # Insertion de la colonne total_gb juste après total_bytes
        try:
            idx_bytes = header_columns.index("total_bytes")
            header_columns.insert(idx_bytes + 1, "total_gb")
        except ValueError:
            idx_bytes = -1

        with self._lock:
            conn = self._get_conn()
            sql = "SELECT " + ", ".join(sql_columns) + " FROM sessions"
            if not include_archived:
                sql += " WHERE archived=0"
            sql += " ORDER BY date_debut DESC, nom_contrat, n_site_tadarida, n_point_fixe"
            rows = conn.execute(sql).fetchall()

        # Écriture avec BOM UTF-8 + séparateur ; (Excel-friendly FR)
        with dest.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header_columns)
            for r in rows:
                values = []
                for col in sql_columns:
                    try:
                        v = r[col]
                    except (KeyError, IndexError):
                        v = None
                    if v is None:
                        values.append("")
                    elif isinstance(v, bool):
                        values.append("1" if v else "0")
                    elif col in DATE_COLUMNS and isinstance(v, str):
                        # ISO 8601 avec 'T' → format Excel-friendly (espace)
                        # → Excel le reconnaît comme date triable/filtrable
                        values.append(v.replace("T", " "))
                    else:
                        values.append(str(v))
                    # Après total_bytes, insère la colonne total_gb calculée
                    if col == "total_bytes":
                        try:
                            n = int(v) if v is not None else 0
                            gb = n / (1024 ** 3)
                            # Locale FR : virgule décimale pour Excel français
                            values.append(f"{gb:.3f}".replace(".", ","))
                        except (TypeError, ValueError):
                            values.append("")
                writer.writerow(values)

        size = dest.stat().st_size
        return {
            "path": str(dest),
            "n_rows": len(rows),
            "size_bytes": size,
        }

    def merge_log(self, limit: int = 20) -> list[dict]:
        with self._lock:
            return [dict(r) for r in self._get_conn().execute(
                "SELECT * FROM merge_log ORDER BY timestamp DESC LIMIT ?",
                (limit,),
            ).fetchall()]

    # -- Import / Export xlsx -----------------------------------------------

    def import_from_suivi_xlsx(self, xlsx_path: Path) -> dict:
        """Importe les lignes du Suivi analyse Chiros vers le registre."""
        from suivi import Suivi
        s = Suivi(xlsx_path)
        result = {"imported": 0, "skipped": 0}
        for row in s.rows:
            if not row.nom_contrat and not row.n_site_tadarida:
                result["skipped"] += 1
                continue
            # Construire un ID unique stable
            date_s = row.date_debut.strftime("%Y%m%d") if row.date_debut else "nodate"
            sid = (f"{date_s}_site{row.n_site_tadarida or 'X'}"
                   f"_{row.n_point_fixe or 'X'}"
                   f"_Pass{row.n_passage or 0}"
                   f"_enr{row.n_enregistreur or 0:02d}")
            data = {
                "id": sid,
                "nom_contrat": row.nom_contrat,
                "cp": row.cp,
                "date_debut": row.date_debut.isoformat() if row.date_debut else "",
                "date_fin": row.date_fin.isoformat() if row.date_fin else "",
                "n_site_tadarida": row.n_site_tadarida,
                "n_point_fixe": row.n_point_fixe,
                "n_passage": row.n_passage,
                "n_enregistreur": row.n_enregistreur,
                "n_serie": row.n_serie,
                "n_micro": row.n_micro,
                "emplacement": row.emplacement,
                "meteo_vent": row.vent,
                "meteo_couverture": row.couverture,
                "commentaires": row.commentaires,
                "renamed": int(bool(row.flag_lupas_kaleidoscope)),
                "te10_done": int(bool(row.flag_lupas_kaleidoscope)),
                "analyzed": int(bool(row.flag_tadarida)),
            }
            self.upsert_session(data)
            result["imported"] += 1

        # Log
        now = datetime.now().isoformat(timespec="seconds")
        self._get_conn().execute(
            "INSERT INTO merge_log (timestamp, source, added, updated, notes) "
            "VALUES (?, ?, ?, 0, ?)",
            (now, str(xlsx_path), result["imported"],
             f"import Suivi xlsx, {result['skipped']} lignes vides ignorées"),
        )
        self._get_conn().commit()
        return result

    def export_to_xlsx(self, dest: Path) -> Path:
        """Exporte le registre filtré (non archivé) en xlsx."""
        import openpyxl
        from openpyxl.styles import PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registre ChiroTool"

        # Colonnes fixes + custom actives
        fixed_cols = [
            ("nom_contrat", "Contrat"),
            ("date_debut", "Date début"),
            ("n_site_tadarida", "N° site Tadarida"),
            ("n_point_fixe", "Point"),
            ("n_passage", "Passage"),
            ("n_enregistreur", "N° enregistreur"),
            ("n_serie", "N° Série"),
            ("etat_global", "État"),
            ("campaign", "Campagne"),
            ("commentaires", "Commentaires"),
        ]
        custom = self.custom_columns(active_only=True)
        all_cols = fixed_cols + [(c["name"], c["display_name"]) for c in custom]

        # Header
        ws.append([label for _, label in all_cols])

        fills = {
            "not_started": PatternFill(fgColor="FFE6E6", fill_type="solid"),
            "prepared":    PatternFill(fgColor="FFE8CC", fill_type="solid"),
            "in_progress": PatternFill(fgColor="FFFDE6", fill_type="solid"),
            "processed":   PatternFill(fgColor="E6FFE6", fill_type="solid"),
            "archived":    PatternFill(fgColor="E0E0E0", fill_type="solid"),
        }

        for s in self.all_sessions(include_archived=True):
            cf = json.loads(s.get("custom_fields") or "{}")
            row_data = []
            for key, _label in all_cols:
                if key in s:
                    row_data.append(s[key])
                elif key in cf:
                    row_data.append(cf[key])
                else:
                    row_data.append("")
            ws.append(row_data)
            etat = s.get("etat_global", "")
            if etat in fills:
                for cell in ws[ws.max_row]:
                    cell.fill = fills[etat]

        dest = Path(dest)
        wb.save(str(dest))
        return dest

    # -- Synchronisation avec l'API Vigie-Chiro ---------------------------

    def sync_from_api(self, token: str, on_progress=None,
                       apply_changes: bool = True) -> dict:
        """Pull les participations depuis l'API et réconcilie avec le registre.

        Si ``apply_changes=False`` : retourne juste le diff sans modifier le
        registre (utile pour afficher un aperçu à l'utilisateur).

        Retourne :
            {
              "total_fetched": int,
              "updates": [...],    # local sessions à mettre à jour
              "ghosts": [...],     # participations serveur sans équivalent local
              "unchanged": [...],  # déjà alignées
              "errors": [...],
            }

        **Règle de mise à jour** : on n'upgrade que les flags (False → True),
        jamais downgrade, pour préserver un état local plus avancé (ex:
        cleanup fait localement même si la participation a été supprimée
        côté serveur).
        """
        from vigiechiro_api import VigieChiroClient

        result = {
            "total_fetched": 0,
            "updates": [],
            "ghosts": [],
            "unchanged": [],
            "errors": [],
        }

        try:
            client = VigieChiroClient(token, source="background")
        except Exception as e:
            result["errors"].append(f"client API : {e}")
            return result

        # Récupération (avec progression)
        participations = []
        try:
            for p in client.iter_participations(mine_only=True):
                participations.append(p)
                if on_progress is not None:
                    on_progress(len(participations), None)
        except Exception as e:
            result["errors"].append(f"pull API : {e}")
            return result

        result["total_fetched"] = len(participations)

        # Index des sessions locales pour matching rapide
        local_sessions = self.all_sessions(include_archived=True)
        by_pid: dict[str, dict] = {}
        by_fuzzy: dict[tuple, dict] = {}
        for s in local_sessions:
            pid = s.get("vigiechiro_participation_id")
            if pid:
                by_pid[pid] = s
            site = s.get("n_site_tadarida") or ""
            point = s.get("n_point_fixe") or ""
            date = (s.get("date_debut") or "")[:10]  # YYYY-MM-DD seulement
            if site and point and date:
                by_fuzzy[(site, point, date)] = s

        # Traitement de chaque participation API
        for raw in participations:
            try:
                p_id = raw.get("_id") or raw.get("id")
                site_obj = raw.get("site") or {}
                point = raw.get("point") or ""
                date_debut = (raw.get("date_debut") or "")[:10]
                traitement = raw.get("traitement") or {}
                etat = traitement.get("etat") or ""

                # Extraire le n° site depuis titre
                numero = None
                if isinstance(site_obj, dict):
                    titre = site_obj.get("titre") or ""
                    import re
                    m = re.search(r"-(\d{5,6})\s*$", titre)
                    if m:
                        numero = m.group(1).zfill(6)
                    site_id_api = site_obj.get("_id")
                else:
                    site_id_api = None

                # Essai 1 : match par participation_id (lien direct)
                local = by_pid.get(p_id) if p_id else None
                # Essai 2 : match fuzzy
                if local is None and numero and point and date_debut:
                    local = by_fuzzy.get((numero, point, date_debut))

                if local is None:
                    # Ghost : participation serveur sans équivalent local
                    result["ghosts"].append({
                        "participation_id": p_id,
                        "site_id": site_id_api,
                        "site_numero": numero,
                        "point": point,
                        "date_debut": date_debut,
                        "etat": etat,
                    })
                    continue

                # Match trouvé → calcul des changements
                changes: dict[str, Any] = {}
                # upgrade flags (jamais downgrade)
                if p_id and not local.get("vigiechiro_participation_id"):
                    changes["vigiechiro_participation_id"] = p_id
                if site_id_api and not local.get("vigiechiro_site_id"):
                    changes["vigiechiro_site_id"] = site_id_api
                if not local.get("uploaded"):
                    changes["uploaded"] = 1
                if etat in ("TERMINE", "FINI") and not local.get("analyzed"):
                    changes["analyzed"] = 1
                changes["api_etat"] = etat
                changes["last_api_sync_at"] = datetime.now().isoformat(timespec="seconds")
                # Source
                if local.get("source") == "local":
                    changes["source"] = "synced"

                if len(changes) > 2:   # plus que juste api_etat + timestamp
                    result["updates"].append({
                        "id": local["id"],
                        "label": f"{local.get('nom_contrat', '?')} / "
                                 f"{local.get('n_point_fixe', '?')} / "
                                 f"Pass{local.get('n_passage', '?')}",
                        "changes": {k: v for k, v in changes.items()
                                    if k not in ("api_etat", "last_api_sync_at")},
                        "api_etat": etat,
                        "_full_changes": changes,
                    })
                else:
                    result["unchanged"].append(local["id"])
            except Exception as e:
                result["errors"].append(f"participation {p_id}: {e}")

        # Application si demandée — sous un seul lock, transaction unique
        # (évite l'interleaving avec un accès GUI concurrent, et garantit
        # l'atomicité de l'ensemble updates + ghosts + log).
        if apply_changes:
            now = datetime.now().isoformat(timespec="seconds")
            with self._lock:
                conn = self._get_conn()
                try:
                    for upd in result["updates"]:
                        # _commit=False : on commit tout en bloc à la fin
                        self._update_fields_locked(upd["id"],
                                                     upd["_full_changes"],
                                                     commit=False)
                    for g in result["ghosts"]:
                        self._add_ghost_session_locked(g, commit=False)
                    conn.execute(
                        "INSERT INTO merge_log (timestamp, source, added, updated, notes) "
                        "VALUES (?, ?, ?, ?, ?)",
                        (now, "api:vigiechiro",
                         len(result["ghosts"]), len(result["updates"]),
                         f"pull API ({result['total_fetched']} participations)"),
                    )
                    conn.commit()
                except Exception:
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    raise

        return result

    def _add_ghost_session(self, ghost: dict) -> None:
        """Ajoute une participation serveur sans équivalent local au registre.

        L'ID du ghost est préfixé 'ghost_' pour ne pas risquer de collision
        avec les ids canoniques de sessions locales.
        """
        with self._lock:
            self._add_ghost_session_locked(ghost, commit=True)

    def _add_ghost_session_locked(self, ghost: dict, *,
                                    commit: bool = True) -> None:
        """Variante interne (lock déjà détenu). commit=False pour batcher."""
        p_id = ghost["participation_id"]
        sid = f"ghost_{p_id}"
        now = datetime.now().isoformat(timespec="seconds")
        # Upsert + override source='server' dans la même transaction
        # pour éviter la fenêtre de crash où un ghost resterait source='local'.
        self._upsert_session_locked({
            "id": sid,
            "canonical_name": f"[ghost] {ghost.get('site_numero', '?')}/"
                              f"{ghost.get('point', '?')}/{ghost.get('date_debut', '?')}",
            "campaign": "(sur serveur Vigie-Chiro)",
            "n_site_tadarida": ghost.get("site_numero") or "",
            "n_point_fixe": ghost.get("point") or "",
            "date_debut": ghost.get("date_debut") or "",
            "uploaded": 1,
            "analyzed": 1 if ghost.get("etat") in ("TERMINE", "FINI") else 0,
            "vigiechiro_participation_id": p_id,
            "vigiechiro_site_id": ghost.get("site_id"),
            "api_etat": ghost.get("etat"),
            "last_api_sync_at": now,
        }, _commit=False)
        conn = self._get_conn()
        conn.execute("UPDATE sessions SET source='server' WHERE id=?", (sid,))
        if commit:
            conn.commit()

    def last_sync_info(self) -> dict | None:
        """Dernière sync API pour affichage dans le footer."""
        with self._lock:
            row = self._get_conn().execute(
                "SELECT MAX(last_api_sync_at) AS ts FROM sessions "
                "WHERE last_api_sync_at IS NOT NULL"
            ).fetchone()
        if not row or not row[0]:
            return None
        return {"last_sync": row[0]}

    # -- Alimentation depuis scan ------------------------------------------

    def feed_from_scan(self, states: list) -> dict:
        """Alimenté depuis le résultat de chiro_core.walk_sessions + analyze.

        ``states`` : liste de SessionState ou dicts avec les champs attendus.

        Mode batch : une seule transaction pour tous les upserts (évite
        ~N fsync sur WAL). Sur 200 sessions / HDD, passe de ~2 s à ~100 ms.
        Rollback automatique si exception en cours de boucle.
        """
        result = {"inserted": 0, "updated": 0}
        with self._lock:
            conn = self._get_conn()
            try:
                for s in states:
                    d = s if isinstance(s, dict) else s.__dict__
                    # Lire le manifest si disponible pour enrichir
                    manifest_meta = {}
                    try:
                        from manifest import Manifest
                        session_path = Path(d.get("path") or "")
                        m = Manifest.load(session_path)
                        if m and m.meta:
                            manifest_meta = dict(m.meta)
                    except Exception:
                        pass

                    sid = d.get("name") or d.get("canonical_name") or str(d.get("path", ""))
                    data = {
                        "id": sid,
                        "path": str(d.get("path", "")),
                        "campaign": d.get("campaign", ""),
                        "canonical_name": sid,
                        "n_wav": d.get("n_wav", 0),
                        "total_bytes": d.get("total_bytes", 0),
                        "flag_renamed": d.get("flag_renamed", False),
                        "flag_te10_done": d.get("flag_te10_done", False),
                        "flag_uploaded_hint": d.get("flag_uploaded_hint", False),
                        "flag_analyzed": d.get("flag_analyzed", False),
                        "flag_cleaned": d.get("flag_cleaned", False),
                        # Enrichissement manifest
                        "nom_contrat": manifest_meta.get("nom_contrat", ""),
                        "n_site_tadarida": manifest_meta.get("n_site_tadarida", ""),
                        "n_point_fixe": manifest_meta.get("n_point_fixe", ""),
                        "n_passage": manifest_meta.get("n_passage"),
                        "n_enregistreur": manifest_meta.get("n_enregistreur"),
                        "n_serie": manifest_meta.get("n_serie", ""),
                        "date_debut": manifest_meta.get("date_debut", ""),
                    }
                    # Mode batch : pas de commit intermédiaire
                    action = self._upsert_session_locked(data, _commit=False)
                    result[action] += 1
                # Commit unique pour toute la batch
                conn.commit()
            except Exception:
                # Annule la batch entière si une erreur survient
                try:
                    conn.rollback()
                except Exception:
                    pass
                raise
        return result
